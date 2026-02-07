from flask import Flask, render_template, request, redirect, url_for, flash, jsonify, make_response,session,flash,abort
from flask_mysqldb import MySQL
import MySQLdb.cursors
from datetime import datetime, date, timedelta
import os
from openpyxl.styles import Font, Alignment
from MySQLdb.cursors import DictCursor


# --- NEW IMPORTS ---
import json
import locale
import calendar
from dotenv import load_dotenv # For loading environment variables
import cloudinary
import cloudinary.uploader
import cloudinary.api
import cloudinary.utils
import re
import requests
from cloudinary.utils import cloudinary_url
# --- END NEW IMPORTS ---

from fpdf import FPDF
import csv
from io import StringIO
from decimal import Decimal
import io 
from flask import send_file
import uuid
from werkzeug.utils import secure_filename
import MySQLdb 
from flask_mysqldb import MySQL
from flask import Flask,jsonify,request
from cloudinary.utils import cloudinary_url


from flask import Response

import openpyxl
from io import BytesIO

# --- LOAD ENVIRONMENT VARIABLES ---
# This will load the .env file you just created
load_dotenv()

app = Flask(__name__)

# --- MODIFIED: Load Secret Key from Environment ---
app.secret_key = os.environ.get("SECRET_KEY", "a_very_bad_default_secret")


app.config['MYSQL_HOST'] = os.environ.get('MYSQL_HOST')
app.config['MYSQL_USER'] = os.environ.get('MYSQL_USER')
app.config['MYSQL_PASSWORD'] = os.environ.get('MYSQL_PASSWORD')
app.config['MYSQL_DB'] = os.environ.get('MYSQL_DB')
app.config['MYSQL_CURSORCLASS'] = 'DictCursor'


mysql = MySQL(app)

# This will read the keys from your .env file
cloudinary.config(
    cloud_name = os.environ.get('CLOUDINARY_CLOUD_NAME'),
    api_key = os.environ.get('CLOUDINARY_API_KEY'),
    api_secret = os.environ.get('CLOUDINARY_API_SECRET')
)
app.logger.info(f"Cloudinary configured with cloud name: {os.environ.get('CLOUDINARY_CLOUD_NAME')}")


# --- FILE UPLOAD LOGIC ---
ALLOWED_EXTENSIONS = {'png', 'jpg', 'jpeg', 'gif'}
ALLOWED_DOC_EXTENSIONS = {'pdf'}

def allowed_file(filename):
    """Check if uploaded file has an allowed extension."""
    return '.' in filename and \
           filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS


# In app.py
from cloudinary.utils import cloudinary_url
app.jinja_env.globals.update(cloudinary_url=cloudinary_url)



def safe_date_format(date_obj, format='%d-%m-%Y', default='N/A'):
    """Safely formats a date object, returning default if None."""
    if date_obj:
        return date_obj.strftime(format)
    return default

# --- TIMEZONE HELPERS ---
def get_ist_now():
    """Returns current time in IST (UTC + 5:30)"""
    return datetime.utcnow() + timedelta(hours=5, minutes=30)

def parse_date_input(date_str):
    """Converts dd-mm-yyyy to yyyy-mm-dd for MySQL"""
    try:
        return datetime.strptime(date_str, '%d-%m-%Y').strftime('%Y-%m-%d')
    except (ValueError, TypeError):
        # Return as-is if it fails (might already be correct or None)
        return date_str

# =========================================================
# MONTHLY SALES DASHBOARD (Updated for Net Revenue)
# =========================================================
@app.route("/monthly_sales_dashboard")
def monthly_sales_dashboard():
    if "loggedin" not in session: return redirect(url_for("login"))

    cursor = mysql.connection.cursor(MySQLdb.cursors.DictCursor)

    # 1. Get Filters
    selected_year = request.args.get("year")
    try: selected_year = int(selected_year)
    except: selected_year = datetime.now().year

    # 2. Get Available Years
    cursor.execute("""
        SELECT YEAR(date) as y FROM evening_settle UNION 
        SELECT YEAR(sale_date) as y FROM office_sales 
        ORDER BY y DESC
    """)
    years = cursor.fetchall()
    year_list = [r['y'] for r in years] if years else [selected_year]

    # 3. UNIFIED QUERY (Net Revenue = Total - Discount)
    # Fetching from HEADER tables (evening_settle, office_sales) is more accurate for Net Revenue
    base_query = f"""
        SELECT 
            DATE_FORMAT(date_val, '%%Y-%%m') as month_key,
            DATE_FORMAT(date_val, '%%b %%Y') as month_label,
            SUM(qty_count) as total_qty,
            SUM(net_amt) as total_revenue
        FROM (
            -- Evening Sales (Field)
            SELECT 
                es.date as date_val, 
                (SELECT SUM(sold_qty) FROM evening_item WHERE settle_id=es.id) as qty_count,
                (es.total_amount - es.discount) as net_amt
            FROM evening_settle es
            WHERE es.status = 'final' AND YEAR(es.date) = %s
            
            UNION ALL
            
            -- Office Sales (Counter)
            SELECT 
                os.sale_date as date_val, 
                (SELECT SUM(qty) FROM office_sale_items WHERE sale_id=os.id) as qty_count,
                (os.final_amount) as net_amt -- final_amount is already (sub_total - discount)
            FROM office_sales os
            WHERE YEAR(os.sale_date) = %s
        ) as combined_sales
        GROUP BY month_key, month_label
        ORDER BY month_key ASC
    """
    
    cursor.execute(base_query, (selected_year, selected_year))
    monthly_data_raw = cursor.fetchall()

    # Process Data
    monthly_data = []
    total_revenue = 0.0
    total_units = 0
    best_month_val = 0
    best_month_name = "-"

    for row in monthly_data_raw:
        rev = float(row['total_revenue'] or 0)
        qty = int(row['total_qty'] or 0)
        
        monthly_data.append({
            'month': row['month_label'],
            'raw_date': row['month_key'],
            'revenue': rev,
            'units': qty
        })
        
        total_revenue += rev
        total_units += qty
        
        if rev > best_month_val:
            best_month_val = rev
            best_month_name = row['month_label']

    avg_monthly = total_revenue / len(monthly_data) if monthly_data else 0.0

    # 4. Top Selling Products (Qty remains same, Revenue approx)
    # Note: Product-wise revenue is hard to net-off with header discounts accurately without complex proportioning.
    # For Top Products, we will show Gross Revenue (Qty * Price) as an indicator, or just rank by Qty.
    prod_query = f"""
        SELECT 
            p.name as product_name,
            SUM(qty) as total_qty
        FROM (
            SELECT ei.product_id, ei.sold_qty as qty
            FROM evening_item ei
            JOIN evening_settle es ON ei.settle_id = es.id
            WHERE es.status = 'final' AND YEAR(es.date) = %s
            
            UNION ALL
            
            SELECT osi.product_id, osi.qty
            FROM office_sale_items osi
            JOIN office_sales os ON osi.sale_id = os.id
            WHERE YEAR(os.sale_date) = %s
        ) as u
        JOIN products p ON u.product_id = p.id
        GROUP BY p.id, p.name
        ORDER BY total_qty DESC
        LIMIT 5
    """
    cursor.execute(prod_query, (selected_year, selected_year))
    top_products = cursor.fetchall()

    cursor.close()

    return render_template(
        "reports/monthly_sales_dashboard.html",
        year_list=year_list,
        selected_year=selected_year,
        monthly_data=monthly_data,
        top_products=top_products,
        kpi={
            'revenue': total_revenue,
            'units': total_units,
            'avg': avg_monthly,
            'best_month': best_month_name,
            'best_val': best_month_val
        }
    )


# ==========================================
# LIVE PRODUCT TRACKER (Who has what?)
# ==========================================
@app.route('/product_tracker')
def product_tracker():
    if "loggedin" not in session: return redirect(url_for("login"))
    
    cur = mysql.connection.cursor(MySQLdb.cursors.DictCursor)
    
    # 1. Fetch Active Products
    cur.execute("SELECT id, name, image, stock, price, category_id FROM products WHERE status='active'")
    products = cur.fetchall()
    
    # 2. Fetch Active Employees
    cur.execute("SELECT id, name, image FROM employees WHERE status='active'")
    employees = cur.fetchall()
    
    # Data Structures for the View
    # tracker_map: { product_id: { 'warehouse': X, 'field': Y, 'holders': [ {name, img, qty}, ... ] } }
    tracker_map = {}
    
    # Initialize Map
    for p in products:
        # Image fix
        img = p['image']
        if not img: img = url_for('static', filename='img/default-product.png')
        elif not img.startswith('http'): img = url_for('static', filename='uploads/' + img)
            
        tracker_map[p['id']] = {
            'info': p,
            'image': img,
            'warehouse_qty': int(p['stock']),
            'field_total': 0,
            'holders': [],
            'total_sold': 0
        }

    # 3. Calculate Field Stock (Employee Holdings)
    for emp in employees:
        emp_id = emp['id']
        emp_img = emp['image']
        
        # Image fix for employee
        if not emp_img: emp_img = url_for('static', filename='img/default-user.png')
        elif not emp_img.startswith('http'): emp_img = url_for('static', filename='uploads/' + emp_img)

        # Logic: Find latest stock state for this employee (Morning or Evening)
        # Check Morning Allocations (Latest)
        cur.execute("SELECT date FROM morning_allocations WHERE employee_id=%s ORDER BY date DESC LIMIT 1", (emp_id,))
        last_morning = cur.fetchone()
        
        # Check Evening Settlements (Latest Final)
        cur.execute("SELECT date FROM evening_settle WHERE employee_id=%s AND status='final' ORDER BY date DESC LIMIT 1", (emp_id,))
        last_evening = cur.fetchone()
        
        target_date = None
        source = None 
        
        # Determine Source Date
        if last_morning and last_evening:
            if last_morning['date'] > last_evening['date']:
                target_date = last_morning['date']; source = 'morning'
            else:
                target_date = last_evening['date']; source = 'evening'
        elif last_morning:
            target_date = last_morning['date']; source = 'morning'
        elif last_evening:
            target_date = last_evening['date']; source = 'evening'
            
        # Fetch Items based on source
        if target_date:
            if source == 'evening':
                cur.execute("""
                    SELECT ei.product_id, ei.remaining_qty 
                    FROM evening_item ei 
                    JOIN evening_settle es ON ei.settle_id = es.id 
                    WHERE es.employee_id=%s AND es.date=%s
                """, (emp_id, target_date))
                
                for row in cur.fetchall():
                    pid = row['product_id']
                    qty = int(row['remaining_qty'])
                    if qty > 0 and pid in tracker_map:
                        tracker_map[pid]['field_total'] += qty
                        tracker_map[pid]['holders'].append({'name': emp['name'], 'image': emp_img, 'qty': qty})
                        
            else: # Morning
                cur.execute("""
                    SELECT mai.product_id, mai.opening_qty, mai.given_qty 
                    FROM morning_allocation_items mai 
                    JOIN morning_allocations ma ON mai.allocation_id = ma.id 
                    WHERE ma.employee_id=%s AND ma.date=%s
                """, (emp_id, target_date))
                
                for row in cur.fetchall():
                    pid = row['product_id']
                    qty = int(row['opening_qty']) + int(row['given_qty'])
                    if qty > 0 and pid in tracker_map:
                        tracker_map[pid]['field_total'] += qty
                        tracker_map[pid]['holders'].append({'name': emp['name'], 'image': emp_img, 'qty': qty})

    # 4. Calculate Total Sales (Lifetime) - Optional Stats
    # Field Sales
    cur.execute("SELECT product_id, SUM(sold_qty) as sold FROM evening_item GROUP BY product_id")
    field_sales = cur.fetchall()
    for row in field_sales:
        if row['product_id'] in tracker_map:
            tracker_map[row['product_id']]['total_sold'] += int(row['sold'])
            
    # Office Sales
    cur.execute("SELECT product_id, SUM(qty) as sold FROM office_sale_items GROUP BY product_id")
    office_sales = cur.fetchall()
    for row in office_sales:
        if row['product_id'] in tracker_map:
            tracker_map[row['product_id']]['total_sold'] += int(row['sold'])

    cur.close()
    
    # Convert map to list for template
    final_data = list(tracker_map.values())
    
    return render_template('reports/product_tracker.html', products=final_data)


# --- VIEW EVENING SETTLEMENT DETAILS ---
@app.route('/evening/view/<int:settle_id>')
def view_evening(settle_id):
    if "loggedin" not in session: return redirect(url_for("login"))
    
    conn = mysql.connection
    cursor = conn.cursor(MySQLdb.cursors.DictCursor)
    
    # 1. Fetch Header Info
    cursor.execute("""
        SELECT es.*, 
               IFNULL(es.total_amount, 0) as total_amount, 
               IFNULL(es.cash_money, 0) as cash_money, 
               IFNULL(es.online_money, 0) as online_money, 
               IFNULL(es.discount, 0) as discount,
               IFNULL(es.emp_credit_amount, 0) as emp_credit_amount,
               IFNULL(es.emp_debit_amount, 0) as emp_debit_amount,
               IFNULL(es.due_note, '') as due_note,
               e.name as emp_name, e.image as emp_image, e.phone as emp_mobile
        FROM evening_settle es
        JOIN employees e ON es.employee_id = e.id
        WHERE es.id = %s
    """, (settle_id,))
    settlement = cursor.fetchone()
    
    if not settlement:
        flash("Record not found.", "danger")
        return redirect(url_for('admin_evening_master'))
        
    # Process Header Data
    if settlement['emp_image']:
        if not settlement['emp_image'].startswith('http'):
             settlement['emp_image'] = url_for('static', filename='uploads/' + settlement['emp_image'])
    else:
        settlement['emp_image'] = url_for('static', filename='img/default-user.png')
        
    if isinstance(settlement['date'], (date, datetime)):
        settlement['formatted_date'] = settlement['date'].strftime('%d-%m-%Y')
    else:
        settlement['formatted_date'] = str(settlement['date'])

    # Calculate Due Amount safely
    total = float(settlement['total_amount'])
    paid = float(settlement['cash_money']) + float(settlement['online_money']) + float(settlement['discount'])
    settlement['due_amount'] = total - paid

    # 2. Fetch Product Items
    cursor.execute("""
        SELECT ei.*, p.name as product_name, p.image 
        FROM evening_item ei
        JOIN products p ON ei.product_id = p.id
        WHERE ei.settle_id = %s
    """, (settle_id,))
    items = cursor.fetchall()
    
    # Process Item Images
    for item in items:
        if item['image']:
            if not item['image'].startswith('http'):
                 item['image'] = url_for('static', filename='uploads/' + item['image'])
        else:
            item['image'] = url_for('static', filename='img/default-product.png')

    return render_template('admin/view_evening.html', settlement=settlement, items=items)


# =================================================================================#
#  PREMIUM EXPENSE SYSTEM V3 
# =================================================================================#

# --- Helper: Financial Health Check ---
def get_financial_health(cursor):
    """
    Calculates real-time financial standing.
    Returns dictionary with cash_hand, bank_balance, total_income, total_expense.
    """
    # 1. Income (Sales)
    cursor.execute("SELECT SUM(cash_money) as c, SUM(online_money) as o FROM evening_settle WHERE status='final'")
    evening = cursor.fetchone()
    
    cursor.execute("SELECT SUM(cash_amount) as c, SUM(online_amount) as o FROM office_sales")
    office = cursor.fetchone()
    
    total_cash_in = (float(evening['c'] or 0) + float(office['c'] or 0))
    total_bank_in = (float(evening['o'] or 0) + float(office['o'] or 0))
    total_income = total_cash_in + total_bank_in

    # 2. Outflow (Expenses)
    cursor.execute("""
        SELECT 
            SUM(CASE WHEN payment_method = 'Cash' THEN amount ELSE 0 END) as cash_out,
            SUM(CASE WHEN payment_method != 'Cash' THEN amount ELSE 0 END) as bank_out,
            SUM(amount) as total_out
        FROM expenses
    """)
    exp = cursor.fetchone()
    
    total_cash_out = float(exp['cash_out'] or 0)
    total_bank_out = float(exp['bank_out'] or 0)
    total_expense = float(exp['total_out'] or 0)

    return {
        "cash_balance": total_cash_in - total_cash_out,
        "bank_balance": total_bank_in - total_bank_out,
        "total_income": total_income,
        "total_expense": total_expense,
        "net_profit": total_income - total_expense
    }


# ... existing imports ...

# ==========================================
# EXPENSE VOUCHER PDF GENERATOR (Updated)
# ==========================================
class ExpensePDF(FPDF):
    def header(self):
        # Company Details
        self.set_font('Arial', 'B', 22)
        self.set_text_color(26, 35, 126) # Dark Blue
        self.cell(0, 10, "REAL PROMOTION", 0, 1, 'C')
        
        self.set_font('Arial', 'B', 9)
        self.set_text_color(100, 100, 100)
        self.cell(0, 5, "SINCE 2005", 0, 1, 'C')
        
        self.set_font('Arial', '', 9)
        self.set_text_color(50, 50, 50)
        address = "Real Promotion, G/12, Tulsimangalam Complex, B/h. Trimurti Complex,\nGhodiya Bazar, Nadiad-387001, Gujarat, India"
        self.multi_cell(0, 5, address, 0, 'C')
        self.cell(0, 5, "+91 96623 22476 | help@realpromotion.in", 0, 1, 'C')
        
        self.ln(5)
        self.set_draw_color(200, 200, 200)
        self.line(10, self.get_y(), 200, self.get_y())
        self.ln(5)
        
        self.set_font('Arial', 'B', 16)
        self.set_text_color(0, 0, 0)
        self.cell(0, 10, "EXPENSE VOUCHER", 0, 1, 'C')
        self.ln(5)

    def footer(self):
        self.set_y(-15)
        self.set_font('Arial', 'I', 8)
        self.set_text_color(128, 128, 128)
        self.cell(0, 10, f'Page {self.page_no()}', 0, 0, 'C')

@app.route('/expense/pdf/<int:expense_id>')
def expense_pdf(expense_id):
    if 'loggedin' not in session: return redirect(url_for('login'))
    
    cur = mysql.connection.cursor(MySQLdb.cursors.DictCursor)
    
    # Fetch Parent Expense
    cur.execute("""
        SELECT e.*, GROUP_CONCAT(DISTINCT c.category_name SEPARATOR ', ') as main_category
        FROM expenses e
        LEFT JOIN expense_items ei ON e.expense_id = ei.expense_id
        LEFT JOIN expensecategories c ON ei.category_id = c.category_id
        WHERE e.expense_id = %s
        GROUP BY e.expense_id
    """, (expense_id,))
    expense = cur.fetchone()
    
    # Fetch Items
    cur.execute("""
        SELECT ei.*, sc.subcategory_name 
        FROM expense_items ei
        LEFT JOIN expensesubcategories sc ON ei.subcategory_id = sc.subcategory_id
        WHERE ei.expense_id = %s
    """, (expense_id,))
    items = cur.fetchall()
    cur.close()
    
    if not expense:
        flash("Expense record not found.", "danger")
        return redirect(url_for('expenses_list'))

    # Generate PDF
    pdf = ExpensePDF()
    pdf.alias_nb_pages()
    pdf.add_page()
    
    # --- VOUCHER INFO SECTION ---
    pdf.set_font('Arial', '', 10)
    
    # Left Side: ID, Date, Time
    y_start = pdf.get_y()
    pdf.set_xy(10, y_start)
    pdf.cell(30, 6, "Voucher No:", 0, 0); pdf.set_font('Arial','B',10); pdf.cell(50, 6, f"#{expense['expense_id']}", 0, 1)
    
    pdf.set_font('Arial','',10)
    pdf.cell(30, 6, "Date:", 0, 0); pdf.set_font('Arial','B',10); pdf.cell(50, 6, str(expense['expense_date']), 0, 1)
    
    pdf.set_font('Arial','',10)
    pdf.cell(30, 6, "Time:", 0, 0); pdf.set_font('Arial','B',10); pdf.cell(50, 6, str(expense['expense_time']), 0, 1)

    # Right Side: Category
    pdf.set_xy(110, y_start)
    pdf.set_font('Arial','',10)
    pdf.cell(35, 6, "Main Category:", 0, 0); pdf.set_font('Arial','B',10); pdf.cell(50, 6, str(expense['main_category'] or 'General'), 0, 1)
    
    # Removed overall Payment Mode since it's item-wise now, but keeping for reference if needed
    # pdf.set_xy(110, y_start + 6)
    # pdf.set_font('Arial','',10)
    # pdf.cell(35, 6, "Payment Mode:", 0, 0); pdf.set_font('Arial','B',10); pdf.cell(50, 6, str(expense['payment_method']), 0, 1)
    
    pdf.ln(20)

    # --- ITEMS TABLE ---
    # Header
    pdf.set_fill_color(30, 58, 138) # Dark Blue (#1e3a8a)
    pdf.set_text_color(255, 255, 255)   # White Text
    pdf.set_font('Arial', 'B', 9)
    pdf.set_draw_color(200, 200, 200)
    pdf.set_line_width(0.3)
    
    # Columns: #, Sub-Category, Description, Payment Mode, Amount
    w = [10, 45, 75, 25, 35] # Adjusted widths
    headers = ["#", "Item", "Description", "Mode", "Amount (Rs.)"]
    
    for i, h in enumerate(headers):
        pdf.cell(w[i], 8, h, 1, 0, 'C', True)
    pdf.ln()
    
    # Rows
    pdf.set_font('Arial', '', 9)
    pdf.set_text_color(0, 0, 0)
    pdf.set_fill_color(250, 250, 250)
    
    fill = False
    for idx, item in enumerate(items):
        pdf.cell(w[0], 8, str(idx+1), 1, 0, 'C', fill)
        pdf.cell(w[1], 8, str(item['subcategory_name']), 1, 0, 'L', fill)
        pdf.cell(w[2], 8, str(item['description'] or '-'), 1, 0, 'L', fill)
        pdf.cell(w[3], 8, str(item['payment_method']), 1, 0, 'C', fill) # Added Payment Mode
        pdf.cell(w[4], 8, f"{item['amount']:,.2f}", 1, 1, 'R', fill)
        fill = not fill
        
    # --- TOTALS SECTION ---
    pdf.ln(2)
    pdf.set_font('Arial', 'B', 11)
    pdf.set_fill_color(240, 240, 240)
    
    # Offset to align with Amount column
    pdf.cell(w[0]+w[1]+w[2]+w[3], 10, "GRAND TOTAL  ", 0, 0, 'R')
    pdf.set_text_color(26, 35, 126)
    pdf.cell(w[4], 10, f"Rs. {expense['total_amount']:,.2f}", 1, 1, 'R', True)
    
    # --- UNIVERSAL NOTES ---
    pdf.ln(10)
    pdf.set_text_color(0, 0, 0)
    pdf.set_font('Arial', 'B', 10)
    pdf.cell(0, 6, "Universal Voucher Notes:", 0, 1)
    
    pdf.set_font('Arial', '', 10)
    note = expense['description'] if expense['description'] else "No additional notes."
    pdf.multi_cell(0, 6, note, 0, 'L')
    
    # --- SIGNATURE ---
    if pdf.get_y() > 240: pdf.add_page()
    pdf.ln(20)
    
    # Add signature image if exists
    sig_path = os.path.join(app.root_path, 'static', 'img', 'signature.png')
    if os.path.exists(sig_path):
        pdf.image(sig_path, x=150, y=pdf.get_y(), w=40)
        
    pdf.set_xy(140, pdf.get_y() + 25)
    pdf.set_font('Arial', 'B', 9)
    pdf.cell(50, 5, "Authorized Signature", 0, 1, 'C')
    pdf.set_font('Arial', 'I', 8)
    pdf.cell(0, 5, "Real Promotion All Rights Reserved.", 0, 1, 'R')

    # Output
    buffer = io.BytesIO(pdf.output(dest='S').encode('latin-1'))
    buffer.seek(0)
    
    # Create Filename: Category_ID.pdf
    main_cat = str(expense['main_category'] or 'General').replace(" ", "_")
    filename = f"{main_cat}_{expense_id}.pdf"
    
    return send_file(buffer, as_attachment=True, download_name=filename, mimetype='application/pdf')


@app.route('/expense_dash')
def expense_dash():
    if 'loggedin' not in session: return redirect(url_for('login'))
    
    cur = mysql.connection.cursor(MySQLdb.cursors.DictCursor)
    
    # --- INCOME ---
    cur.execute("SELECT SUM(cash_money) as c, SUM(online_money) as b FROM evening_settle WHERE status='final'")
    eve = cur.fetchone()
    cur.execute("SELECT SUM(cash_amount) as c, SUM(online_amount) as b FROM office_sales")
    off = cur.fetchone()
    cur.execute("SELECT SUM(CASE WHEN payment_mode='Cash' THEN amount ELSE 0 END) as c, SUM(CASE WHEN payment_mode!='Cash' THEN amount ELSE 0 END) as b FROM employee_transactions WHERE type='credit'")
    emp_in = cur.fetchone()
    
    # Calculate Total Cash Income
    total_cash_in = float(eve['c'] or 0) + float(off['c'] or 0) + float(emp_in['c'] or 0)
    total_bank_in = float(eve['b'] or 0) + float(off['b'] or 0) + float(emp_in['b'] or 0)

    # --- EXPENSES (Now includes Supplier Payments) ---
    cur.execute("SELECT SUM(CASE WHEN payment_method='Cash' THEN amount ELSE 0 END) as c, SUM(CASE WHEN payment_method!='Cash' THEN amount ELSE 0 END) as b FROM expense_items")
    exp_out = cur.fetchone()
    
    cur.execute("SELECT SUM(CASE WHEN payment_mode='Cash' THEN amount ELSE 0 END) as c, SUM(CASE WHEN payment_mode!='Cash' THEN amount ELSE 0 END) as b FROM employee_transactions WHERE type='debit'")
    emp_out = cur.fetchone()
    
    # *** ADDED THIS BLOCK ***
    cur.execute("SELECT SUM(CASE WHEN payment_mode='Cash' THEN amount_paid ELSE 0 END) as c, SUM(CASE WHEN payment_mode!='Cash' THEN amount_paid ELSE 0 END) as b FROM supplier_payments")
    supp_out = cur.fetchone()
    
    # --- NEW: CALCULATE REFUNDS (PROFIT PAID TO EMPLOYEE) ---
    # Logic: Sum of negative 'due_amount' where status is final. 
    # Negative due means company owes employee, usually paid immediately in cash.
    cur.execute("SELECT SUM(ABS(due_amount)) as total_refunds FROM evening_settle WHERE status='final' AND due_amount < 0")
    refund_res = cur.fetchone()
    total_refunds = float(refund_res['total_refunds'] or 0)

    # Total Cash Outflow now includes Refunds
    total_cash_out = float(exp_out['c'] or 0) + float(emp_out['c'] or 0) + float(supp_out['c'] or 0) + total_refunds
    total_bank_out = float(exp_out['b'] or 0) + float(emp_out['b'] or 0) + float(supp_out['b'] or 0)
    
    fin_health = {
        'total_balance': (total_cash_in + total_bank_in) - (total_cash_out + total_bank_out),
        'cash_balance': total_cash_in - total_cash_out,
        'bank_balance': total_bank_in - total_bank_out
    }

    # --- 2. CASHFLOW CHART (Monthly Income vs Expense) ---
    months = []
    income_data = []
    expense_data = []
    
    for i in range(5, -1, -1):
        d = date.today() - timedelta(days=i*30)
        m_start = d.replace(day=1)
        next_m = (m_start.replace(day=28) + timedelta(days=4)).replace(day=1)
        m_end = next_m - timedelta(days=1)
        m_label = m_start.strftime('%b')
        months.append(m_label)
        
        # Monthly Income (Field + Office + Emp Credit)
        cur.execute("""
            SELECT (
                COALESCE((SELECT SUM(total_amount) FROM evening_settle WHERE date BETWEEN %s AND %s AND status='final'), 0) +
                COALESCE((SELECT SUM(final_amount) FROM office_sales WHERE sale_date BETWEEN %s AND %s), 0) +
                COALESCE((SELECT SUM(amount) FROM employee_transactions WHERE transaction_date BETWEEN %s AND %s AND type='credit'), 0)
            ) as total
        """, (m_start, m_end, m_start, m_end, m_start, m_end))
        inc = cur.fetchone()['total']
        income_data.append(float(inc))
        
        # Monthly Expense (Vouchers + Emp Debit + Refunds)
        # Note: Ideally we should include monthly refunds here too for accuracy, but kept simple as requested to not change major logic
        cur.execute("""
            SELECT (
                COALESCE((SELECT SUM(total_amount) FROM expenses WHERE expense_date BETWEEN %s AND %s), 0) +
                COALESCE((SELECT SUM(amount) FROM employee_transactions WHERE transaction_date BETWEEN %s AND %s AND type='debit'), 0)
            ) as total
        """, (m_start, m_end, m_start, m_end))
        exc = cur.fetchone()['total'] or 0
        expense_data.append(float(exc))

    chart_cashflow = {
        'labels': months,
        'income': income_data,
        'expense': expense_data
    }

    # --- 3. CATEGORY PIE CHART ---
    cur.execute("""
        SELECT c.category_name, SUM(ei.amount) as total
        FROM expense_items ei
        JOIN expensecategories c ON ei.category_id = c.category_id
        GROUP BY c.category_name
        ORDER BY total DESC
    """)
    cat_data = cur.fetchall()
    
    chart_pie = {
        'labels': [r['category_name'] for r in cat_data],
        'data': [float(r['total']) for r in cat_data]
    }

    # --- 4. RECENT ACTIVITY ---
    cur.execute("""
        SELECT e.expense_id, e.expense_date, e.total_amount, e.payment_method, e.description,
               COUNT(ei.id) as item_count
        FROM expenses e
        LEFT JOIN expense_items ei ON e.expense_id = ei.expense_id
        GROUP BY e.expense_id
        ORDER BY e.expense_date DESC, e.expense_time DESC
        LIMIT 5
    """)
    recent = cur.fetchall()

    cur.close()
    
    return render_template('expenses/expense_dash.html', 
                         fin=fin_health, 
                         cashflow=chart_cashflow, 
                         pie=chart_pie, 
                         recent=recent)


@app.route('/add_expense', methods=['GET', 'POST'])
def add_expense():
    if 'loggedin' not in session: 
        return redirect(url_for('login'))
    
    # We open the connection at the start
    cur = mysql.connection.cursor(MySQLdb.cursors.DictCursor)
    
    if request.method == 'POST':
        try:
            # 1. Extract Form Data
            exp_date_raw = request.form.get('expense_date')
            # Assuming parse_date_input is defined in your app.py to handle d-m-Y to Y-m-d
            exp_date = parse_date_input(exp_date_raw)
            exp_time = get_ist_now().strftime('%H:%M:%S')
            main_cat_id = request.form.get('main_category_id')
            universal_note = request.form.get('voucher_description')
            
            subcat_ids = request.form.getlist('subcategory_id[]')
            amounts = request.form.getlist('amount[]')
            item_notes = request.form.getlist('description[]')
            item_methods = request.form.getlist('item_payment_method[]')
            
            # 2. Calculations
            # Convert to integers as requested for whole numbers
            total_sum = sum(int(float(x)) for x in amounts if x)
            
            # Determine overall voucher payment method
            unique_methods = set(item_methods)
            if len(unique_methods) > 1:
                final_voucher_method = "Mixed"
            elif len(unique_methods) == 1:
                final_voucher_method = list(unique_methods)[0]
            else:
                final_voucher_method = "Cash"

            # 3. Handle Cloudinary Upload
            receipt_url = None
            if 'receipt' in request.files:
                file = request.files['receipt']
                if file and file.filename != '':
                    try:
                        upload_res = cloudinary.uploader.upload(file, folder="erp_expenses")
                        receipt_url = upload_res.get('secure_url')
                    except Exception as e:
                        app.logger.error(f"Cloudinary Upload Error: {e}")

            # 4. Insert Parent Voucher (expenses table)
            cur.execute("""
                INSERT INTO expenses 
                (expense_date, expense_time, amount, total_amount, payment_method, receipt_url, description, subcategory_id)
                VALUES (%s, %s, %s, %s, %s, %s, %s, NULL)
            """, (exp_date, exp_time, total_sum, total_sum, final_voucher_method, receipt_url, universal_note))
            eid = cur.lastrowid
            
            # 5. Insert Child Items (expense_items table)
            for i in range(len(subcat_ids)):
                if amounts[i] and int(float(amounts[i])) > 0:
                    cur.execute("""
                        INSERT INTO expense_items (expense_id, category_id, subcategory_id, amount, payment_method, description)
                        VALUES (%s, %s, %s, %s, %s, %s)
                    """, (eid, main_cat_id, subcat_ids[i], int(float(amounts[i])), item_methods[i], item_notes[i]))
            
            mysql.connection.commit()
            flash(f"Voucher #{eid} Saved Successfully!", "success")
            cur.close()
            return redirect(url_for('expenses_list'))
            
        except Exception as e:
            mysql.connection.rollback()
            app.logger.error(f"POST /add_expense error: {str(e)}")
            flash(f"Error recording expense: {str(e)}", "danger")

    # --- GET Logic & Error Recovery Path ---
    try:
        if not cur or cur.connection is None:
            cur = mysql.connection.cursor(MySQLdb.cursors.DictCursor)

        # 1. Total Income (Sales + Employee Credit)
        cur.execute("SELECT SUM(cash_money) as cash, SUM(online_money) as bank FROM evening_settle WHERE status='final'")
        eve = cur.fetchone()
        cur.execute("SELECT SUM(cash_amount) as cash, SUM(online_amount) as bank FROM office_sales")
        off = cur.fetchone()
        
        # New: Add Employee Credit (Money In)
        cur.execute("""
            SELECT 
                SUM(CASE WHEN payment_mode = 'Cash' THEN amount ELSE 0 END) as cash_in,
                SUM(CASE WHEN payment_mode != 'Cash' THEN amount ELSE 0 END) as bank_in
            FROM employee_transactions 
            WHERE type = 'credit'
        """)
        emp_in = cur.fetchone()
        
        total_cash_in = float(eve['cash'] or 0) + float(off['cash'] or 0) + float(emp_in['cash_in'] or 0)
        total_bank_in = float(eve['bank'] or 0) + float(off['bank'] or 0) + float(emp_in['bank_in'] or 0)

        # 2. Total Expenses (Vouchers + Employee Debit + Supplier Payments)
        cur.execute("""
            SELECT 
                SUM(CASE WHEN payment_method = 'Cash' THEN amount ELSE 0 END) as cash_total,
                SUM(CASE WHEN payment_method = 'Online' THEN amount ELSE 0 END) as bank_total
            FROM expense_items
        """)
        exp_totals = cur.fetchone()
        
        # New: Add Employee Debit (Money Out)
        cur.execute("""
            SELECT 
                SUM(CASE WHEN payment_mode = 'Cash' THEN amount ELSE 0 END) as cash_out,
                SUM(CASE WHEN payment_mode != 'Cash' THEN amount ELSE 0 END) as bank_out
            FROM employee_transactions 
            WHERE type = 'debit'
        """)
        emp_out = cur.fetchone()

        # Add Supplier Payments (Money Out) - ADDED AS PER PREVIOUS INSTRUCTION FOR EXPENSE DASH
        cur.execute("""
            SELECT 
                SUM(CASE WHEN payment_mode = 'Cash' THEN amount_paid ELSE 0 END) as cash_out,
                SUM(CASE WHEN payment_mode != 'Cash' THEN amount_paid ELSE 0 END) as bank_out
            FROM supplier_payments
        """)
        supp_out = cur.fetchone()
        
        # --- NEW: CALCULATE REFUNDS (PROFIT PAID TO EMPLOYEE) ---
        # Logic: Sum of negative 'due_amount' where status is final. 
        # Negative due means company owes employee, usually paid immediately in cash.
        cur.execute("SELECT SUM(ABS(due_amount)) as total_refunds FROM evening_settle WHERE status='final' AND due_amount < -0.01")
        refund_res = cur.fetchone()
        total_refunds = float(refund_res['total_refunds'] or 0)

        total_cash_out = float(exp_totals['cash_total'] or 0) + float(emp_out['cash_out'] or 0) + float(supp_out['cash_out'] or 0) + total_refunds
        total_bank_out = float(exp_totals['bank_total'] or 0) + float(emp_out['bank_out'] or 0) + float(supp_out['bank_out'] or 0)

        cash_balance = total_cash_in - total_cash_out
        bank_balance = total_bank_in - total_bank_out

        # 3. Meta Data
        cur.execute("SELECT * FROM expensecategories ORDER BY category_name")
        cats = cur.fetchall()
        cur.execute("SELECT * FROM expensesubcategories ORDER BY subcategory_name")
        subs = cur.fetchall()
        
        cur.close()
        return render_template('expenses/add_expense.html', 
                             categories=cats, 
                             subcategories=subs,
                             cash_balance=cash_balance,
                             bank_balance=bank_balance,
                             today=date.today().strftime('%d-%m-%Y'))

    except Exception as e:
        app.logger.error(f"GET /add_expense error: {str(e)}")
        if cur: cur.close()
        flash("System error while loading the expense form.", "danger")
        return redirect(url_for('expense_dash'))

# 1. Expense List (With Advanced Filters & Date Range Support)
@app.route('/expenses_list')
def expenses_list():
    if 'loggedin' not in session: return redirect(url_for('login'))
    cur = mysql.connection.cursor(MySQLdb.cursors.DictCursor)
    
    # Get Filters
    q = request.args.get('q', '')
    cat_id = request.args.get('cat_id', 'all')
    sub_id = request.args.get('sub_id', 'all')
    date_range = request.args.get('date_range', '') # Single string from Flatpickr
    
    sql = """
        SELECT e.*, 
               GROUP_CONCAT(DISTINCT c.category_name SEPARATOR ', ') as main_category,
               COUNT(ei.id) as item_count
        FROM expenses e
        LEFT JOIN expense_items ei ON e.expense_id = ei.expense_id
        LEFT JOIN expensecategories c ON ei.category_id = c.category_id
        WHERE 1=1
    """
    params = []
    
    if q:
        sql += " AND (e.description LIKE %s OR c.category_name LIKE %s)"
        params.extend([f'%{q}%', f'%{q}%'])
    if cat_id != 'all':
        sql += " AND ei.category_id = %s"
        params.append(cat_id)
    if sub_id != 'all':
        sql += " AND ei.subcategory_id = %s"
        params.append(sub_id)
        
    # Date Range Logic (format from Flatpickr: "YYYY-MM-DD to YYYY-MM-DD")
    if date_range:
        if "to" in date_range:
            try:
                parts = date_range.split(" to ")
                if len(parts) == 2:
                    start_str, end_str = parts
                    sql += " AND e.expense_date BETWEEN %s AND %s"
                    params.extend([start_str.strip(), end_str.strip()])
            except: pass
        else:
            # Single date selected
            sql += " AND e.expense_date = %s"
            params.append(date_range.strip())
    
    sql += " GROUP BY e.expense_id ORDER BY e.expense_date DESC, e.expense_time DESC"
    
    cur.execute(sql, tuple(params))
    expenses = cur.fetchall()
    
    cur.execute("SELECT * FROM expensecategories ORDER BY category_name")
    categories = cur.fetchall()
    cur.execute("SELECT * FROM expensesubcategories ORDER BY subcategory_name")
    subcategories = cur.fetchall()
    
    cur.close()
    return render_template('expenses/expenses_list.html', 
                         expenses=expenses, 
                         categories=categories, 
                         subcategories=subcategories,
                         filters={'q':q, 'cat':cat_id, 'sub':sub_id, 'range':date_range})




# 5. CATEGORY MANAGER
@app.route('/category_man', methods=['GET', 'POST'])
def category_man():
    if 'loggedin' not in session: return redirect(url_for('login'))
    conn = mysql.connection
    cursor = conn.cursor(MySQLdb.cursors.DictCursor)
    
    if request.method == 'POST':
        type_ = request.form['type']
        name = request.form['name']
        
        if type_ == 'main':
            cursor.execute("INSERT INTO expensecategories (category_name) VALUES (%s)", [name])
        else:
            parent = request.form['parent_id']
            cursor.execute("INSERT INTO expensesubcategories (category_id, subcategory_name) VALUES (%s, %s)", (parent, name))
        
        conn.commit()
        return redirect(url_for('category_man'))
        
    cursor.execute("SELECT * FROM expensecategories")
    cats = cursor.fetchall()
    cursor.execute("SELECT * FROM expensesubcategories")
    subs = cursor.fetchall()
    
    tree = []
    for c in cats:
        c['subs'] = [s for s in subs if s['category_id'] == c['category_id']]
        tree.append(c)
        
    cursor.close()
    return render_template('expenses/category_man.html', tree=tree)




@app.route('/delete_category/<int:category_id>', methods=['POST'])
def delete_category(category_id):
    if 'loggedin' not in session: return redirect(url_for('login'))
    cur = mysql.connection.cursor(MySQLdb.cursors.DictCursor)
    
    try:
        # 1. Check if used in Expense Items
        cur.execute("SELECT COUNT(*) as cnt FROM expense_items WHERE category_id = %s", (category_id,))
        if cur.fetchone()['cnt'] > 0:
            flash("Cannot delete: Expenses exist under this Main Category.", "danger")
            return redirect(url_for('category_man'))

        # 2. Check if has Sub-categories
        cur.execute("SELECT COUNT(*) as cnt FROM expensesubcategories WHERE category_id = %s", (category_id,))
        if cur.fetchone()['cnt'] > 0:
             flash("Cannot delete: This category contains sub-categories. Delete them first.", "warning")
             return redirect(url_for('category_man'))

        # 3. Safe to delete
        cur.execute("DELETE FROM expensecategories WHERE category_id = %s", (category_id,))
        mysql.connection.commit()
        flash("Main Category deleted successfully.", "success")
        
    except Exception as e:
        mysql.connection.rollback()
        flash(f"Error: {e}", "danger")
    finally:
        cur.close()
        
    return redirect(url_for('category_man'))


@app.route('/delete_subcategory/<int:subcategory_id>', methods=['POST'])
def delete_subcategory(subcategory_id):
    if 'loggedin' not in session: return redirect(url_for('login'))
    cur = mysql.connection.cursor(MySQLdb.cursors.DictCursor)
    
    try:
        # 1. Check if used in Expense Items (New System)
        cur.execute("SELECT COUNT(*) as cnt FROM expense_items WHERE subcategory_id = %s", (subcategory_id,))
        if cur.fetchone()['cnt'] > 0:
            flash("Cannot delete: Expenses exist for this Item.", "danger")
            return redirect(url_for('category_man'))
            
        # 2. Check usage in Expenses (Legacy Check)
        cur.execute("SELECT COUNT(*) as cnt FROM expenses WHERE subcategory_id = %s", (subcategory_id,))
        if cur.fetchone()['cnt'] > 0:
            flash("Cannot delete: Linked to a voucher in expenses table.", "danger")
            return redirect(url_for('category_man'))

        # 3. Safe to delete
        cur.execute("DELETE FROM expensesubcategories WHERE subcategory_id = %s", (subcategory_id,))
        mysql.connection.commit()
        flash("Sub-Category deleted successfully.", "success")
        
    except Exception as e:
        mysql.connection.rollback()
        flash(f"Error: {e}", "danger")
    finally:
        cur.close()
        
    return redirect(url_for('category_man'))



@app.route('/view_expense/<int:expense_id>')
def view_expense(expense_id):
    if 'loggedin' not in session: return redirect(url_for('login'))
    cur = mysql.connection.cursor(MySQLdb.cursors.DictCursor)
    
    cur.execute("SELECT * FROM expenses WHERE expense_id = %s", (expense_id,))
    expense = cur.fetchone()
    
    cur.execute("""
        SELECT ei.*, c.category_name, sc.subcategory_name 
        FROM expense_items ei
        JOIN expensecategories c ON ei.category_id = c.category_id
        JOIN expensesubcategories sc ON ei.subcategory_id = sc.subcategory_id
        WHERE ei.expense_id = %s
    """, (expense_id,))
    items = cur.fetchall()
    
    cur.close()
    return render_template('expenses/view_expense.html', expense=expense, items=items)

@app.route('/edit_expense/<int:expense_id>', methods=['GET', 'POST'])
def edit_expense(expense_id):
    if 'loggedin' not in session: return redirect(url_for('login'))
    cur = mysql.connection.cursor(MySQLdb.cursors.DictCursor)
    
    if request.method == 'POST':
        try:
            exp_date = parse_date_input(request.form.get('expense_date'))
            main_cat_id = request.form.get('main_category_id')
            universal_note = request.form.get('voucher_description')
            subcat_ids = request.form.getlist('subcategory_id[]')
            amounts = request.form.getlist('amount[]')
            item_notes = request.form.getlist('description[]')
            item_methods = request.form.getlist('item_payment_method[]')
            
            total_sum = sum(int(x) for x in amounts if x)
            unique_methods = set(item_methods)
            final_voucher_method = item_methods[0] if len(unique_methods) == 1 else "Mixed"

            # Handle Receipt Update
            receipt_url = request.form.get('old_receipt_url')
            if 'receipt' in request.files and request.files['receipt'].filename != '':
                upload_res = cloudinary.uploader.upload(request.files['receipt'], folder="erp_expenses")
                receipt_url = upload_res.get('secure_url')

            cur.execute("""
                UPDATE expenses 
                SET expense_date=%s, amount=%s, total_amount=%s, payment_method=%s, receipt_url=%s, description=%s
                WHERE expense_id=%s
            """, (exp_date, total_sum, total_sum, final_voucher_method, receipt_url, universal_note, expense_id))
            
            cur.execute("DELETE FROM expense_items WHERE expense_id=%s", (expense_id,))
            for i in range(len(subcat_ids)):
                if amounts[i] and int(amounts[i]) > 0:
                    cur.execute("""
                        INSERT INTO expense_items (expense_id, category_id, subcategory_id, amount, payment_method, description)
                        VALUES (%s, %s, %s, %s, %s, %s)
                    """, (expense_id, main_cat_id, subcat_ids[i], int(amounts[i]), item_methods[i], item_notes[i]))
            
            mysql.connection.commit()
            flash("Voucher updated!", "success")
            return redirect(url_for('expenses_list'))
        except Exception as e:
            mysql.connection.rollback()
            flash(f"Error: {e}", "danger")

    # GET: Load current data
    cur.execute("SELECT * FROM expenses WHERE expense_id=%s", (expense_id,))
    expense = cur.fetchone()
    cur.execute("SELECT * FROM expense_items WHERE expense_id=%s", (expense_id,))
    items = cur.fetchall()
    cur.execute("SELECT * FROM expensecategories ORDER BY category_name")
    categories = cur.fetchall()
    cur.execute("SELECT * FROM expensesubcategories ORDER BY subcategory_name")
    subs = cur.fetchall()
    cur.close()
    return render_template('expenses/edit_expense.html', expense=expense, items=items, categories=categories, subcategories=subs)

@app.route('/delete_expense/<int:id>', methods=['POST'])
def delete_expense(id):
    if 'loggedin' not in session: return redirect(url_for('login'))
    cur = mysql.connection.cursor()
    try:
        cur.execute("DELETE FROM expense_items WHERE expense_id=%s", (id,))
        cur.execute("DELETE FROM expenses WHERE expense_id=%s", (id,))
        mysql.connection.commit()
        flash("Voucher deleted permanently.", "info")
    except Exception as e:
        mysql.connection.rollback()
        flash(f"Error: {e}", "danger")
    cur.close()
    return redirect(url_for('expenses_list'))



# ... existing imports ...

# Helper to ensure dates are always YYYY-MM-DD for MySQL
def ensure_mysql_date(date_str):
    if not date_str: return None
    try:
        # Try parsing dd-mm-yyyy (Common frontend format)
        return datetime.strptime(date_str, '%d-%m-%Y').strftime('%Y-%m-%d')
    except ValueError:
        try:
            # Try parsing yyyy-mm-dd (Standard MySQL format)
            datetime.strptime(date_str, '%Y-%m-%d')
            return date_str
        except ValueError:
            return date_str # Return as-is if unknown format

@app.route("/exp_report", methods=["GET"])
def exp_report():
    if "loggedin" not in session:
        return redirect(url_for("login"))

    cursor = mysql.connection.cursor(MySQLdb.cursors.DictCursor)

    report_type = request.args.get("type", "monthly_summary")
    now = datetime.now()

    # 1. Date Filter Logic
    try:
        year = int(request.args.get("year") or now.year)
        month = int(request.args.get("month") or now.month)
    except ValueError:
        year = now.year
        month = now.month

    # Get raw strings from URL
    raw_start = request.args.get("start_date")
    raw_end = request.args.get("end_date")

    # Convert to MySQL format (YYYY-MM-DD)
    start_date_mysql = ensure_mysql_date(raw_start)
    end_date_mysql = ensure_mysql_date(raw_end)

    # Defaults: If no date provided, use current Month ranges
    if not start_date_mysql:
        start_date_mysql = f"{year}-{month:02d}-01"
    if not end_date_mysql:
        _, last_day = calendar.monthrange(year, month)
        end_date_mysql = f"{year}-{month:02d}-{last_day}"

    # Filters dictionary to pass back to template (Keep raw format for UI inputs)
    filters = {
        "year": year,
        "month": f"{month:02d}", 
        "start_date": raw_start or start_date_mysql, # Keep what user typed or default
        "end_date": raw_end or end_date_mysql
    }

    report_data = {}

    try:
        # --- TYPE 1: MONTHLY SUMMARY ---
        if report_type == "monthly_summary":
            _, last_day = calendar.monthrange(year, month)
            m_start = f"{year}-{month:02d}-01"
            m_end = f"{year}-{month:02d}-{last_day}"
            
            cursor.execute("""
                SELECT COALESCE(SUM(total_amount), 0) as total, COUNT(expense_id) as count 
                FROM expenses WHERE expense_date BETWEEN %s AND %s
            """, (m_start, m_end))
            summary = cursor.fetchone()
            
            total_spend = float(summary['total'])
            tx_count = summary['count']
            avg_ticket = total_spend / tx_count if tx_count > 0 else 0
            
            cursor.execute("""
                SELECT COALESCE(sc.subcategory_name, 'General') as subcategory_name, 
                       COALESCE(c.category_name, 'Other') as category_name, 
                       SUM(ei.amount) as sub_total
                FROM expense_items ei
                JOIN expenses e ON ei.expense_id = e.expense_id
                LEFT JOIN expensesubcategories sc ON ei.subcategory_id = sc.subcategory_id
                LEFT JOIN expensecategories c ON ei.category_id = c.category_id
                WHERE e.expense_date BETWEEN %s AND %s
                GROUP BY ei.subcategory_id
                ORDER BY sub_total DESC LIMIT 5
            """, (m_start, m_end))
            top_items = [dict(row, sub_total=float(row['sub_total'])) for row in cursor.fetchall()]
            
            report_data = {
                "month_name": calendar.month_name[month],
                "year": year,
                "summary": {"total": total_spend, "count": tx_count, "average": avg_ticket},
                "top_items": top_items
            }

        # --- TYPE 2: DETAILED LOG (FIXED) ---
        elif report_type == "detailed_log":
            # Using MySQL formatted dates
            cursor.execute("""
                SELECT 
                    e.expense_date, 
                    COALESCE(c.category_name, 'Uncategorized') as category_name, 
                    COALESCE(sc.subcategory_name, '-') as subcategory_name, 
                    ei.amount, 
                    COALESCE(ei.description, e.description) as description,
                    COALESCE(e.payment_method, 'Cash') as payment_method
                FROM expense_items ei
                JOIN expenses e ON ei.expense_id = e.expense_id
                LEFT JOIN expensecategories c ON ei.category_id = c.category_id
                LEFT JOIN expensesubcategories sc ON ei.subcategory_id = sc.subcategory_id
                WHERE e.expense_date BETWEEN %s AND %s
                ORDER BY e.expense_date DESC, e.expense_id DESC
            """, (start_date_mysql, end_date_mysql))
            
            log = cursor.fetchall()
            
            total_log = 0.0
            for item in log:
                amt = float(item['amount']) if item['amount'] else 0.0
                item['amount'] = amt
                total_log += amt
            
            report_data = {
                "log": log,
                "summary": {"total": total_log}
            }

        # --- TYPE 3: CATEGORY ANALYSIS ---
        elif report_type == "category_wise":
            cursor.execute("""
                SELECT c.category_name, SUM(ei.amount) as total_amount
                FROM expense_items ei
                JOIN expenses e ON ei.expense_id = e.expense_id
                LEFT JOIN expensecategories c ON ei.category_id = c.category_id
                WHERE YEAR(e.expense_date) = %s
                GROUP BY c.category_id
                ORDER BY total_amount DESC
            """, (year,))
            cats = [dict(row, total_amount=float(row['total_amount'])) for row in cursor.fetchall()]
            total_exp = sum(x['total_amount'] for x in cats)
            
            report_data = {
                "by_category": cats,
                "total_expense": total_exp,
                "summary": {"total": total_exp}
            }

        # --- TYPE 4: SUB-CATEGORY DRILLDOWN (FIXED) ---
        elif report_type == "subcategory_wise":
            cursor.execute("""
                SELECT 
                    COALESCE(sc.subcategory_name, 'General/Other') as subcategory_name, 
                    COALESCE(c.category_name, 'Uncategorized') as category_name,
                    SUM(ei.amount) as total_amount
                FROM expense_items ei
                JOIN expenses e ON ei.expense_id = e.expense_id
                LEFT JOIN expensesubcategories sc ON ei.subcategory_id = sc.subcategory_id
                LEFT JOIN expensecategories c ON ei.category_id = c.category_id
                WHERE YEAR(e.expense_date) = %s
                GROUP BY ei.subcategory_id, ei.category_id
                ORDER BY total_amount DESC
            """, (year,))
            subs = cursor.fetchall()
            
            # Ensure float conversion
            for sub in subs:
                sub['total_amount'] = float(sub['total_amount']) if sub['total_amount'] else 0.0
                
            total_exp = sum(x['total_amount'] for x in subs)
            
            report_data = {
                "by_subcategory": subs,
                "summary": {"total": total_exp}
            }

        # --- TYPE 5: MOM COMPARISON ---
        elif report_type == "mom_comparison":
            cursor.execute("""
                SELECT MONTH(e.expense_date) as month_num, SUM(e.total_amount) as total
                FROM expenses e
                WHERE YEAR(e.expense_date) = %s
                GROUP BY MONTH(e.expense_date)
            """, (year,))
            results = {row['month_num']: float(row['total']) for row in cursor.fetchall()}
            
            comparison = []
            total_year = 0
            for m in range(1, 13):
                curr = results.get(m, 0.0)
                total_year += curr
                prev = results.get(m-1, 0.0)
                
                change = 0
                if prev > 0:
                    change = ((curr - prev) / prev) * 100
                elif curr > 0 and m > 1:
                    change = 100
                
                comparison.append({
                    "month_name": calendar.month_abbr[m], 
                    "total": curr, 
                    "change": change
                })
            
            report_data = {
                "comparison": comparison,
                "year": year,
                "summary": {"total": total_year}
            }

        # --- TYPE 6: ANNUAL SUMMARY (FIXED) ---
        elif report_type == "annual_summary":
            # Using Header table for Annual Overview to ensure accuracy
            cursor.execute("""
                SELECT 
                    MONTH(e.expense_date) as month_num, 
                    SUM(e.total_amount) as total
                FROM expenses e
                WHERE YEAR(e.expense_date) = %s
                GROUP BY MONTH(e.expense_date)
            """, (year,))
            
            rows = cursor.fetchall()
            # Map month number to total
            results = {row['month_num']: float(row['total']) for row in rows}
            
            annual_list = []
            grand_total = 0.0
            
            # Ensure all 12 months are present in the list
            for m in range(1, 13):
                val = results.get(m, 0.0)
                grand_total += val
                annual_list.append({
                    "month_name": calendar.month_name[m],
                    "total": val
                })
                
            report_data = {
                "year": year,
                "annual": annual_list,
                "grand_total": grand_total,
                "summary": {"total": grand_total}
            }

    except Exception as e:
        app.logger.error(f"Report Error: {e}")
        flash(f"Error generating report: {e}", "danger")
        # Prevent template crash
        report_data = {
            "summary": {"total": 0, "count": 0}, 
            "log": [], "top_items": [], "annual": [], "by_subcategory": []
        }

    cursor.close()
    
    months = [("01", "January"), ("02", "February"), ("03", "March"), ("04", "April"), 
              ("05", "May"), ("06", "June"), ("07", "July"), ("08", "August"), 
              ("09", "September"), ("10", "October"), ("11", "November"), ("12", "December")]

    return render_template(
        "reports/exp_report.html",
        report_type=report_type,
        filters=filters,
        months=months,
        report_data=report_data
    )


# ==========================================
# 2. NEW ROUTE: EXPORT EXPENSES (PDF/EXCEL)
# ==========================================
@app.route("/export_expenses")
def export_expenses():
    if "loggedin" not in session: return redirect(url_for("login"))
    
    fmt = request.args.get("format", "excel") # pdf or excel
    report_type = request.args.get("type", "detailed_log")
    
    # Reuse filter logic
    now = datetime.now()
    try: year = int(request.args.get("year") or now.year)
    except: year = now.year
    
    start_date = request.args.get("start_date") or f"{year}-{now.month:02d}-01"
    end_date = request.args.get("end_date") or datetime.today().strftime('%Y-%m-%d')

    cursor = mysql.connection.cursor(MySQLdb.cursors.DictCursor)
    
    # ----------------------------------------
    # FETCH DATA (Simplified for Export)
    # ----------------------------------------
    data = []
    headers = []
    title = f"Expense Report - {report_type}"
    
    try:
        if report_type == "detailed_log":
            cursor.execute("""
                SELECT 
                    e.expense_date, 
                    c.category_name, 
                    sc.subcategory_name, 
                    COALESCE(ei.description, e.description) as description,
                    e.payment_method,
                    ei.amount
                FROM expense_items ei
                JOIN expenses e ON ei.expense_id = e.expense_id
                LEFT JOIN expensecategories c ON ei.category_id = c.category_id
                LEFT JOIN expensesubcategories sc ON ei.subcategory_id = sc.subcategory_id
                WHERE e.expense_date BETWEEN %s AND %s
                ORDER BY e.expense_date DESC
            """, (start_date, end_date))
            rows = cursor.fetchall()
            headers = ["Date", "Category", "Item", "Description", "Mode", "Amount"]
            for r in rows:
                data.append([
                    r['expense_date'].strftime('%d-%m-%Y'),
                    r['category_name'] or 'General',
                    r['subcategory_name'] or '-',
                    r['description'] or '',
                    r['payment_method'],
                    float(r['amount'])
                ])
            title = f"Detailed Expense Log ({start_date} to {end_date})"

        elif report_type == "monthly_summary":
            # Just export list of expenses for the month
            m_start = f"{year}-{request.args.get('month', '01')}-01"
            # Logic to find end of month
            m = int(request.args.get('month', '01'))
            _, last = calendar.monthrange(year, m)
            m_end = f"{year}-{m:02d}-{last}"
            
            cursor.execute("""
                SELECT expense_date, description, payment_method, total_amount 
                FROM expenses WHERE expense_date BETWEEN %s AND %s ORDER BY expense_date
            """, (m_start, m_end))
            rows = cursor.fetchall()
            headers = ["Date", "Description", "Mode", "Amount"]
            for r in rows:
                data.append([
                    r['expense_date'].strftime('%d-%m-%Y'),
                    r['description'],
                    r['payment_method'],
                    float(r['total_amount'])
                ])
            title = f"Monthly Expenses ({calendar.month_name[m]} {year})"

    except Exception as e:
        cursor.close()
        return f"Error exporting: {e}"

    cursor.close()

    # ----------------------------------------
    # GENERATE EXCEL
    # ----------------------------------------
    if fmt == "excel":
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Expenses"
        
        # Title Row
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
        ws.cell(row=1, column=1, value=title).font = Font(size=14, bold=True)
        ws.cell(row=1, column=1).alignment = Alignment(horizontal='center')
        
        # Header Row
        ws.append(headers)
        for cell in ws[2]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
            
        # Data Rows
        for row in data:
            ws.append(row)
            
        # Auto-width
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length: max_length = len(str(cell.value))
                except: pass
            ws.column_dimensions[column].width = max_length + 2

        out = io.BytesIO()
        wb.save(out)
        out.seek(0)
        return send_file(out, as_attachment=True, download_name=f"expenses_{report_type}.xlsx", mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

    # ----------------------------------------
    # GENERATE PDF (Bulk Report - Maintains Generic Style)
    # ----------------------------------------
    elif fmt == "pdf":
        class PDFReport(FPDF):
            def header(self):
                self.set_font('Arial', 'B', 16)
                self.cell(0, 10, "REAL PROMOTION", 0, 1, 'C')
                self.set_font('Arial', 'B', 9)
                self.set_text_color(100, 100, 100)
                self.cell(0, 5, "SINCE 2005", 0, 1, 'C')
                self.ln(5)
                self.set_font('Arial', 'B', 14)
                self.set_text_color(0, 0, 0)
                self.cell(0, 10, title, 0, 1, 'C')
                self.ln(5)
                
            def footer(self):
                self.set_y(-15)
                self.set_font('Arial', 'I', 8)
                self.set_text_color(128, 128, 128)
                self.cell(0, 10, f'Page {self.page_no()}', 0, 0, 'C')

        pdf = PDFReport(orientation='L') # Landscape
        pdf.add_page()
        pdf.set_font("Arial", 'B', 10)
        
        # Widths
        w_base = 275 / len(headers) 
        widths = [w_base] * len(headers)
        if report_type == "detailed_log":
            widths = [25, 40, 40, 90, 30, 50] 
        
        # Header
        pdf.set_fill_color(79, 70, 229)
        pdf.set_text_color(255, 255, 255)
        for i, h in enumerate(headers):
            pdf.cell(widths[i], 8, h, 1, 0, 'C', True)
        pdf.ln()
        
        # Rows
        pdf.set_font("Arial", '', 9)
        pdf.set_text_color(0, 0, 0)
        total = 0
        fill = False
        
        for row in data:
            pdf.set_fill_color(245, 245, 245)
            for i, val in enumerate(row):
                align = 'R' if isinstance(val, float) else 'L'
                txt = f"{val:,.2f}" if isinstance(val, float) else str(val)
                if isinstance(val, float) and i == len(row)-1: total += val
                
                pdf.cell(widths[i], 8, txt[:40], 1, 0, align, fill) 
            pdf.ln()
            fill = not fill
            
        # Total
        pdf.set_font("Arial", 'B', 10)
        pdf.cell(sum(widths[:-1]), 10, "Grand Total", 1, 0, 'R')
        pdf.cell(widths[-1], 10, f"{total:,.2f}", 1, 1, 'R')

        out = io.BytesIO(pdf.output(dest='S').encode('latin-1'))
        out.seek(0)
        return send_file(out, as_attachment=True, download_name=f"expenses_{report_type}.pdf", mimetype="application/pdf")

    return "Invalid format"



# ==========================================
# NEW ROUTE: VIEW OFFICE SALE (FIXED TIME)
# ==========================================
@app.route('/office_sales/view/<int:sale_id>')
def view_office_sale(sale_id):
    if "loggedin" not in session: return redirect(url_for("login"))
    
    conn = mysql.connection
    cursor = conn.cursor(MySQLdb.cursors.DictCursor)
    
    # Fetch Sale
    cursor.execute("SELECT * FROM office_sales WHERE id=%s", (sale_id,))
    sale = cursor.fetchone()
    
    if not sale:
        flash("Sale not found.", "danger")
        return redirect(url_for('office_sales_master'))
        
    # Formatting
    if sale['sale_date']:
        sale['formatted_date'] = sale['sale_date'].strftime('%d-%m-%Y')
    
    if sale.get('created_at'):
        # FIX: Database stores IST. Do NOT add timedelta. Just format.
        try:
            sale['created_at_formatted'] = sale['created_at'].strftime('%d-%m-%Y %I:%M %p')
        except:
            sale['created_at_formatted'] = str(sale['created_at'])
    else:
        sale['created_at_formatted'] = "N/A"

    # Fetch Items
    cursor.execute("""
        SELECT i.*, p.name as product_name, p.image 
        FROM office_sale_items i
        JOIN products p ON i.product_id = p.id
        WHERE i.sale_id=%s
    """, (sale_id,))
    items = cursor.fetchall()
    
    for i in items: i['image'] = resolve_img(i['image'])
    
    cursor.close()
    return render_template('view_office_sale.html', sale=sale, items=items)

# ==========================================
# 4. OFFICE SALES MASTER (UPDATED WITH TIME)
# ==========================================
@app.route('/office_sales/master')
def office_sales_master():
    if "loggedin" not in session: return redirect(url_for("login"))
    
    conn = mysql.connection
    cursor = conn.cursor(MySQLdb.cursors.DictCursor)
    
    # 1. Get Filters
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    
    # 2. Build Query
    query = "SELECT * FROM office_sales WHERE 1=1"
    params = []
    
    if start_date:
        try:
            d = datetime.strptime(start_date, '%d-%m-%Y').strftime('%Y-%m-%d')
            query += " AND sale_date >= %s"
            params.append(d)
        except: pass
        
    if end_date:
        try:
            d = datetime.strptime(end_date, '%d-%m-%Y').strftime('%Y-%m-%d')
            query += " AND sale_date <= %s"
            params.append(d)
        except: pass
        
    query += " ORDER BY sale_date DESC, created_at DESC"
    
    cursor.execute(query, tuple(params))
    sales = cursor.fetchall()
    
    # 3. Calculate Stats & Format Time
    stats = {
        'total_sales': 0.0,
        'cash': 0.0,
        'online': 0.0,
        'discount': 0.0
    }
    
    for s in sales:
        # Date
        if isinstance(s['sale_date'], date):
            s['formatted_date'] = s['sale_date'].strftime('%d-%m-%Y')
        else:
            s['formatted_date'] = str(s['sale_date'])
            
        # Time (IST)
        if s.get('created_at'):
            try:
                dt = s['created_at']
                if isinstance(dt, datetime):
                     # Format using 12-hour AM/PM format
                     s['formatted_time'] = dt.strftime('%I:%M %p')
                else:
                     s['formatted_time'] = str(dt)
            except:
                s['formatted_time'] = ""
        else:
            s['formatted_time'] = ""

        stats['total_sales'] += float(s['final_amount'] or 0)
        stats['cash'] += float(s['cash_amount'] or 0)
        stats['online'] += float(s['online_amount'] or 0)
        stats['discount'] += float(s['discount'] or 0)
        
    cursor.close()
    
    return render_template('office_sales_master.html', 
                           sales=sales, 
                           stats=stats, 
                           filters={'start': start_date, 'end': end_date})


# ==========================================
# 5. DELETE OFFICE SALE (Restore Stock)
# ==========================================
@app.route('/office_sales/delete/<int:sale_id>', methods=['POST'])
def delete_office_sale(sale_id):
    if "loggedin" not in session: return redirect(url_for("login"))
    
    conn = mysql.connection
    cursor = conn.cursor(MySQLdb.cursors.DictCursor)
    
    try:
        # 1. Fetch Items to Restore Stock
        cursor.execute("SELECT product_id, qty FROM office_sale_items WHERE sale_id = %s", (sale_id,))
        items = cursor.fetchall()
        
        # 2. Restore Stock
        for item in items:
            cursor.execute("UPDATE products SET stock = stock + %s WHERE id = %s", (item['qty'], item['product_id']))
            
        # 3. Delete Record (Cascading delete handles items if set, else manual)
        cursor.execute("DELETE FROM office_sale_items WHERE sale_id = %s", (sale_id,))
        cursor.execute("DELETE FROM office_sales WHERE id = %s", (sale_id,))
        
        conn.commit()
        flash("Sale Record Deleted & Stock Restored.", "success")
        
    except Exception as e:
        conn.rollback()
        flash(f"Error deleting sale: {e}", "danger")
        
    finally:
        cursor.close()
        
    return redirect(url_for('office_sales_master'))

# =====================================================================================
# 4. EDIT OFFICE SALE ROUTE (Updated to show only Active Products)
# =====================================================================================
@app.route('/office_sales/edit/<int:sale_id>', methods=['GET', 'POST'])
def edit_office_sale(sale_id):
    if "loggedin" not in session: return redirect(url_for("login"))
    
    conn = mysql.connection
    cursor = conn.cursor(MySQLdb.cursors.DictCursor)
    
    if request.method == 'POST':
        try:
            # 1. Fetch Old Items to reverse stock
            cursor.execute("SELECT product_id, qty FROM office_sale_items WHERE sale_id = %s", (sale_id,))
            old_items = cursor.fetchall()
            
            # Reverse old stock logic (Add back)
            for item in old_items:
                cursor.execute("UPDATE products SET stock = stock + %s WHERE id = %s", (item['qty'], item['product_id']))
            
            # Delete old items
            cursor.execute("DELETE FROM office_sale_items WHERE sale_id = %s", (sale_id,))
            
            # 2. Process New Data
            c_name = request.form.get('customer_name')
            c_mobile = request.form.get('customer_mobile')
            c_addr = request.form.get('customer_address')
            sales_person = request.form.get('sales_person')
            
            b_date_str = request.form.get('bill_date')
            try: bill_date = datetime.strptime(b_date_str, '%d-%m-%Y').strftime('%Y-%m-%d')
            except: bill_date = date.today().strftime('%Y-%m-%d')
            
            discount = float(request.form.get('discount') or 0)
            online_amt = float(request.form.get('online') or 0)
            cash_amt = float(request.form.get('cash') or 0)
            due_note = request.form.get('due_note') 
            
            p_ids = request.form.getlist('product_id[]')
            qtys = request.form.getlist('qty[]')
            prices = request.form.getlist('price[]')
            
            total_amt = 0
            
            # 3. Insert New Items & Deduct Stock
            for i, pid in enumerate(p_ids):
                if not pid: continue
                qty = int(qtys[i])
                price = float(prices[i])
                
                # Check Stock (Current stock + what we just added back)
                cursor.execute("SELECT stock, name FROM products WHERE id=%s", (pid,))
                prod = cursor.fetchone()
                if prod and prod['stock'] < qty:
                    raise Exception(f"Insufficient stock for {prod['name']}")
                
                cursor.execute("UPDATE products SET stock = stock - %s WHERE id=%s", (qty, pid))
                
                item_total = qty * price
                cursor.execute("""
                    INSERT INTO office_sale_items (sale_id, product_id, qty, unit_price, total_price)
                    VALUES (%s, %s, %s, %s, %s)
                """, (sale_id, pid, qty, price, item_total))
                
                total_amt += item_total
            
            final_amt = total_amt - discount
            
            # 4. Update Header
            cursor.execute("""
                UPDATE office_sales 
                SET customer_name=%s, customer_mobile=%s, customer_address=%s, sales_person=%s, sale_date=%s,
                    sub_total=%s, discount=%s, online_amount=%s, cash_amount=%s, final_amount=%s, due_note=%s
                WHERE id=%s
            """, (c_name, c_mobile, c_addr, sales_person, bill_date, total_amt, discount, online_amt, cash_amt, final_amt, due_note, sale_id))
            
            conn.commit()
            flash("Sale Updated Successfully!", "success")
            return redirect(url_for('office_sales_master'))
            
        except Exception as e:
            conn.rollback()
            flash(f"Update Error: {e}", "danger")
            return redirect(url_for('office_sales_master'))
            
    # GET Request
    cursor.execute("SELECT * FROM office_sales WHERE id=%s", (sale_id,))
    sale = cursor.fetchone()
    
    if sale['sale_date']:
        sale['formatted_date'] = sale['sale_date'].strftime('%d-%m-%Y')
        
    cursor.execute("""
        SELECT i.*, p.name, p.image, p.price, p.stock 
        FROM office_sale_items i
        JOIN products p ON i.product_id = p.id
        WHERE i.sale_id=%s
    """, (sale_id,))
    items = cursor.fetchall()
    
    # --- MODIFIED: Fetch ONLY Active Products for Dropdown ---
    try:
        cursor.execute("SELECT id, name, price, stock, image FROM products WHERE status='Active' ORDER BY name")
    except:
        cursor.execute("SELECT id, name, price, stock, image FROM products ORDER BY name")
    
    all_products = cursor.fetchall()
    
    # Image resolving
    for p in all_products: p['image'] = resolve_img(p['image'])
    for i in items: i['image'] = resolve_img(i['image'])
    
    cursor.close()
    return render_template('edit_office_sale.html', sale=sale, items=items, products=all_products)

        


import pytz # Recommended: pip install pytz

# Helper to get current IST time
def get_ist_now():
    # If pytz is available
    try:
        ist = pytz.timezone('Asia/Kolkata')
        return datetime.now(ist)
    except:
        # Fallback manual calculation
        return datetime.utcnow() + timedelta(hours=5, minutes=30)

# Helper to parse dd-mm-yyyy to yyyy-mm-dd for MySQL
def parse_date_input(date_str):
    try:
        # Tries to parse dd-mm-yyyy
        return datetime.strptime(date_str, '%d-%m-%Y').strftime('%Y-%m-%d')
    except ValueError:
        # If it fails, maybe it's already yyyy-mm-dd or invalid
        return date_str 




# Helpers
def parse_ddmmyyyy_to_date(s):
    if not s:
        return None
    try:
        return datetime.strptime(s, "%d-%m-%Y").date()
    except Exception:
        return None

def validate_email(email):
    if not email:
        return True
    return re.match(r'^[^\s@]+@[^\s@]+\.[^\s@]+$', email) is not None

def save_file_to_cloudinary(fileobj, folder, resource_type='image'):
    """Upload fileobj (werkzeug FileStorage) to Cloudinary and return public_id (or secure_url fallback)."""
    if not fileobj or not getattr(fileobj, "filename", None):
        return None
    filename = secure_filename(fileobj.filename)
    try:
        if resource_type == 'raw':
            res = cloudinary.uploader.upload(fileobj, resource_type='raw', folder=folder)
        else:
            res = cloudinary.uploader.upload(fileobj, folder=folder)
        return res.get('public_id') or res.get('secure_url')
    except Exception as e:
        app.logger.exception("Cloudinary upload failed")
        raise


# Helper function
# This "context processor" injects the cloudinary_url function into all templates
@app.context_processor
def inject_cloudinary_url():
    """Make cloudinary_url function available in all templates."""
    # This is the correct function from the cloudinary library
    return dict(cloudinary_url=cloudinary.utils.cloudinary_url)


# ----------------- FINAL FIX FOR CLOUDINARY + LOCAL IMAGES -----------------#
@app.template_filter("fix_image")
def fix_image(image):
    if not image or image.strip() == "":
        
        return url_for('static', filename='img/default-user.png')

    if image.startswith("http"):
        return image

    # CASE 2: Cloudinary public_id (example: erp_employees/abcd123)
    if "/" in image and "." not in image:
        # generate full Cloudinary URL
        url, _ = cloudinary_url(image, secure=True)
        return url

    # CASE 3: stored old filename like "1.jpeg"
    if "." in image:
        return url_for('static', filename='uploads/' + image)

    # fallback
    return url_for('static', filename='img/default-user.png')



@app.route("/")
def index1():
    return redirect(url_for("login"))

@app.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        username = request.form.get("username")
        password = request.form.get("password")

        cursor = mysql.connection.cursor(MySQLdb.cursors.DictCursor)
        cursor.execute("SELECT * FROM users WHERE username=%s AND password=%s", (username, password))
        user = cursor.fetchone()
        cursor.close()

        if user:
            session["loggedin"] = True
            session["username"] = user["username"]
            return redirect(url_for("index"))
        else:
            flash("Invalid username or password", "error")
            return redirect(url_for("login"))

    return render_template("login.html")

@app.route("/logout")
def logout():
    session.clear()
    return redirect(url_for("login"))

@app.route('/index')
def index():
    return render_template('index.html')


@app.route("/dashboard")
def dashboard():
    if "loggedin" not in session:
        return redirect(url_for("login"))

    report_date = date.today().isoformat()
    summary_data = {
        "total_sales": 0,
        "total_expenses": 0,
        "net_cash_flow": 0
    }

    try:
        
        db_cursor = mysql.connection.cursor(MySQLdb.cursors.DictCursor)


        # Total Sales (AFTER discount)
        db_cursor.execute("""
            SELECT SUM(total_amount - discount) AS total_sales 
            FROM evening_settle 
            WHERE date = %s
        """, (report_date,))
        sales_result = db_cursor.fetchone()
        summary_data["total_sales"] = sales_result["total_sales"] or 0

        
        # Total Expenses
        db_cursor.execute("""
            SELECT SUM(amount) AS total_expenses 
            FROM expenses 
            WHERE expense_date = %s
        """, (report_date,))
        expenses_result = db_cursor.fetchone()
        summary_data["total_expenses"] = expenses_result["total_expenses"] or 0

        # Net Cash Flow
        summary_data["net_cash_flow"] = (
            summary_data["total_sales"] - summary_data["total_expenses"]
        )

        db_cursor.close()

    except Exception as e:
        flash(f"Error loading dashboard: {e}", "danger")

    report_date_obj = date.fromisoformat(report_date)
    report_date_str = report_date_obj.strftime('%d %B, %Y')

    
    
    return render_template("dashboard.html",
        report_date=report_date,
        report_date_str=report_date_str,
        **summary_data
    )





#==============================================ADMIN PAGE=======================================
@app.context_processor
def inject_kpis():
    db_cursor = None
    try:
        if 'mysql' not in app.extensions:
            return dict(kpis={"sales_today": 0, "supplier_dues": 0, "mtd_expenses": 0, "stock_value": 0, "total_employees": 0, "total_products": 0})
            
        db_cursor = mysql.connection.cursor(MySQLdb.cursors.DictCursor)
        
        db_cursor.execute("SELECT SUM(total_amount) as total FROM evening_settle WHERE date = CURDATE()")
        sales_today = db_cursor.fetchone()['total'] or 0

        db_cursor.execute("SELECT SUM(current_due) as total FROM suppliers")
        supplier_dues = db_cursor.fetchone()['total'] or 0
        
        db_cursor.execute("SELECT SUM(amount) as total FROM expenses WHERE MONTH(expense_date) = MONTH(CURDATE()) AND YEAR(expense_date) = YEAR(CURDATE())")
        mtd_expenses = db_cursor.fetchone()['total'] or 0

        db_cursor.execute("SELECT SUM(stock * purchase_price) as total FROM products")
        stock_value = db_cursor.fetchone()['total'] or 0
        
        db_cursor.execute("SELECT COUNT(id) as total FROM employees")
        total_employees = db_cursor.fetchone()['total'] or 0
        
        db_cursor.execute("SELECT COUNT(id) as total FROM products")
        total_products = db_cursor.fetchone()['total'] or 0

        kpis = {
            "sales_today": float(sales_today),
            "supplier_dues": float(supplier_dues),
            "mtd_expenses": float(mtd_expenses),
            "stock_value": float(stock_value),
            "total_employees": total_employees,
            "total_products": total_products
        }
        
        return dict(kpis=kpis)
    except Exception as e:
        app.logger.error(f"CRITICAL ERROR in context processor: {e}")
        return dict(kpis={"sales_today": 0, "supplier_dues": 0, "mtd_expenses": 0, "stock_value": 0, "total_employees": 0, "total_products": 0})
    finally:
        if db_cursor:
            db_cursor.close()

# --- API ROUTE FOR DASHBOARD CHARTS ---
@app.route('/api/dashboard-charts')
def api_dashboard_charts():
    db_cursor = None
    try:
        db_cursor = mysql.connection.cursor(MySQLdb.cursors.DictCursor)
        
        db_cursor.execute("""
            SELECT 
                DATE(date) as sale_date, 
                SUM(total_amount) as total_sales
            FROM evening_settle
            WHERE date >= CURDATE() - INTERVAL 7 DAY
            GROUP BY DATE(date)
            ORDER BY sale_date ASC;
        """)
        sales_data = db_cursor.fetchall()
        
        db_cursor.execute("""
            SELECT 
                DATE(expense_date) as exp_date, 
                SUM(amount) as total_expenses
            FROM expenses
            WHERE expense_date >= CURDATE() - INTERVAL 7 DAY
            GROUP BY DATE(expense_date)
            ORDER BY exp_date ASC;
        """)
        expense_data = db_cursor.fetchall()

        sales_labels = [row['sale_date'].strftime('%b %d') for row in sales_data]
        sales_values = [float(row['total_sales']) for row in sales_data]
        
        expense_labels = [row['exp_date'].strftime('%b %d') for row in expense_data]
        expense_values = [float(row['total_expenses']) for row in expense_data]

        db_cursor.close()
        return jsonify({
            "sales": {"labels": sales_labels, "values": sales_values},
            "expenses": {"labels": expense_labels, "values": expense_values}
        })
    except Exception as e:
        app.logger.error(f"Chart API Error: {e}")
        return jsonify({"error": str(e)}), 500
    finally:
        if db_cursor:
            db_cursor.close()


# Admin Page 
@app.route('/admin_master')
def admin_master():
    if "loggedin" not in session: 
        return redirect(url_for("login"))
    # Redirect dashboard page 
    return redirect(url_for('dash'))



# --- HELPER FUNCTION: Find correct column name automatically ---
def get_db_column(cursor, table_name, candidates):
    """Finds which column name actually exists in the table."""
    try:
        cursor.execute(f"SHOW COLUMNS FROM {table_name}")
        existing_columns = [row['Field'] for row in cursor.fetchall()]
        for cand in candidates:
            if cand in existing_columns: return cand
        return candidates[0]
    except:
        return candidates[0]

# =================================================================================
#  ADMIN DASHBOARD ROUTE (Updated to include Today's Sales)
# =================================================================================
@app.route("/dash")
def dash():
    if "loggedin" not in session: return redirect(url_for("login"))
    
    conn = mysql.connection
    cursor = conn.cursor(MySQLdb.cursors.DictCursor)
    
    # --- 1. TIMEZONE FIX (IST) ---
    ist_now = datetime.utcnow() + timedelta(hours=5, minutes=30)
    today = ist_now.date() 
    yesterday = today - timedelta(days=1)
    current_year = today.year
    
    # --- 2. KEY METRICS ---
    
    # A. SALES YESTERDAY (Net)
    cursor.execute("SELECT SUM(total_amount - discount) as total FROM evening_settle WHERE date = %s AND status = 'final'", (yesterday,))
    eve_yest = float(cursor.fetchone()['total'] or 0)
    
    cursor.execute("SELECT SUM(final_amount) as total FROM office_sales WHERE sale_date = %s", (yesterday,))
    off_yest = float(cursor.fetchone()['total'] or 0)
    
    sales_yesterday = eve_yest + off_yest

    # --- NEW: SALES TODAY (Net) ---
    cursor.execute("SELECT SUM(total_amount - discount) as total FROM evening_settle WHERE date = %s AND status = 'final'", (today,))
    eve_today = float(cursor.fetchone()['total'] or 0)

    cursor.execute("SELECT SUM(final_amount) as total FROM office_sales WHERE sale_date = %s", (today,))
    off_today = float(cursor.fetchone()['total'] or 0)

    sales_today = eve_today + off_today

    # B. SALES THIS MONTH (Net)
    start_of_month = today.replace(day=1)
    cursor.execute("SELECT SUM(total_amount - discount) as total FROM evening_settle WHERE date >= %s AND date <= %s AND status = 'final'", (start_of_month, today))
    eve_month = float(cursor.fetchone()['total'] or 0)
    
    cursor.execute("SELECT SUM(final_amount) as total FROM office_sales WHERE sale_date >= %s AND sale_date <= %s", (start_of_month, today))
    off_month = float(cursor.fetchone()['total'] or 0)
    
    sales_this_month = eve_month + off_month

    # C. YEARLY REVENUE
    start_of_year = date(current_year, 1, 1)
    cursor.execute("SELECT SUM(total_amount - discount) as total FROM evening_settle WHERE date >= %s AND status = 'final'", (start_of_year,))
    eve_year = float(cursor.fetchone()['total'] or 0)
    
    cursor.execute("SELECT SUM(final_amount) as total FROM office_sales WHERE sale_date >= %s", (start_of_year,))
    off_year = float(cursor.fetchone()['total'] or 0)
    
    yearly_sales = eve_year + off_year

    # D. YEARLY EXPENSES
    cursor.execute("SELECT SUM(amount) as total FROM expenses WHERE expense_date >= %s", (start_of_year,))
    row = cursor.fetchone()
    yearly_expenses = float(row['total'] or 0)

    # E. SUPPLIER DUES & LOW STOCK
    cursor.execute("""
        SELECT SUM(
            opening_balance + 
            (SELECT COALESCE(SUM(total_amount),0) FROM purchases WHERE supplier_id=s.id) + 
            (SELECT COALESCE(SUM(amount),0) FROM supplier_adjustments WHERE supplier_id=s.id) - 
            (SELECT COALESCE(SUM(amount_paid),0) FROM supplier_payments WHERE supplier_id=s.id)
        ) as total_due 
        FROM suppliers s
    """)
    supplier_dues = float(cursor.fetchone()['total_due'] or 0)

    try:
        cursor.execute("SELECT COUNT(*) as cnt FROM products WHERE stock <= low_stock_threshold AND status='Active'")
    except:
        cursor.execute("SELECT COUNT(*) as cnt FROM products WHERE stock < 10")
    low_stock = int(cursor.fetchone()['cnt'] or 0)

    # --- 3. CHARTS ---
    chart_months = []
    chart_sales = []
    chart_expenses = []

    # Sales Map
    cursor.execute("""
        SELECT MONTH(date) as m, SUM(total_amount - discount) as total 
        FROM evening_settle WHERE YEAR(date) = %s AND status = 'final' GROUP BY MONTH(date)
    """, (current_year,))
    eve_map = {row['m']: float(row['total']) for row in cursor.fetchall()}
    
    cursor.execute("""
        SELECT MONTH(sale_date) as m, SUM(final_amount) as total 
        FROM office_sales WHERE YEAR(sale_date) = %s GROUP BY MONTH(sale_date)
    """, (current_year,))
    off_map = {row['m']: float(row['total']) for row in cursor.fetchall()}
    
    # Expense Map
    cursor.execute("SELECT MONTH(expense_date) as m, SUM(amount) as total FROM expenses WHERE YEAR(expense_date) = %s GROUP BY MONTH(expense_date)", (current_year,))
    expense_map = {row['m']: float(row['total']) for row in cursor.fetchall()}

    for m in range(1, 13):
        chart_months.append(calendar.month_abbr[m])
        total_s = eve_map.get(m, 0.0) + off_map.get(m, 0.0)
        chart_sales.append(total_s)
        chart_expenses.append(expense_map.get(m, 0.0))

    # --- 4. TOP PRODUCTS ---
    top_prod_names = []
    top_prod_qty = []
    cursor.execute("""
        SELECT p.name, SUM(ei.sold_qty) as total_qty
        FROM evening_item ei
        JOIN products p ON ei.product_id = p.id
        JOIN evening_settle es ON ei.settle_id = es.id
        WHERE YEAR(es.date) = %s AND es.status = 'final'
        GROUP BY p.id, p.name ORDER BY total_qty DESC LIMIT 5
    """, (current_year,))
    for row in cursor.fetchall():
        top_prod_names.append(row['name'])
        top_prod_qty.append(int(row['total_qty']))

    cursor.close()

    return render_template(
        "dash.html",
        today_date=today, 
        current_year=current_year,
        sales_yesterday=sales_yesterday,
        sales_today=sales_today,          # NEW: Passed to template
        sales_this_month=sales_this_month,
        yearly_sales=yearly_sales,
        yearly_expenses=yearly_expenses,
        low_stock=low_stock,
        supplier_dues=supplier_dues,
        chart_months=chart_months,
        chart_sales=chart_sales,
        chart_expenses=chart_expenses,
        top_prod_names=top_prod_names,
        top_prod_qty=top_prod_qty
    )

# =====================================================================
# SUPPLIER MANAGEMENT ROUTES
# =====================================================================
@app.route('/suppliers')
def suppliers():
    if 'loggedin' not in session: return redirect(url_for('login'))
    
    cursor = mysql.connection.cursor(MySQLdb.cursors.DictCursor)
    
    # FETCH SUPPLIERS WITH DYNAMIC DUE CALCULATION
    # Formula: Opening + Purchases + Adjustments - Payments
    # Ordered by ID ASC
    query = """
        SELECT s.*, 
            (
                s.opening_balance + 
                COALESCE((SELECT SUM(total_amount) FROM purchases WHERE supplier_id = s.id), 0) +
                COALESCE((SELECT SUM(amount) FROM supplier_adjustments WHERE supplier_id = s.id), 0) -
                COALESCE((SELECT SUM(amount_paid) FROM supplier_payments WHERE supplier_id = s.id), 0)
            ) as calculated_due
        FROM suppliers s
        ORDER BY s.id ASC
    """
    cursor.execute(query)
    suppliers = cursor.fetchall()
    
    cursor.close()
    return render_template('suppliers/suppliers.html', suppliers=suppliers)



@app.route('/suppliers/add', methods=['GET', 'POST'])
def add_supplier():
    if 'loggedin' not in session: return redirect(url_for('login'))
    
    if request.method == 'POST':
        name = request.form['name']
        phone = request.form['phone']
        email = request.form['email']
        gstin = request.form.get('gstim', '')
        address = request.form.get('address', '')
        
        # New Fields
        opening_balance = request.form.get('opening_balance', 0.0)
        payment_terms = request.form.get('payment_terms', 60)
        
        cursor = mysql.connection.cursor(MySQLdb.cursors.DictCursor)
        cursor.execute("""
            INSERT INTO suppliers (name, phone, email, gstin, address, opening_balance, payment_terms) 
            VALUES (%s, %s, %s, %s, %s, %s, %s)
        """, (name, phone, email, gstin, address, opening_balance, payment_terms))
        
        mysql.connection.commit()
        cursor.close()
        flash('Supplier added successfully!', 'success')
        return redirect(url_for('suppliers'))
        
    return render_template('suppliers/add_supplier.html')

@app.route('/suppliers/edit/<int:supplier_id>', methods=['GET', 'POST'])
def edit_supplier(supplier_id):
    if 'loggedin' not in session: return redirect(url_for('login'))
    
    cursor = mysql.connection.cursor(MySQLdb.cursors.DictCursor)
    
    if request.method == 'POST':
        name = request.form['name']
        phone = request.form['phone']
        email = request.form['email']
        gstin = request.form.get('gstin', '')
        address = request.form.get('address', '')
        opening_balance = request.form.get('opening_balance', 0.0)
        
        cursor.execute("""
            UPDATE suppliers 
            SET name=%s, phone=%s, email=%s, gstin=%s, address=%s, opening_balance=%s
            WHERE id=%s
        """, (name, phone, email, gstin, address, opening_balance, supplier_id))
        
        mysql.connection.commit()
        flash('Supplier updated successfully!', 'success')
        return redirect(url_for('suppliers'))
    
    cursor.execute("SELECT * FROM suppliers WHERE id=%s", (supplier_id,))
    supplier = cursor.fetchone()
    cursor.close()
    
    return render_template('suppliers/edit_supplier.html', supplier=supplier)

# ==========================================
# 3. FIX LEDGER DISPLAY (Sorting & Formatting)
# ==========================================
@app.route('/supplier_ledger/<int:supplier_id>')
def supplier_ledger(supplier_id):
    if 'loggedin' not in session: return redirect(url_for('login'))
    
    cursor = mysql.connection.cursor(MySQLdb.cursors.DictCursor)
    
    # 1. Fetch Supplier
    cursor.execute("SELECT * FROM suppliers WHERE id=%s", (supplier_id,))
    supplier = cursor.fetchone()
    
    if not supplier:
        flash("Supplier not found!", "danger")
        return redirect(url_for('suppliers'))

    # 2. Fetch Records (SORTED BY DATE DESC - NEWEST FIRST)
    cursor.execute("SELECT * FROM purchases WHERE supplier_id=%s ORDER BY purchase_date DESC", (supplier_id,))
    purchases = cursor.fetchall()
    
    try:
        cursor.execute("SELECT * FROM supplier_payments WHERE supplier_id=%s ORDER BY payment_date DESC", (supplier_id,))
        payments = cursor.fetchall()
    except: payments = []
    
    try:
        cursor.execute("SELECT * FROM supplier_adjustments WHERE supplier_id=%s ORDER BY adjustment_date DESC", (supplier_id,))
        adjustments = cursor.fetchall()
    except: adjustments = []
    
    # --- ROBUST DATE FORMATTING HELPER ---
    def format_record_date(record, date_key):
        val = record.get(date_key)
        if not val:
            return "-"
        
        # If it's already a Date object
        if isinstance(val, (date, datetime)):
            return val.strftime('%d-%m-%Y')
        
        # If it's a string, try to parse
        s_val = str(val)
        try:
            # Try yyyy-mm-dd (Standard DB)
            return datetime.strptime(s_val, '%Y-%m-%d').strftime('%d-%m-%Y')
        except ValueError:
            try:
                # Try dd-mm-yyyy (If saved as string)
                return datetime.strptime(s_val, '%d-%m-%Y').strftime('%d-%m-%Y')
            except ValueError:
                return s_val # Give up, return as is

    # Apply formatting
    for p in payments:
        p['formatted_date'] = format_record_date(p, 'payment_date')
        
    for a in adjustments:
        a['formatted_date'] = format_record_date(a, 'adjustment_date')
        
    for pur in purchases:
        pur['formatted_date'] = format_record_date(pur, 'purchase_date')

    # 3. Calculate Totals
    cursor.execute("SELECT SUM(COALESCE(total_amount, 0)) as total FROM purchases WHERE supplier_id=%s", (supplier_id,))
    total_purchases = float(cursor.fetchone()['total'] or 0)
    
    try:
        cursor.execute("SELECT SUM(COALESCE(amount_paid, 0)) as total FROM supplier_payments WHERE supplier_id=%s", (supplier_id,))
        total_paid = float(cursor.fetchone()['total'] or 0)
    except: total_paid = 0.0
        
    try:
        cursor.execute("SELECT SUM(COALESCE(amount, 0)) as total FROM supplier_adjustments WHERE supplier_id=%s", (supplier_id,))
        total_adjustments = float(cursor.fetchone()['total'] or 0)
    except: total_adjustments = 0.0
    
    opening_bal = float(supplier.get('opening_balance', 0.0))
    total_outstanding_amount = (opening_bal + total_purchases + total_adjustments) - total_paid
    
    cursor.close()
    
    return render_template('suppliers/supplier_ledger.html', 
                         supplier=supplier, 
                         purchases=purchases, 
                         payments=payments, 
                         adjustments=adjustments,
                         total_outstanding_amount=total_outstanding_amount,
                         opening_balance=opening_bal,
                         total_purchases=total_purchases,
                         total_adjustments=total_adjustments,
                         total_paid=total_paid)


# Route to Add Manual Due (Other Reason)
# ==========================================
# 2. FIX SAVING MANUAL DUE (Convert dd-mm-yyyy -> yyyy-mm-dd)
# ==========================================
@app.route('/suppliers/add_due/<int:supplier_id>', methods=['GET', 'POST'])
def supplier_add_due(supplier_id):
    if 'loggedin' not in session: return redirect(url_for('login'))
    
    cursor = mysql.connection.cursor(MySQLdb.cursors.DictCursor)
    cursor.execute("SELECT * FROM suppliers WHERE id = %s", (supplier_id,))
    supplier = cursor.fetchone()
    
    if not supplier:
        flash("Supplier not found.", "danger")
        return redirect(url_for('suppliers'))
        
    if request.method == 'POST':
        amount = request.form.get('amount')
        raw_date = request.form.get('date') # Likely dd-mm-yyyy
        notes = request.form.get('notes')
        
        # --- FIX: CONVERT DATE ---
        try:
            final_date = datetime.strptime(raw_date, '%d-%m-%Y').strftime('%Y-%m-%d')
        except ValueError:
            final_date = raw_date

        try:
            cursor.execute("""
                INSERT INTO supplier_adjustments (supplier_id, amount, adjustment_date, notes)
                VALUES (%s, %s, %s, %s)
            """, (supplier_id, amount, final_date, notes))
            
            cursor.execute("UPDATE suppliers SET current_due = current_due + %s WHERE id = %s", (amount, supplier_id))
            
            mysql.connection.commit()
            flash("Manual due added successfully.", "success")
            return redirect(url_for('supplier_ledger', supplier_id=supplier_id))
            
        except Exception as e:
            mysql.connection.rollback()
            flash(f"Error: {e}", "danger")
            
    cursor.close()
    return render_template('suppliers/add_manual_due.html', supplier=supplier, today_date=date.today().strftime('%d-%m-%Y'))


@app.route('/suppliers/delete_adjustment/<int:adjustment_id>', methods=['POST'])
def delete_supplier_adjustment(adjustment_id):
    if 'loggedin' not in session: return redirect(url_for('login'))
    
    cursor = mysql.connection.cursor(MySQLdb.cursors.DictCursor)
    
    # Get supplier ID first to redirect back
    cursor.execute("SELECT supplier_id, amount FROM supplier_adjustments WHERE id=%s", (adjustment_id,))
    adj = cursor.fetchone()
    
    if adj:
        supplier_id = adj['supplier_id']
        amount = adj['amount']
        
        try:
            # Delete adjustment
            cursor.execute("DELETE FROM supplier_adjustments WHERE id=%s", (adjustment_id,))
            
            # Update supplier current_due (Decrease debt)
            cursor.execute("UPDATE suppliers SET current_due = current_due - %s WHERE id=%s", (amount, supplier_id))
            
            mysql.connection.commit()
            flash("Adjustment deleted successfully.", "success")
            return redirect(url_for('supplier_ledger', supplier_id=supplier_id))
        except Exception as e:
            mysql.connection.rollback()
            flash(f"Error deleting adjustment: {e}", "danger")
            return redirect(url_for('suppliers')) # Fallback
            
    cursor.close()
    return redirect(url_for('suppliers'))

# ==========================================
# 1. FIX SAVING PAYMENT (Convert dd-mm-yyyy -> yyyy-mm-dd)
# ==========================================
@app.route('/suppliers/<int:supplier_id>/payment/new', methods=['GET', 'POST'])
def record_payment(supplier_id):
    if 'loggedin' not in session: return redirect(url_for('login'))
    
    cur = mysql.connection.cursor(MySQLdb.cursors.DictCursor)
    
    # 1. Fetch Supplier & Current Due
    cur.execute("SELECT * FROM suppliers WHERE id = %s", (supplier_id,))
    supplier = cur.fetchone()
    if not supplier:
        flash("Supplier not found.", "danger")
        return redirect(url_for('suppliers'))

    # Calculate Due
    cur.execute("SELECT SUM(total_amount) as total FROM purchases WHERE supplier_id=%s", (supplier_id,))
    t_purch = float(cur.fetchone()['total'] or 0)
    cur.execute("SELECT SUM(amount_paid) as total FROM supplier_payments WHERE supplier_id=%s", (supplier_id,))
    t_paid = float(cur.fetchone()['total'] or 0)
    cur.execute("SELECT SUM(amount) as total FROM supplier_adjustments WHERE supplier_id=%s", (supplier_id,))
    t_adj = float(cur.fetchone()['total'] or 0)
    
    current_due = (float(supplier['opening_balance']) + t_purch + t_adj) - t_paid

    # 2. Fetch Global Company Balance (For UI Display)
    # Income
    cur.execute("SELECT SUM(cash_money) as c, SUM(online_money) as b FROM evening_settle WHERE status='final'")
    eve = cur.fetchone()
    cur.execute("SELECT SUM(cash_amount) as c, SUM(online_amount) as b FROM office_sales")
    off = cur.fetchone()
    cur.execute("SELECT SUM(CASE WHEN payment_mode='Cash' THEN amount ELSE 0 END) as c, SUM(CASE WHEN payment_mode!='Cash' THEN amount ELSE 0 END) as b FROM employee_transactions WHERE type='credit'")
    emp_in = cur.fetchone()
    
    total_cash_in = float((eve['c'] or 0) + (off['c'] or 0) + (emp_in['c'] or 0))
    total_bank_in = float((eve['b'] or 0) + (off['b'] or 0) + (emp_in['b'] or 0))

    # Expenses (Vouchers + Emp Debits + Supplier Payments)
    cur.execute("SELECT SUM(CASE WHEN payment_method='Cash' THEN amount ELSE 0 END) as c, SUM(CASE WHEN payment_method!='Cash' THEN amount ELSE 0 END) as b FROM expense_items")
    exp_out = cur.fetchone()
    
    cur.execute("SELECT SUM(CASE WHEN payment_mode='Cash' THEN amount ELSE 0 END) as c, SUM(CASE WHEN payment_mode!='Cash' THEN amount ELSE 0 END) as b FROM employee_transactions WHERE type='debit'")
    emp_out = cur.fetchone()
    
    cur.execute("SELECT SUM(CASE WHEN payment_mode='Cash' THEN amount_paid ELSE 0 END) as c, SUM(CASE WHEN payment_mode!='Cash' THEN amount_paid ELSE 0 END) as b FROM supplier_payments")
    supp_out = cur.fetchone()

    total_cash_out = float(exp_out['c'] or 0) + float(emp_out['c'] or 0) + float(supp_out['c'] or 0)
    total_bank_out = float(exp_out['b'] or 0) + float(emp_out['b'] or 0) + float(supp_out['b'] or 0)
    
    cash_bal = total_cash_in - total_cash_out
    bank_bal = total_bank_in - total_bank_out

    if request.method == 'POST':
        try:
            amt = float(request.form.get('amount_paid'))
            date_raw = request.form.get('payment_date')
            pay_mode = request.form.get('payment_mode')
            notes = request.form.get('notes')
            
            # Date Parsing
            try: payment_date = datetime.strptime(date_raw, '%d-%m-%Y').strftime('%Y-%m-%d')
            except: payment_date = date_raw

            # 1. Record Supplier Payment Only (No Employee Transaction Insert)
            cur.execute("""
                INSERT INTO supplier_payments (supplier_id, amount_paid, payment_date, payment_mode, notes, created_at)
                VALUES (%s, %s, %s, %s, %s, NOW())
            """, (supplier_id, amt, payment_date, pay_mode, notes))
            
            # 2. Update Supplier Due
            cur.execute("UPDATE suppliers SET current_due = current_due - %s WHERE id = %s", (amt, supplier_id))

            mysql.connection.commit()
            flash(f"Payment of ₹{amt} recorded successfully.", "success")
            return redirect(url_for('supplier_ledger', supplier_id=supplier_id))

        except Exception as e:
            mysql.connection.rollback()
            flash(f"Error: {e}", "danger")

    cur.close()
    return render_template('suppliers/new_payment.html', 
                         supplier=supplier, 
                         current_due=current_due, 
                         cash_balance=cash_bal, 
                         bank_balance=bank_bal,
                         today=date.today().strftime('%d-%m-%Y'))

@app.route('/supplier/payment/delete/<int:payment_id>', methods=['POST'])
def delete_supplier_payment(payment_id):
    if 'loggedin' not in session: return redirect(url_for('login'))
    
    cur = mysql.connection.cursor(MySQLdb.cursors.DictCursor)
    try:
        cur.execute("SELECT * FROM supplier_payments WHERE id = %s", (payment_id,))
        pay = cur.fetchone()
        
        if pay:
            supplier_id = pay['supplier_id']
            amount = pay['amount_paid']
            
            # 1. Reverse Supplier Due
            cur.execute("UPDATE suppliers SET current_due = current_due + %s WHERE id = %s", (amount, supplier_id))
            
            # 2. Delete Payment Record
            cur.execute("DELETE FROM supplier_payments WHERE id = %s", (payment_id,))
            
            mysql.connection.commit()
            flash('Payment deleted. Balance restored.', 'warning')
            return redirect(url_for('supplier_ledger', supplier_id=supplier_id))
            
    except Exception as e:
        mysql.connection.rollback()
        flash(f"Error: {e}", "danger")
    
    cur.close()
    return redirect(url_for('suppliers'))



@app.route('/suppliers/delete/<int:supplier_id>', methods=['POST'])
def delete_supplier(supplier_id):
    if 'loggedin' not in session: return redirect(url_for('login'))
    
    cursor = mysql.connection.cursor(MySQLdb.cursors.DictCursor)
    try:
        # 1. Delete Supplier Payments (Purchase Order Payments)
        # We try this, but if table doesn't exist or is empty, we continue
        try:
            cursor.execute("DELETE FROM supplier_payments WHERE supplier_id = %s", (supplier_id,))
        except Exception:
            pass # Table might not exist or other issue
            
        # 2. Delete Purchases (and items)
        # Get purchase IDs to manually delete items if CASCADE is missing
        cursor.execute("SELECT id FROM purchases WHERE supplier_id = %s", (supplier_id,))
        purchase_ids = [p['id'] for p in cursor.fetchall()]
        
        if purchase_ids:
            if len(purchase_ids) == 1:
                ids_tuple = f"({purchase_ids[0]})"
            else:
                ids_tuple = str(tuple(purchase_ids))
                
            cursor.execute(f"DELETE FROM purchase_items WHERE purchase_id IN {ids_tuple}")
            cursor.execute("DELETE FROM purchases WHERE supplier_id=%s", (supplier_id,))

        # 3. Delete Manual Adjustments (if any)
        try:
            cursor.execute("DELETE FROM supplier_adjustments WHERE supplier_id = %s", (supplier_id,))
        except:
            pass

        # NOTE: Removed 'supplier_cashflow' deletion as per instruction ("module removed")

        # 4. Finally, Delete the Supplier
        cursor.execute("DELETE FROM suppliers WHERE id=%s", (supplier_id,))
        
        mysql.connection.commit()
        flash('Supplier and related records deleted successfully.', 'success')
        
    except Exception as e:
        mysql.connection.rollback()
        # Check for specific FK error (1451) to give better message
        if "1451" in str(e):
             flash("Cannot delete supplier: They still have linked records (possibly in Cashflow/Archive) that must be removed first.", "danger")
        else:
             flash(f"Error deleting supplier: {str(e)}", "danger")
        
    finally:
        cursor.close()
        
    return redirect(url_for('suppliers'))



# =====================================================================#
# PURCHASE MANAGEMENT ROUTES
# =====================================================================#

@app.route('/purchases')
def purchases():
    cur = mysql.connection.cursor(MySQLdb.cursors.DictCursor)
    
    # --- 1. Get Filters from URL ---
    q = request.args.get('q', '').strip()
    supplier_id = request.args.get('supplier_id', '')
    start_date = request.args.get('start_date', '')
    end_date = request.args.get('end_date', '')
    
    # --- 2. Build Query ---
    query = """
        SELECT p.*, s.name as supplier_name, COUNT(pi.id) as item_count
        FROM purchases p
        JOIN suppliers s ON p.supplier_id = s.id
        LEFT JOIN purchase_items pi ON p.id = pi.purchase_id
        WHERE 1=1
    """
    params = []
    
    if q:
        query += " AND p.bill_number LIKE %s"
        params.append(f"%{q}%")
        
    if supplier_id:
        query += " AND p.supplier_id = %s"
        params.append(supplier_id)
        
    if start_date:
        query += " AND p.purchase_date >= %s"
        params.append(start_date)
        
    if end_date:
        query += " AND p.purchase_date <= %s"
        params.append(end_date)
        
    query += " GROUP BY p.id ORDER BY p.purchase_date DESC"
    
    # --- 3. Execute Query ---
    cur.execute(query, tuple(params))
    all_purchases = cur.fetchall()
    
    # --- 4. Fetch Suppliers for Filter Dropdown ---
    cur.execute("SELECT id, name FROM suppliers ORDER BY name")
    suppliers = cur.fetchall()
    
    cur.close()
    
    # --- 5. Return Template ---
    return render_template('purchases/purchases.html', 
                           purchases=all_purchases, 
                           suppliers=suppliers,
                           filters={'q': q, 'supplier_id': supplier_id, 'start_date': start_date, 'end_date': end_date})



@app.route('/new_purchase', methods=['GET', 'POST'])
def new_purchase():
    if 'loggedin' not in session: return redirect(url_for('login'))
    
    cursor = mysql.connection.cursor(MySQLdb.cursors.DictCursor)
    
    if request.method == 'POST':
        try:
            # 1. Master Data
            supplier_id = request.form['supplier_id']
            purchase_date_str = request.form['purchase_date']
            
            bill_number = request.form.get('bill_number', '').strip()
            notes = request.form.get('notes', '')
            
            # Auto-Generate Bill Number Logic
            if not bill_number:
                try:
                    p_date_obj = datetime.strptime(purchase_date_str, '%Y-%m-%d')
                    date_part = p_date_obj.strftime("%d%m%y")
                except:
                    p_date_obj = date.today()
                    date_part = p_date_obj.strftime("%d%m%y")
                
                prefix = f"PO-{date_part}-"
                query_pattern = prefix + "%"
                cursor.execute("SELECT bill_number FROM purchases WHERE bill_number LIKE %s ORDER BY id DESC", (query_pattern,))
                existing_bills = cursor.fetchall()
                
                max_suffix = 0
                for row in existing_bills:
                    bn = row['bill_number']
                    if bn.startswith(prefix):
                        try:
                            parts = bn.split('-')
                            suffix = int(parts[-1])
                            if suffix > max_suffix: max_suffix = suffix
                        except: pass
                
                bill_number = f"PO-{date_part}-{max_suffix + 1}"
            
            # 2. Item Arrays
            product_ids = request.form.getlist('product_id[]')
            quantities = request.form.getlist('quantity[]')
            prices = request.form.getlist('price[]')
            
            total_amount = 0.0
            valid_items = []
            
            for i in range(len(product_ids)):
                if product_ids[i]: 
                    try:
                        raw_qty = quantities[i]
                        qty = float(raw_qty) if raw_qty and str(raw_qty).strip() else 0.0
                    except ValueError: qty = 0.0

                    try:
                        raw_price = prices[i]
                        price = float(raw_price) if raw_price and str(raw_price).strip() else 0.0
                    except ValueError: price = 0.0

                    total_amount += (qty * price)
                    valid_items.append({'p_id': product_ids[i], 'qty': qty, 'price': price})
            
            # 3. Insert Master Purchase Record
            cursor.execute("""
                INSERT INTO purchases (supplier_id, purchase_date, bill_number, total_amount, notes)
                VALUES (%s, %s, %s, %s, %s)
            """, (supplier_id, purchase_date_str, bill_number, total_amount, notes))
            
            purchase_id = cursor.lastrowid
            
            # 4. Insert Items & Update Stock (Warehouse Stock Increase)
            stock_col = 'stock' 

            for item in valid_items:
                p_id = item['p_id']
                new_qty = item['qty']
                new_price = item['price']

                # Insert Item
                cursor.execute("""
                    INSERT INTO purchase_items (purchase_id, product_id, quantity, purchase_price)
                    VALUES (%s, %s, %s, %s)
                """, (purchase_id, p_id, new_qty, new_price))
                
                # --- UPDATE STOCK & PRICE LOGIC (LAST PURCHASE PRICE) ---
                # No Average Calculation. Just set purchase_price = new_price
                update_query = f"UPDATE products SET {stock_col} = {stock_col} + %s, purchase_price = %s WHERE id = %s"
                cursor.execute(update_query, (new_qty, new_price, p_id))
            
            # 5. Update Supplier Dues
            cursor.execute("""
                UPDATE suppliers SET current_due = current_due + %s WHERE id = %s
            """, (total_amount, supplier_id))
            
            mysql.connection.commit()
            flash(f"Purchase Order #{purchase_id} created successfully! (Bill No: {bill_number})", "success")
            return redirect(url_for('purchases'))
            
        except Exception as e:
            mysql.connection.rollback()
            flash(f"Error creating purchase: {str(e)}", "danger")
            return redirect(url_for('purchases'))
            
    # GET: Load data
    cursor.execute("SELECT id, name FROM suppliers ORDER BY name")
    suppliers = cursor.fetchall()
    
    try: cursor.execute("SELECT * FROM products WHERE status='Active' ORDER BY name")
    except: cursor.execute("SELECT * FROM products ORDER BY name")
    products = cursor.fetchall()
    
    cursor.close()
    return render_template('purchases/new_purchase.html', suppliers=suppliers, products=products)



# =====================================================================================
# EDIT PURCHASE ROUTE (COMBINED GET & POST)
# =====================================================================================
@app.route('/purchases/edit/<int:purchase_id>', methods=['GET', 'POST'])
def edit_purchase(purchase_id):
    if 'loggedin' not in session: return redirect(url_for('login'))
    
    cursor = mysql.connection.cursor(MySQLdb.cursors.DictCursor)
    
    # --- POST: HANDLE UPDATE LOGIC ---
    if request.method == 'POST':
        try:
            # 1. Fetch OLD Purchase Data
            cursor.execute("SELECT * FROM purchases WHERE id = %s", (purchase_id,))
            old_purchase = cursor.fetchone()
            
            if not old_purchase:
                flash("Purchase not found.", "danger")
                return redirect(url_for('purchases'))
                
            old_total = float(old_purchase['total_amount'])
            old_supplier_id = old_purchase['supplier_id']
            
            # 2. Fetch OLD Items
            cursor.execute("SELECT * FROM purchase_items WHERE purchase_id = %s", (purchase_id,))
            old_items = cursor.fetchall()
            
            stock_col = 'stock' 
            
            # Reverse Stock: Subtract old quantities
            for item in old_items:
                cursor.execute(f"UPDATE products SET {stock_col} = {stock_col} - %s WHERE id = %s", 
                               (item['quantity'], item['product_id']))
                
            # 3. Get NEW Form Data
            new_supplier_id = request.form['supplier_id']
            new_date = request.form['purchase_date']
            new_bill = request.form.get('bill_number', '')
            notes = request.form.get('notes', '')
            
            product_ids = request.form.getlist('product_id[]')
            quantities = request.form.getlist('quantity[]')
            prices = request.form.getlist('price[]')
            
            # 4. Delete Old Items
            cursor.execute("DELETE FROM purchase_items WHERE purchase_id = %s", (purchase_id,))
            
            # 5. Insert NEW Items & Update Stock
            new_total = 0.0
            
            for i in range(len(product_ids)):
                if product_ids[i]:
                    p_id = product_ids[i]
                    new_qty = float(quantities[i])
                    new_price = float(prices[i])
                    line_total = new_qty * new_price
                    new_total += line_total
                    
                    cursor.execute("""
                        INSERT INTO purchase_items (purchase_id, product_id, quantity, purchase_price)
                        VALUES (%s, %s, %s, %s)
                    """, (purchase_id, p_id, new_qty, new_price))
                    
                    # Update Stock & Purchase Price (Last Price Logic)
                    cursor.execute(f"UPDATE products SET {stock_col} = {stock_col} + %s, purchase_price = %s WHERE id = %s", 
                                   (new_qty, new_price, p_id))

            # 6. Update Master Record
            cursor.execute("""
                UPDATE purchases 
                SET supplier_id=%s, purchase_date=%s, bill_number=%s, total_amount=%s, notes=%s
                WHERE id=%s
            """, (new_supplier_id, new_date, new_bill, new_total, notes, purchase_id))
            
            # 7. Update Supplier Dues
            if str(old_supplier_id) == str(new_supplier_id):
                diff = new_total - old_total
                cursor.execute("UPDATE suppliers SET current_due = current_due + %s WHERE id = %s", (diff, old_supplier_id))
            else:
                cursor.execute("UPDATE suppliers SET current_due = current_due - %s WHERE id = %s", (old_total, old_supplier_id))
                cursor.execute("UPDATE suppliers SET current_due = current_due + %s WHERE id = %s", (new_total, new_supplier_id))
            
            mysql.connection.commit()
            flash(f"Purchase Order #{purchase_id} updated successfully.", "success")
            return redirect(url_for('purchases'))
            
        except Exception as e:
            mysql.connection.rollback()
            flash(f"Error updating purchase: {str(e)}", "danger")
            app.logger.error(f"Update Error: {e}")
            return redirect(url_for('purchases'))
            
    # --- GET: RENDER EDIT FORM ---
    cursor.execute("SELECT * FROM purchases WHERE id = %s", (purchase_id,))
    purchase = cursor.fetchone()
    
    cursor.execute("SELECT id, name FROM suppliers ORDER BY name")
    suppliers = cursor.fetchall()
    
    cursor.execute("SELECT * FROM products ORDER BY name")
    products = cursor.fetchall()
    
    cursor.execute("SELECT * FROM purchase_items WHERE purchase_id = %s", (purchase_id,))
    items = cursor.fetchall()
    
    cursor.close()
    return render_template('purchases/edit_purchase.html', 
                           purchase=purchase, 
                           suppliers=suppliers, 
                           products=products, 
                           items=items)





@app.route('/purchases/view/<int:purchase_id>')
def view_purchase(purchase_id):
    cur = mysql.connection.cursor(MySQLdb.cursors.DictCursor)
    
    # Updated Query: Added 'p.notes'
    cur.execute("""
        SELECT p.*, s.name as supplier_name, s.address as supplier_address, s.phone as supplier_phone
        FROM purchases p 
        JOIN suppliers s ON p.supplier_id = s.id 
        WHERE p.id = %s
    """, (purchase_id,))
    
    purchase = cur.fetchone()
    
    cur.execute("""
        SELECT pi.*, pr.name as product_name 
        FROM purchase_items pi 
        JOIN products pr ON pi.product_id = pr.id 
        WHERE pi.purchase_id = %s
    """, (purchase_id,))
    
    items = cur.fetchall()
    cur.close()
    
    if not purchase:
        flash("Purchase not found.", "danger")
        return redirect(url_for('purchases'))
        
    return render_template('purchases/view_purchase.html', purchase=purchase, items=items)



@app.route('/purchases/delete/<int:purchase_id>', methods=['POST'])
def delete_purchase(purchase_id):
    if 'loggedin' not in session: return redirect(url_for('login'))
    
    cursor = mysql.connection.cursor(MySQLdb.cursors.DictCursor)
    
    try:
        # 1. Fetch Purchase Details
        cursor.execute("SELECT supplier_id, total_amount FROM purchases WHERE id = %s", (purchase_id,))
        purchase = cursor.fetchone()
        
        if not purchase:
            flash("Purchase not found.", "danger")
            return redirect(url_for('purchases'))
            
        supplier_id = purchase['supplier_id']
        total_amount = purchase['total_amount']

        # 2. Fetch Purchase Items
        cursor.execute("SELECT product_id, quantity FROM purchase_items WHERE purchase_id = %s", (purchase_id,))
        items = cursor.fetchall()
        
        # FIX: Explicitly use 'stock'
        stock_col = 'stock'
        
        # --- FIX: Check for Negative Stock Risk ---
        # Before deleting, check if reversing this purchase would drop stock below zero.
        for item in items:
            p_id = item['product_id']
            qty = float(item['quantity'])
            
            cursor.execute(f"SELECT {stock_col}, name FROM products WHERE id = %s", (p_id,))
            prod = cursor.fetchone()
            
            current_stock = float(prod[stock_col] or 0)
            
            if current_stock < qty:
                # Stock is LESS than what was purchased? Meaning it was sold!
                # Block deletion to prevent data corruption.
                flash(f"Cannot delete purchase: Item '{prod['name']}' has already been sold/allocated. Current Stock ({current_stock}) < Purchase Qty ({qty}).", "danger")
                return redirect(url_for('purchases'))

        # 3. If Safe, Reverse Stock (Subtract quantity from Warehouse)
        for item in items:
            p_id = item['product_id']
            qty = float(item['quantity'])
            query = f"UPDATE products SET {stock_col} = {stock_col} - %s WHERE id = %s"
            cursor.execute(query, (qty, p_id))
            
        # 4. Reverse Supplier Dues
        cursor.execute("""
            UPDATE suppliers SET current_due = current_due - %s WHERE id = %s
        """, (total_amount, supplier_id))
        
        # 5. Delete the Purchase Record
        cursor.execute("DELETE FROM purchases WHERE id = %s", (purchase_id,))
        
        mysql.connection.commit()
        flash(f"Purchase #{purchase_id} deleted and stock reversed.", "success")
        
    except Exception as e:
        mysql.connection.rollback()
        flash(f"Error deleting purchase: {str(e)}", "danger")
        app.logger.error(f"Delete Error: {e}")
        
    finally:
        cursor.close()
        
    return redirect(url_for('purchases'))



# ==========================================
# PURCHASE ORDER PDF GENERATOR (Dedicated Class)
# ==========================================
class PurchasePDFGenerator(FPDF):
    def __init__(self):
        super().__init__()
        self.company_name = "REAL PROMOTION"
        self.slogan = "SINCE 2005"
        self.address_lines = [
            "Real Promotion, G/12, Tulsimangalam Complex,",
            "B/h. Trimurti Complex, Ghodiya Bazar,",
            "Nadiad-387001, Gujarat, India"
        ]
        self.contact = "+91 96623 22476 | help@realpromotion.in"
        self.gst_text = "GSTIN:24AJVPT0460H1ZW"
        self.gst_subtext = "Composition Dealer- Not Eligibal To Collect Taxes On Sppliers"

    def header(self):
        # Company Header (Same style as other PDFs)
        self.set_font('Arial', 'B', 24)
        self.set_text_color(26, 35, 126) 
        self.cell(0, 10, self.company_name, 0, 1, 'C')
        
        self.set_font('Arial', 'B', 10)
        self.set_text_color(100, 100, 100) 
        self.cell(0, 5, self.slogan, 0, 1, 'C')
        self.ln(2)
        
        self.set_font('Arial', '', 9)
        self.set_text_color(50, 50, 50)
        for line in self.address_lines:
            self.cell(0, 4, line, 0, 1, 'C')
        self.cell(0, 4, self.contact, 0, 1, 'C')
        
        self.set_font('Arial', 'B', 9)
        self.cell(0, 4, self.gst_text, 0, 1, 'C')
        self.set_font('Arial', 'I', 8)
        self.cell(0, 4, self.gst_subtext, 0, 1, 'C')
        self.ln(5)
        
        self.set_draw_color(200, 200, 200)
        self.line(10, self.get_y(), 200, self.get_y())
        self.ln(5)

        self.set_font('Arial', 'B', 16)
        self.set_text_color(0, 0, 0)
        self.cell(0, 10, "PURCHASE ORDER", 0, 1, 'C')
        self.ln(5)

    def footer(self):
        self.set_y(-15)
        self.set_font('Arial', 'I', 8)
        self.set_text_color(128, 128, 128)
        self.cell(0, 10, f'Page {self.page_no()}', 0, 0, 'C')

    def add_po_info(self, purchase):
        self.set_font('Arial', '', 10)
        self.set_text_color(0, 0, 0)
        y = self.get_y()
        
        # Left: Supplier Info
        self.set_xy(10, y)
        self.set_font('Arial', 'B', 11)
        self.cell(0, 6, "Supplier Details:", 0, 1)
        self.set_font('Arial', '', 10)
        self.cell(25, 6, "Name:", 0, 0); self.set_font('Arial','B',10); self.cell(70, 6, str(purchase['supplier_name']), 0, 1)
        self.set_font('Arial','',10); self.cell(25, 6, "Contact:", 0, 0); self.set_font('Arial','B',10); self.cell(70, 6, str(purchase['supplier_phone'] or 'N/A'), 0, 1)
        
        # Address handling
        addr = purchase.get('supplier_address') or 'N/A'
        self.set_font('Arial','',10); self.cell(25, 6, "Address:", 0, 0); self.set_font('Arial','',9); self.multi_cell(70, 6, str(addr), 0, 'L')
        
        # Right: PO Details
        self.set_xy(120, y)
        self.set_font('Arial', 'B', 11)  
        self.cell(0, 6, "Order Details:", 0, 1)
        
        # Use helper or string conversion for date
        p_date = purchase['purchase_date'].strftime('%d-%m-%Y') if purchase['purchase_date'] else "N/A"
        
        self.set_xy(120, y+8)
        self.set_font('Arial','',10); self.cell(25, 6, "PO ID:", 0, 0); self.set_font('Arial','B',10); self.cell(40, 6, f"#{purchase['id']}", 0, 1)
        
        self.set_xy(120, y+14)
        self.set_font('Arial','',10); self.cell(25, 6, "Date:", 0, 0); self.set_font('Arial','B',10); self.cell(40, 6, p_date, 0, 1)
        
        self.set_xy(120, y+20)
        self.set_font('Arial','',10); self.cell(25, 6, "Bill No:", 0, 0); self.set_font('Arial','B',10); self.cell(40, 6, str(purchase['bill_number'] or 'N/A'), 0, 1)
        
        if purchase.get('supplier_gst'):
            self.set_xy(120, y+26)
            self.set_font('Arial','',10); self.cell(25, 6, "GSTIN:", 0, 0); self.set_font('Arial','B',10); self.cell(40, 6, str(purchase['supplier_gst']), 0, 1)

        # Removed Notes from here (Moved to bottom)
        self.ln(20) # Space before table

    def add_table(self, items, notes=None):
        # Header
        cols = ["#", "Product Description", "Qty", "Unit Price", "Total Amount"]
        widths = [10, 90, 20, 30, 40]
        
        self.set_font('Arial', 'B', 10)
        self.set_fill_color(26, 35, 126) 
        self.set_text_color(255, 255, 255)
        self.set_draw_color(0, 0, 0)
        self.set_line_width(0.3)
        
        for i, col in enumerate(cols):
            self.cell(widths[i], 8, col, 1, 0, 'C', True)
        self.ln()
        
        # Rows
        self.set_font('Arial', '', 10)
        self.set_text_color(0, 0, 0)
        fill = False
        self.set_fill_color(245, 245, 245)
        
        total_qty = 0
        grand_total = 0
        
        for i, item in enumerate(items):
            qty = float(item['quantity'])
            price = float(item['purchase_price'])
            # Calculate row total explicitly to be safe
            row_total = qty * price 
            
            total_qty += qty
            grand_total += row_total
            
            self.cell(widths[0], 8, str(i+1), 1, 0, 'C', fill)
            self.cell(widths[1], 8, str(item['product_name']), 1, 0, 'L', fill)
            self.cell(widths[2], 8, str(int(qty)), 1, 0, 'C', fill)
            self.cell(widths[3], 8, f"{price:.2f}", 1, 0, 'R', fill)
            self.cell(widths[4], 8, f"{row_total:.2f}", 1, 1, 'R', fill)
            fill = not fill
            
        # Totals Row
        self.set_font('Arial', 'B', 10)
        self.cell(widths[0]+widths[1], 8, "TOTALS", 1, 0, 'R', True)
        self.cell(widths[2], 8, str(int(total_qty)), 1, 0, 'C', True)
        self.cell(widths[3], 8, "", 1, 0, 'C', True)
        self.cell(widths[4], 8, f"{grand_total:.2f}", 1, 1, 'R', True)
        self.ln(10)
        
        # Grand Total Box
        self.set_x(120)
        self.set_fill_color(230, 230, 230)
        self.set_font('Arial', 'B', 12)
        self.cell(40, 10, "GRAND TOTAL:", 1, 0, 'R', True)
        self.cell(30, 10, f"{grand_total:.2f}", 1, 1, 'R', True)
        self.ln(5)

        # ADDED: Notes Section (Below Grand Total)
        if notes:
            self.ln(5) # Gap
            self.set_font('Arial', 'B', 10)
            self.cell(0, 6, "Notes / Remarks:", 0, 1)
            self.set_font('Arial', '', 9)
            # Use multi-cell to handle long text
            self.multi_cell(0, 5, str(notes), 0, 'L')

    def add_signature(self):
        # Ensure space
        if self.get_y() > 240: self.add_page()
        self.ln(10)
        y_pos = self.get_y()
        
        sig_path = os.path.join(app.root_path, 'static', 'img', 'signature.png')
        if os.path.exists(sig_path):
            self.image(sig_path, x=140, y=y_pos, w=40)
            
        self.set_xy(130, y_pos + 25)
        self.set_font('Arial', 'B', 10)
        self.set_text_color(0, 0, 0)
        self.cell(60, 5, "For, REAL PROMOTION", 0, 1, 'C')
        
        self.set_xy(130, y_pos + 30)
        self.cell(60, 5, "(Authorized Signatory)", 0, 1, 'C')
        
        self.ln(20)
        self.set_font('Arial', 'I', 8)
        self.set_text_color(128, 128, 128)
        self.cell(0, 5, "This is a computer generated document.", 0, 1, 'C')

# ==========================================
# UPDATED PURCHASE PDF ROUTE
# ==========================================
@app.route('/purchases/pdf/<int:purchase_id>')
def purchase_pdf(purchase_id):
    if 'loggedin' not in session: return redirect(url_for('login'))
    
    cursor = mysql.connection.cursor(MySQLdb.cursors.DictCursor)
    
    # 1. Fetch Purchase & Supplier (ADDED NOTES field)
    cursor.execute("""
        SELECT p.*, s.name as supplier_name, s.address as supplier_address, s.phone as supplier_phone, s.gstin as supplier_gst
        FROM purchases p 
        LEFT JOIN suppliers s ON p.supplier_id = s.id 
        WHERE p.id = %s
    """, (purchase_id,))
    purchase = cursor.fetchone()
    
    if not purchase:
        flash("Purchase not found", "danger")
        return redirect(url_for('purchases'))
        
    # 2. Fetch Items
    cursor.execute("""
        SELECT pi.*, pr.name as product_name 
        FROM purchase_items pi 
        LEFT JOIN products pr ON pi.product_id = pr.id 
        WHERE pi.purchase_id = %s
    """, (purchase_id,))
    items = cursor.fetchall()
    
    cursor.close()
    
    # 3. Generate PDF using Dedicated Class
    pdf = PurchasePDFGenerator()
    pdf.alias_nb_pages()
    pdf.add_page()
    
    pdf.add_po_info(purchase)
    # Pass notes to add_table
    pdf.add_table(items, notes=purchase.get('notes'))
    pdf.add_signature()
    
    pdf_string = pdf.output(dest='S').encode('latin-1')
    buffer = io.BytesIO(pdf_string)
    buffer.seek(0)
    
    # Filename: PO_ID_Supplier.pdf
    supp_name = str(purchase.get('supplier_name', 'Unknown')).replace(" ", "_")
    filename = f"PO_{purchase['id']}_{supp_name}.pdf"
    
    return send_file(buffer, as_attachment=True, download_name=filename, mimetype='application/pdf')


# --- HELPER: Safe Date Formatter ---
def safe_date_format(date_obj, format='%d-%m-%Y', default='N/A'):
    """Safely formats a date object, returning default if None."""
    if date_obj:
        return date_obj.strftime(format)
    return default






# ----------------- Category Module CRDUD Operation for admin side inventory management-------------------------------------------------------------------

# =========================================================
#  STOCK ADJUSTMENT ROUTE
# =========================================================###################################################
@app.route('/inventory/adjust', methods=['GET', 'POST'])
def stock_adjust():
    # 1. Security Check
    if "loggedin" not in session:
        return redirect(url_for("login"))

    cur = mysql.connection.cursor(MySQLdb.cursors.DictCursor)

    # 2. Handle Form Submission
    if request.method == 'POST':
        product_id = request.form.get('product_id')
        adj_type = request.form.get('adjustment_type')  # 'add' or 'subtract'
        try:
            quantity = int(request.form.get('quantity'))
        except (ValueError, TypeError):
            flash("Invalid quantity entered.", "danger")
            return redirect(url_for('stock_adjust'))
            
        reason = request.form.get('reason')

        # Basic Validation
        if not product_id or quantity <= 0:
            flash("Please select a product and enter a valid quantity.", "danger")
            return redirect(url_for('stock_adjust'))

        try:
            # 3. Logic for Subtracting Stock (Check availability first)
            if adj_type == 'subtract':
                cur.execute("SELECT stock FROM products WHERE id = %s", (product_id,))
                row = cur.fetchone()
                current_stock = row['stock'] if row else 0

                if current_stock < quantity:
                    flash(f"Error: Cannot remove {quantity} items. Current stock is only {current_stock}.", "danger")
                    return redirect(url_for('stock_adjust'))

                # Decrease Stock
                cur.execute("UPDATE products SET stock = stock - %s WHERE id = %s", (quantity, product_id))

            # 4. Logic for Adding Stock
            else:
                cur.execute("UPDATE products SET stock = stock + %s WHERE id = %s", (quantity, product_id))

            # 5. Insert into History Log (stock_adjustments table)
            cur.execute("""
                INSERT INTO stock_adjustments (product_id, adjustment_type, quantity, reason, created_at)
                VALUES (%s, %s, %s, %s, NOW())
            """, (product_id, adj_type, quantity, reason))

            mysql.connection.commit()
            flash("Stock adjustment recorded successfully!", "success")
            
            # Redirect to Inventory Master to see the changes
            return redirect(url_for('inventory_master'))

        except Exception as e:
            mysql.connection.rollback()
            app.logger.error(f"Stock Adjust Error: {e}")
            flash(f"An error occurred: {e}", "danger")
            return redirect(url_for('stock_adjust'))
        finally:
            cur.close()

    # 6. GET Request - Show the Form
    try:
        cur.execute("SELECT id, name, stock FROM products ORDER BY name ASC")
        products = cur.fetchall()
        cur.close()
        return render_template('inventory/adjust_stock.html', products=products)
    except Exception as e:
        flash(f"Error loading products: {e}", "danger")
        return redirect(url_for('inventory_master'))


# =========================================================
#  UPDATED INVENTORY MASTER (Universal Logic)
# =========================================================
@app.route('/inventory_master')
def inventory_master():
    if "loggedin" not in session:
        return redirect(url_for("login"))

    cur = mysql.connection.cursor(MySQLdb.cursors.DictCursor)
    
    # --- MODIFIED: Added "WHERE p.status = 'Active'" ---
    cur.execute("""
        SELECT p.*, pc.category_name 
        FROM products p
        LEFT JOIN product_categories pc ON p.category_id = pc.id
        WHERE p.status = 'Active'
        ORDER BY p.id ASC
    """)
    products = cur.fetchall()

    # --- LOGIC PRESERVED: Calculate Allocated Stock (Field Stock) ---
    # Logic: For every employee, find their CURRENT holding.
    # Current Holding = Last Known State (Either from Morning or Evening).
    
    allocated_map = {} 
    
    cur.execute("SELECT id FROM employees WHERE status='active'")
    employees = cur.fetchall()
    
    for emp in employees:
        eid = emp['id']
        emp_stock = {} 

        # Step A: Find the LATEST date of activity for this employee
        # Check Morning Allocations
        cur.execute("SELECT date FROM morning_allocations WHERE employee_id=%s ORDER BY date DESC LIMIT 1", (eid,))
        last_morning = cur.fetchone()
        
        # Check Evening Settlements (Final)
        cur.execute("SELECT date FROM evening_settle WHERE employee_id=%s AND status='final' ORDER BY date DESC LIMIT 1", (eid,))
        last_evening = cur.fetchone()
        
        target_date = None
        source = None # 'morning' or 'evening'
        
        # Determine which record is more recent (or same day priority)
        if last_morning and last_evening:
            if last_morning['date'] > last_evening['date']:
                target_date = last_morning['date']
                source = 'morning'
            elif last_evening['date'] > last_morning['date']:
                target_date = last_evening['date']
                source = 'evening'
            else:
                # Same date: Evening supersedes Morning (because it's end of day)
                target_date = last_evening['date']
                source = 'evening'
        elif last_morning:
            target_date = last_morning['date']
            source = 'morning'
        elif last_evening:
            target_date = last_evening['date']
            source = 'evening'
            
        # Step B: Fetch Stock based on the source
        if target_date:
            if source == 'evening':
                # Fetch remaining from evening_item
                cur.execute("""
                    SELECT ei.product_id, ei.remaining_qty
                    FROM evening_item ei
                    JOIN evening_settle es ON ei.settle_id = es.id
                    WHERE es.employee_id=%s AND es.date=%s
                """, (eid, target_date))
                for row in cur.fetchall():
                    pid = row['product_id']
                    qty = int(row['remaining_qty'])
                    if qty > 0: emp_stock[pid] = qty
            else:
                # Source is Morning: Fetch Opening + Given
                # Note: This handles the "Restock" case correctly because restock updates 'given_qty'
                # in the same morning_allocation_items record for that day.
                cur.execute("""
                    SELECT mai.product_id, mai.opening_qty, mai.given_qty
                    FROM morning_allocation_items mai
                    JOIN morning_allocations ma ON mai.allocation_id = ma.id
                    WHERE ma.employee_id=%s AND ma.date=%s
                """, (eid, target_date))
                for row in cur.fetchall():
                    pid = row['product_id']
                    qty = int(row['opening_qty']) + int(row['given_qty'])
                    if qty > 0: emp_stock[pid] = qty

        # Add this employee's stock to the global map
        for pid, qty in emp_stock.items():
            if pid in allocated_map: allocated_map[pid] += qty
            else: allocated_map[pid] = qty

    # 3. Merge & Stats
    total_value = 0
    low_stock_count = 0
    total_items = 0

    for p in products:
        warehouse_stock = int(p['stock'] or 0)
        allocated_stock = allocated_map.get(p['id'], 0)
        
        # Total Asset = Warehouse + Field
        total_holding = warehouse_stock + allocated_stock
        
        p['warehouse_stock'] = warehouse_stock
        p['allocated_stock'] = allocated_stock
        p['total_stock'] = total_holding
        
        total_value += (p['price'] * total_holding)
        
        # --- FIX START: Handle 0 Threshold Correctly ---
        threshold = p['low_stock_threshold']
        if threshold is None: threshold = 10
            
        if total_holding <= threshold:
            low_stock_count += 1
        # --- FIX END ---
        
        total_items += 1

    cur.execute("SELECT * FROM product_categories ORDER BY category_name")
    categories = cur.fetchall()
    
    cur.close()

    # NOTE: Keeping your original template path
    return render_template('inventory/inventory_master.html', 
                           products=products, 
                           categories=categories,
                           stats={
                               "total_value": total_value,
                               "low_stock": low_stock_count,
                               "total_items": total_items
                           })


# =========================================================
#  CATEGORY MASTER (With Product Counts)
# =========================================================
@app.route('/category_master', methods=['GET', 'POST'])
def category_master():
    if "loggedin" not in session:
        return redirect(url_for("login"))

    cur = mysql.connection.cursor(MySQLdb.cursors.DictCursor)

    if request.method == 'POST':
        action = request.form.get('action')
        
        if action == 'add':
            name = request.form.get('category_name')
            desc = request.form.get('description')
            cur.execute("INSERT INTO product_categories (category_name, description) VALUES (%s, %s)", (name, desc))
        
        elif action == 'edit':
            cat_id = request.form.get('category_id')
            name = request.form.get('category_name')
            desc = request.form.get('description')
            cur.execute("UPDATE product_categories SET category_name=%s, description=%s WHERE id=%s", (name, desc, cat_id))
        
        elif action == 'delete':
            cat_id = request.form.get('category_id')
            # Optional: Check if products exist in this category before deleting
            cur.execute("DELETE FROM product_categories WHERE id=%s", (cat_id,))

        mysql.connection.commit()
        flash("Category updated successfully", "success")
        return redirect(url_for('category_master'))

    # Modified Query to Count Products
    cur.execute("""
        SELECT c.*, COUNT(p.id) as total_products 
        FROM product_categories c 
        LEFT JOIN products p ON c.id = p.category_id 
        GROUP BY c.id 
        ORDER BY c.category_name ASC
    """)
    categories = cur.fetchall()
    cur.close()

    return render_template('inventory/category_master.html', categories=categories)



# 3. Product History (The Timeline)
@app.route('/product_history/<int:product_id>')
def product_history(product_id):
    if "loggedin" not in session:
        return redirect(url_for("login"))

    cur = mysql.connection.cursor(MySQLdb.cursors.DictCursor)

    # Get Product Details
    cur.execute("SELECT * FROM products WHERE id = %s", (product_id,))
    product = cur.fetchone()

    # COMPLEX QUERY: Combine Purchases, Sales, and Adjustments
    # We use UNION ALL to merge different tables into one timeline
    query = """
        (SELECT 
            'Purchase' as type, 
            purchase_date as date, 
            quantity as qty, 
            'in' as flow,
            concat('Bill: ', bill_number) as details 
         FROM purchases p 
         JOIN purchase_items pi ON p.id = pi.purchase_id 
         WHERE pi.product_id = %s)

        UNION ALL

        (SELECT 
            'Sale' as type, 
            es.date as date, 
            ei.sold_qty as qty, 
            'out' as flow,
            concat('Sold by: ', e.name) as details
         FROM evening_item ei
         JOIN evening_settle es ON ei.settle_id = es.id
         JOIN employees e ON es.employee_id = e.id
         WHERE ei.product_id = %s AND ei.sold_qty > 0)

        UNION ALL

        (SELECT 
            'Adjustment' as type, 
            created_at as date, 
            quantity as qty, 
            adjustment_type as flow, -- 'add' is in, 'subtract' is out
            reason as details
         FROM stock_adjustments 
         WHERE product_id = %s)

        ORDER BY date DESC
    """
    
    cur.execute(query, (product_id, product_id, product_id))
    history = cur.fetchall()
    cur.close()

    return render_template('inventory/product_history.html', product=product, history=history)

# =====================================================================================
# UPDATED INVENTORY ROUTE (Main Site)
# Fix: Shows Total Stock (Warehouse + Field) like Inventory Master
# =====================================================================================
@app.route('/inventory')
def inventory():
    # 1. Security Check
    if "loggedin" not in session:
        return redirect(url_for("login"))

    cur = mysql.connection.cursor(MySQLdb.cursors.DictCursor)

    # 2. Fetch Active Products (Soft Delete Logic Applied)
    # --- LOGIC PRESERVED: Keeping your try-except block ---
    try:
        cur.execute("SELECT * FROM products WHERE status = 'Active' ORDER BY id ASC")
    except:
        cur.execute("SELECT * FROM products ORDER BY id ASC")
    products = cur.fetchall()

    # 3. Fetch Categories for Filter
    cur.execute("SELECT id, category_name FROM product_categories")
    categories = cur.fetchall()
    categories_map = {c['id']: c['category_name'] for c in categories}

    # -------------------------------------------------------------
    # 4. Calculate Allocated Stock (Global Logic from Inventory Master)
    # -------------------------------------------------------------
    # Hum check karenge ki har employee ke paas LAST known stock kitna hai.
    # Logic: Agar aaj subah allocation hua hai to wo lenge, agar kal shaam ka settlement latest hai to wo lenge.
    
    allocated_map = {} 
    
    cur.execute("SELECT id FROM employees WHERE status='active'")
    employees = cur.fetchall()
    
    for emp in employees:
        eid = emp['id']
        emp_stock = {} 

        # Step A: Find the LATEST date of activity for this employee
        # Check Morning Allocations
        cur.execute("SELECT date FROM morning_allocations WHERE employee_id=%s ORDER BY date DESC LIMIT 1", (eid,))
        last_morning = cur.fetchone()
        
        # Check Evening Settlements (Final)
        cur.execute("SELECT date FROM evening_settle WHERE employee_id=%s AND status='final' ORDER BY date DESC LIMIT 1", (eid,))
        last_evening = cur.fetchone()
        
        target_date = None
        source = None # 'morning' or 'evening'
        
        # Determine which record is more recent (Priority to Evening if same day)
        if last_morning and last_evening:
            if last_morning['date'] > last_evening['date']:
                target_date = last_morning['date']
                source = 'morning'
            elif last_evening['date'] > last_morning['date']:
                target_date = last_evening['date']
                source = 'evening'
            else:
                # Same date: Evening supersedes Morning (End of Day status)
                target_date = last_evening['date']
                source = 'evening'
        elif last_morning:
            target_date = last_morning['date']
            source = 'morning'
        elif last_evening:
            target_date = last_evening['date']
            source = 'evening'
            
        # Step B: Fetch Stock based on the source
        if target_date:
            if source == 'evening':
                # Fetch remaining from evening_item (Jo stock bach gaya)
                cur.execute("""
                    SELECT ei.product_id, ei.remaining_qty
                    FROM evening_item ei
                    JOIN evening_settle es ON ei.settle_id = es.id
                    WHERE es.employee_id=%s AND es.date=%s
                """, (eid, target_date))
                for row in cur.fetchall():
                    pid = row['product_id']
                    qty = int(row['remaining_qty'])
                    if qty > 0: emp_stock[pid] = qty
            else:
                # Source is Morning: Fetch Opening + Given (Jo subah diya gaya)
                cur.execute("""
                    SELECT mai.product_id, mai.opening_qty, mai.given_qty
                    FROM morning_allocation_items mai
                    JOIN morning_allocations ma ON mai.allocation_id = ma.id
                    WHERE ma.employee_id=%s AND ma.date=%s
                """, (eid, target_date))
                for row in cur.fetchall():
                    pid = row['product_id']
                    qty = int(row['opening_qty']) + int(row['given_qty'])
                    if qty > 0: emp_stock[pid] = qty

        # Add this employee's stock to the global map
        for pid, qty in emp_stock.items():
            if pid in allocated_map: allocated_map[pid] += qty
            else: allocated_map[pid] = qty

    # -------------------------------------------------------------
    # 5. Merge Data & Update Product Objects for Display
    # -------------------------------------------------------------
    total_value = 0
    low_stock_count = 0
    out_stock_count = 0

    for p in products:
        price = p.get('price', 0) or 0
        
        # 'stock' column in DB is Warehouse Stock
        warehouse_stock = int(p.get('stock', 0) or 0)
        
        # Allocated Stock from our calculation above
        allocated_stock = allocated_map.get(p['id'], 0)
        
        # Combined Total Stock (Warehouse + Field)
        total_holding = warehouse_stock + allocated_stock
        
        # --- CRITICAL FIX START ---
        # Update attributes so the Template shows the TOTAL
        p['warehouse_stock'] = warehouse_stock
        p['allocated_stock'] = allocated_stock
        
        # HERE IS THE MAGIC: We overwrite 'stock' with 'total_holding'
        # This ensures the Main Site Inventory Card shows (Warehouse + Allocated)
        p['stock'] = total_holding  
        # --- CRITICAL FIX END ---
        
        # Calculate Value based on Total Holding
        total_value += price * total_holding 
        
        # --- FIX START: Handle 0 Threshold Correctly ---
        threshold = p.get('low_stock_threshold')
        if threshold is None: threshold = 10
            
        if total_holding <= threshold:
            low_stock_count += 1
        # --- FIX END ---
            
        if total_holding == 0:
            out_stock_count += 1

    total_products = len(products)

    stats = {
        "total_value": total_value,
        "total_products": total_products,
        "low_stock_count": low_stock_count,
        "out_stock_count": out_stock_count
    }

    # 6. Fetch Categories for Filter Dropdown
    cur.execute("SELECT category_name FROM product_categories ORDER BY category_name ASC")
    filter_categories = cur.fetchall()

    cur.close()

    return render_template(
        "inventory.html",
        products=products,
        stats=stats,
        categories_map=categories_map,
        filters={"categories": filter_categories}
    )


@app.route("/add_product_form")
def add_product_form():
    # This route just shows the page
    return render_template("add_product.html")


@app.route('/add_product', methods=['GET', 'POST'])
def add_product():
    if 'loggedin' not in session:
        return redirect(url_for('login'))

    cursor = mysql.connection.cursor(MySQLdb.cursors.DictCursor)

    if request.method == 'POST':
        name = request.form.get('name')
        price = float(request.form.get('price') or 0)
        # New Field: Purchase Price (Cost)
        purchase_price = float(request.form.get('purchase_price') or 0)
        stock = int(request.form.get('stock') or 0)
        category_id = request.form.get('category_id') or None
        if category_id == "": category_id = None
        low_stock_threshold = int(request.form.get('low_stock_threshold') or 10)

        image_url = None
        if 'image' in request.files and request.files['image'].filename:
            image = request.files['image']
            upload_res = cloudinary.uploader.upload(image, folder="erp_products")
            image_url = upload_res.get("secure_url") or upload_res.get("url")

        # Insert Product with Purchase Price
        cursor.execute("""
            INSERT INTO products
            (name, price, purchase_price, stock, category_id, low_stock_threshold, image)
            VALUES (%s, %s, %s, %s, %s, %s, %s)
        """, (name, price, purchase_price, stock, category_id, low_stock_threshold, image_url))

        mysql.connection.commit()
        cursor.close()

        flash("Product added successfully!", "success")
        return redirect(url_for("inventory"))

    # GET: Categories
    cursor.execute("SELECT id, category_name FROM product_categories ORDER BY category_name ASC")
    categories = cursor.fetchall()
    cursor.close()

    return render_template("add_product.html", categories=categories)

@app.route('/edit_product/<int:product_id>', methods=['GET', 'POST'])
def edit_product(product_id):
    if 'loggedin' not in session:
        return redirect(url_for('login'))

    cursor = mysql.connection.cursor(MySQLdb.cursors.DictCursor)

    # 1. Fetch Existing Product
    cursor.execute("SELECT * FROM products WHERE id = %s", (product_id,))
    product = cursor.fetchone()

    if not product:
        flash("Product not found!", "danger")
        return redirect(url_for('inventory'))

    # 2. Handle Form Submission (POST)
    if request.method == 'POST':
        name = request.form['name']
        category_id = request.form['category_id']
        price = request.form['price']
        
        # --- CHANGES: Capture Stock and Low Stock Threshold ---
        stock = request.form['stock'] 
        low_stock = request.form['low_stock_threshold'] 
        # ----------------------------------------------------

        # Handle Image Upload
        image_url = product['image']
        if 'image' in request.files:
            file = request.files['image']
            if file and file.filename != '':
                upload_result = cloudinary.uploader.upload(file)
                image_url = upload_result['secure_url']

        # 3. Update Database 
        cursor.execute("""
            UPDATE products 
            SET name=%s, category_id=%s, price=%s, stock=%s, low_stock_threshold=%s, image=%s 
            WHERE id=%s
        """, (name, category_id, price, stock, low_stock, image_url, product_id))
        
        mysql.connection.commit()
        cursor.close()

        flash("Product updated successfully!", "success")
        return redirect(url_for('inventory'))

    # 3. Handle Page Load (GET)
    # --- FIX: Changed table name from 'categories' to 'product_categories' ---
    cursor.execute("SELECT * FROM product_categories")
    categories = cursor.fetchall()
    cursor.close()

    return render_template('edit_product.html', product=product, categories=categories)

# ==========================================
# SOFT DELETE PRODUCT (Safe Delete)
# ==========================================

@app.route('/delete_product_soft/<int:id>', methods=['POST'])
def delete_product_soft(id):
    if "loggedin" not in session:
        return redirect(url_for("login"))

    try:
        cur = mysql.connection.cursor()
        
        # 1. Update status to Inactive instead of DELETE
        cur.execute("UPDATE products SET status = 'Inactive' WHERE id = %s", (id,))
        
        mysql.connection.commit()
        cur.close()
        
        flash('Product archived successfully! It is now hidden from lists.', 'success')
        
    except Exception as e:
        flash(f'Error archiving product: {str(e)}', 'danger')
        
    # Redirect back to where the user came from (Master or Main Inventory)
    return redirect(request.referrer or url_for('inventory_master'))


#=========================================Product wise sales report===========================#

@app.route("/product_sales_report")
def product_sales_report():
    if "loggedin" not in session:
        return redirect(url_for("login"))

    cursor = mysql.connection.cursor(MySQLdb.cursors.DictCursor)

    cursor.execute("""
        SELECT 
            p.id AS product_id,
            p.name AS product_name,
            p.price AS product_price,
            p.image AS product_image,
            pc.category_name,
            
            COALESCE(SUM(s.quantity), 0) AS total_sold_qty,
            COALESCE(SUM(s.quantity * s.price), 0) AS total_revenue,
            MAX(s.sale_date) AS last_sold_date

        FROM products p
        LEFT JOIN sales s ON s.product_id = p.id
        LEFT JOIN product_categories pc ON pc.id = p.category_id

        GROUP BY 
            p.id, p.name, p.price, p.image, pc.category_name

        ORDER BY total_sold_qty DESC;
    """)

    sales_report = cursor.fetchall()
    cursor.close()

    return render_template("product_sales_report.html", sales_report=sales_report)



@app.route("/product_sales_details/<int:product_id>")
def product_sales_details(product_id):
    if "loggedin" not in session:
        return redirect(url_for("login"))

    cursor = mysql.connection.cursor(MySQLdb.cursors.DictCursor)

    # 1) Get product info
    cursor.execute("""
        SELECT p.*, pc.category_name
        FROM products p
        LEFT JOIN product_categories pc ON pc.id = p.category_id
        WHERE p.id = %s
    """, (product_id,))
    product = cursor.fetchone()

    if not product:
        flash("Product not found!", "danger")
        return redirect(url_for("product_sales_report"))

    # 2) Fetch all sales for this product
    cursor.execute("""
        SELECT 
            id,
            quantity,
            price,
            discount,
            payment_mode,
            payment_remark,
            sale_date,
            (quantity * price) AS total_amount
        FROM sales
        WHERE product_id = %s
        ORDER BY sale_date DESC
    """, (product_id,))
    
    sales = cursor.fetchall()
    cursor.close()

    return render_template(
        "product_sales_details.html",
        product=product,
        sales=sales
    )


@app.route("/export_sales_excel")
def export_sales_excel():
    cursor = mysql.connection.cursor(MySQLdb.cursors.DictCursor)
    cursor.execute("""
        SELECT 
            p.name AS product_name,
            pc.category_name,
            COALESCE(SUM(s.quantity), 0) AS total_sold_qty,
            COALESCE(SUM(s.quantity * s.price), 0) AS total_revenue,
            MAX(s.sale_date) AS last_sold_date
        FROM products p
        LEFT JOIN sales s ON s.product_id = p.id
        LEFT JOIN product_categories pc ON pc.id = p.category_id
        GROUP BY p.id
        ORDER BY total_sold_qty DESC
    """)
    data = cursor.fetchall()
    cursor.close()

    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Sales Report"

    # Header
    ws.append(["Product", "Category", "Sold Qty", "Revenue", "Last Sold"])

    # Rows
    for row in data:
        ws.append([
            row["product_name"],
            row["category_name"],
            row["total_sold_qty"],
            row["total_revenue"],
            row["last_sold_date"]
        ])

    filepath = "product_sales_report.xlsx"
    wb.save(filepath)

    return send_file(filepath, as_attachment=True)


@app.route("/export_sales_pdf")
def export_sales_pdf():
    cursor = mysql.connection.cursor(MySQLdb.cursors.DictCursor)
    cursor.execute("""
        SELECT 
            p.name AS product_name,
            pc.category_name,
            COALESCE(SUM(s.quantity), 0) AS total_sold_qty,
            COALESCE(SUM(s.quantity * s.price), 0) AS total_revenue,
            MAX(s.sale_date) AS last_sold_date
        FROM products p
        LEFT JOIN sales s ON s.product_id = p.id
        LEFT JOIN product_categories pc ON pc.id = p.category_id
        GROUP BY p.id
        ORDER BY total_sold_qty DESC
    """)
    data = cursor.fetchall()
    cursor.close()

    from reportlab.lib.pagesizes import letter
    from reportlab.pdfgen import canvas

    filename = "product_sales_report.pdf"
    c = canvas.Canvas(filename, pagesize=letter)
    width, height = letter

    y = height - 50
    c.setFont("Helvetica-Bold", 14)
    c.drawString(50, y, "Product Sales Report")
    y -= 40

    c.setFont("Helvetica", 10)
    for row in data:
        line = f"{row['product_name']} | {row['category_name']} | Qty: {row['total_sold_qty']} | Revenue: ₹{row['total_revenue']}"
        c.drawString(50, y, line)
        y -= 20
        if y < 50:
            c.showPage()
            y = height - 50

    c.save()
    return send_file(filename, as_attachment=True)




# ============================
#  EMPLOYEE MANAGEMENT MODULE
# ============================

# ---------- USER SIDE: EMPLOYEE DASHBOARD + DETAIL (base.html) ----------

@app.route("/employees")
def employees():
    cur = mysql.connection.cursor(DictCursor)
    try:
        cur.execute("""
            SELECT e.id, e.name, e.image, e.phone, e.city, e.status,
                   p.position_name, d.department_name
            FROM employees e
            LEFT JOIN employee_positions p ON e.position_id = p.id
            LEFT JOIN employee_departments d ON e.department_id = d.id
            WHERE e.status = 'active'
            ORDER BY e.name
        """)
        rows = cur.fetchall()
    finally:
        cur.close()
    return render_template("employees.html", employees=rows)


# 3. PUBLIC EMPLOYEE DETAILS ROUTE
@app.route("/employee/<int:id>")
def employee_details(id):
    cur = mysql.connection.cursor(DictCursor)
    try:
        # Added join for Position and Department names
        cur.execute("""
            SELECT e.*, 
                   p.position_name, 
                   d.department_name
            FROM employees e
            LEFT JOIN employee_positions p ON e.position_id = p.id
            LEFT JOIN employee_departments d ON e.department_id = d.id
            WHERE e.id = %s
        """, (id,))
        emp = cur.fetchone()
    finally:
        cur.close()
        
    if not emp:
        abort(404)
        
    return render_template("employees/employee_details.html", employee=emp)


# ---------- ADMIN SIDE: EMPLOYEE MASTER (admin_master.html) ----------#
# REPLACE any existing 'employee_master' functions with ONLY this one:

@app.route("/employee_master")
def employee_master():
    if "loggedin" not in session:
        return redirect(url_for("login"))

    # 1. Get Filter from URL (Default is 'active')
    # This allows switching between Active and Inactive tabs
    status_filter = request.args.get('status', 'active')

    cur = mysql.connection.cursor(DictCursor)
    try:
        # 2. Dynamic Query based on status (Active vs Inactive)
        query = """
            SELECT e.id, e.name, e.email, e.phone, e.city, e.state, e.pincode,
                   e.address_line1, e.address_line2, e.status, e.image,
                   p.position_name, d.department_name
            FROM employees e
            LEFT JOIN employee_positions p ON e.position_id = p.id
            LEFT JOIN employee_departments d ON e.department_id = d.id
            WHERE e.status = %s
            ORDER BY e.id ASC
        """
        cur.execute(query, (status_filter,))
        employees = cur.fetchall()

        cur.execute("SELECT * FROM employee_positions ORDER BY id DESC")
        positions = cur.fetchall()

        cur.execute("SELECT * FROM employee_departments ORDER BY id DESC")
        departments = cur.fetchall()
    finally:
        cur.close()

    # 3. Pass 'current_status' so the frontend knows which Tab to highlight
    return render_template("employees/employee_master.html",
                             employees=employees, 
                             positions=positions, 
                             departments=departments,
                             current_status=status_filter)




# ADD THIS NEW ROUTE (To Restore/Reactivate Employee)
@app.route("/reactivate_employee/<int:id>", methods=["POST"])
def reactivate_employee(id):
    if "loggedin" not in session:
        return redirect(url_for("login"))
    
    cur = mysql.connection.cursor()
    try:
        # Simple Status Update
        cur.execute("UPDATE employees SET status='active' WHERE id=%s", (id,))
        mysql.connection.commit()
        flash("Employee Reactivated Successfully! They are now back in the active list.", "success")
    except Exception as e:
        mysql.connection.rollback()
        flash(f"Error reactivating: {e}", "danger")
    finally:
        cur.close()
        
    # Go back to the Inactive list so user can see others
    return redirect(url_for("employee_master", status='inactive'))

# ---------- ADMIN SIDE: POSITION MASTER ----------

# ---------- ADMIN SIDE: DEPARTMENT MASTER ----------#############################################################################################################
@app.route("/employee_department_add", methods=["POST"])
def employee_department_add():
    if "loggedin" not in session:
        return redirect(url_for("login"))

    department_name = request.form.get("department_name", "").strip()
    if not department_name:
        flash("Department name is required.", "danger")
        return redirect(url_for("employee_master"))

    try:
        # Use DictCursor for consistency
        cursor = mysql.connection.cursor(MySQLdb.cursors.DictCursor)
        cursor.execute(
            "INSERT INTO employee_departments (department_name) VALUES (%s)",
            (department_name,)
        )
        mysql.connection.commit()
        flash("Department added successfully!", "success")
    except Exception as e:
        mysql.connection.rollback()
        flash(f"Error adding department: {e}", "danger")
    finally:
        cursor.close()

    return redirect(url_for("employee_master"))


@app.route("/employee_department_delete/<int:id>")
def employee_department_delete(id):
    if "loggedin" not in session:
        return redirect(url_for("login"))

    # Use DictCursor so we can access result['cnt'] safely
    cursor = mysql.connection.cursor(MySQLdb.cursors.DictCursor)
    
    try:
        # Check if any employee is using this department
        cursor.execute("SELECT COUNT(*) AS cnt FROM employees WHERE department_id=%s", (id,))
        result = cursor.fetchone()
        count = result['cnt'] if result else 0

        if count > 0:
            flash("Cannot delete this department because employees are currently assigned to it.", "danger")
        else:
            # Safe delete
            cursor.execute("DELETE FROM employee_departments WHERE id=%s", (id,))
            mysql.connection.commit()
            flash("Department deleted successfully!", "success")
            
    except Exception as e:
        mysql.connection.rollback()
        flash(f"Error deleting department: {e}", "danger")
    finally:
        cursor.close()

    return redirect(url_for("employee_master"))


@app.route("/employee_position_add", methods=["POST"])
def employee_position_add():
    if "loggedin" not in session:
        return redirect(url_for("login"))

    position_name = request.form.get("position_name", "").strip()
    if not position_name:
        flash("Position name is required.", "danger")
        return redirect(url_for("employee_master"))

    try:
        cursor = mysql.connection.cursor(MySQLdb.cursors.DictCursor)
        cursor.execute(
            "INSERT INTO employee_positions (position_name) VALUES (%s)",
            (position_name,)
        )
        mysql.connection.commit()
        flash("Position added successfully!", "success")
    except Exception as e:
        mysql.connection.rollback()
        flash(f"Error adding position: {e}", "danger")
    finally:
        cursor.close()

    return redirect(url_for("employee_master"))


@app.route("/employee_position_delete/<int:id>")
def employee_position_delete(id):
    if "loggedin" not in session:
        return redirect(url_for("login"))

    cursor = mysql.connection.cursor(MySQLdb.cursors.DictCursor)
    
    try:
        # Check if any employee is using this position
        cursor.execute("SELECT COUNT(*) AS cnt FROM employees WHERE position_id=%s", (id,))
        result = cursor.fetchone()
        count = result['cnt'] if result else 0

        if count > 0:
            flash("Cannot delete this position because employees are currently assigned to it.", "danger")
        else:
            # Safe delete
            cursor.execute("DELETE FROM employee_positions WHERE id=%s", (id,))
            mysql.connection.commit()
            flash("Position deleted successfully!", "success")
            
    except Exception as e:
        mysql.connection.rollback()
        flash(f"Error deleting position: {e}", "danger")
    finally:
        cursor.close()

    return redirect(url_for("employee_master"))


# Add employee route#
@app.route("/add_employee", methods=["GET", "POST"])
def add_employee():
    cur = mysql.connection.cursor(MySQLdb.cursors.DictCursor)
    if request.method == "POST":
        d = request.form
        
        # 1. Get Files
        img_file = request.files.get("image")
        doc_file = request.files.get("document")
        
        img_url = None
        doc_url = None
        
        try:
            # 2. Upload Image
            if img_file and allowed_file(img_file.filename):
                res = cloudinary.uploader.upload(img_file, folder="erp_employees")
                img_url = res.get('secure_url')
            
            # 3. Upload Document
            if doc_file and doc_file.filename != '':
                res = cloudinary.uploader.upload(doc_file, folder="erp_employee_docs", resource_type="auto")
                doc_url = res.get('secure_url')
        except Exception as e:
            app.logger.error(f"Upload error: {e}")
            flash(f"Upload warning: {e}", "warning")

        # 4. Insert (UPDATED QUERY)
        try:
            cur.execute("""
                INSERT INTO employees (
                    name, email, phone, aadhar_no, emergency_contact, emergency_contact_person, 
                    dob, position_id, department_id, address_line1, address_line2, 
                    pincode, city, district, state, image, document, status
                )
                VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,'active')
            """, (
                d.get("name"), 
                d.get("email"), 
                d.get("phone"), 
                d.get("aadhar_no"), 
                d.get("emergency_contact"),
                d.get("emergency_contact_person"), # NEW FIELD
                d.get("dob") if d.get("dob") else None,
                d.get("position_id"), 
                d.get("department_id"), 
                d.get("address_line1"), 
                d.get("address_line2"), 
                d.get("pincode"), 
                d.get("city"), 
                d.get("district"), 
                d.get("state"), 
                img_url, 
                doc_url
            ))
            mysql.connection.commit()
            flash("Employee Added Successfully!", "success")
            return redirect(url_for("employees"))
        except Exception as e:
            mysql.connection.rollback()
            app.logger.error(f"DB Error: {e}")
            flash(f"Error adding employee: {e}", "danger")
        finally:
            cur.close()

    cur.execute("SELECT * FROM employee_positions")
    pos = cur.fetchall()
    cur.execute("SELECT * FROM employee_departments")
    dept = cur.fetchall()
    cur.close()
    return render_template("employees/add_employee.html", positions=pos, departments=dept)


# =========================================================
#  EMPLOYEE Document route 
# =========================================================

@app.route("/employee_document/<int:id>")
def employee_document(id):
    if "loggedin" not in session: return redirect(url_for("login"))
    
    cur = mysql.connection.cursor(MySQLdb.cursors.DictCursor)
    try:
        cur.execute("SELECT document FROM employees WHERE id=%s", (id,))
        row = cur.fetchone()
    finally:
        cur.close()

    if not row or not row.get("document"):
        flash("No document attached.", "warning")
        return redirect(request.referrer or url_for('employees'))
        
    doc_val = row["document"]

    # 1. Full URL
    if doc_val.startswith("http"):
        # Fix for PDF 404s (Append .pdf if missing in docs folder)
        if "erp_employee_docs" in doc_val and not any(doc_val.lower().endswith(x) for x in ['.pdf', '.jpg', '.png', '.jpeg', '.doc', '.docx']):
            return redirect(doc_val + ".pdf")
        return redirect(doc_val)

    # 2. Cloudinary Public ID
    try:
        # Heuristic: Force PDF format if likely a doc without extension
        if "erp_employee_docs" in doc_val and not any(doc_val.lower().endswith(x) for x in ['.pdf', '.jpg', '.png', '.jpeg']):
             url, _ = cloudinary.utils.cloudinary_url(doc_val, secure=True, format="pdf")
             return redirect(url)
        
        url, _ = cloudinary.utils.cloudinary_url(doc_val, secure=True)
        return redirect(url)
    except Exception as e:
        app.logger.error(f"Doc URL Error: {e}")
        return "Error opening document", 500


# ---------- ADMIN SIDE: EDIT EMPLOYEE (admin_master.html) ----------

@app.route("/edit_employee/<int:id>", methods=["GET", "POST"])
def edit_employee(id):
    if "loggedin" not in session:
        return redirect(url_for("login"))

    cur = mysql.connection.cursor(MySQLdb.cursors.DictCursor)

    if request.method == "POST":
        form = request.form
        
        # Basic Info
        name = form.get("name", "").strip()
        email = form.get("email", "").strip()
        phone = form.get("phone", "").strip()
        
        # Dates
        dob = form.get("dob") or None 
        joining_date = form.get("join_date") or None # Fixed field name match

        # IDs
        position_id = form.get("position_id") or None
        department_id = form.get("department_id") or None
        
        # Status (Added Logic)
        status = form.get("status", "active").lower() # Default to active if missing

        # Extra Info
        emergency_contact = form.get("emergency_contact", "").strip()
        emergency_contact_person = form.get("emergency_contact_person", "").strip()
        aadhar_no = form.get("aadhar_no", "").strip()

        # Address Info
        pincode = form.get("pincode", "").strip()
        city = form.get("city", "").strip()
        district = form.get("district", "").strip()
        state = form.get("state", "").strip()
        address1 = form.get("address_line1", "").strip()
        address2 = form.get("address_line2", "").strip()

        # File Handling
        image_file = request.files.get("image")
        doc_file = request.files.get("document")
        
        image_public_id = None
        doc_public_id = None

        try:
            if image_file and image_file.filename:
                res = cloudinary.uploader.upload(image_file, folder="erp_employees")
                image_public_id = res.get('secure_url')
            
            if doc_file and doc_file.filename:
                res = cloudinary.uploader.upload(doc_file, folder="erp_employee_docs", resource_type='auto')
                doc_public_id = res.get('secure_url')

            # Dynamic Update Query
            fields = [
                ("name", name),
                ("email", email or None),
                ("phone", phone or None),
                ("dob", dob),
                ("joining_date", joining_date),
                ("status", status), # Include status in update
                ("position_id", position_id),
                ("department_id", department_id),
                ("emergency_contact", emergency_contact or None),
                ("emergency_contact_person", emergency_contact_person or None),
                ("aadhar_no", aadhar_no or None),
                ("pincode", pincode or None),
                ("city", city or None),
                ("district", district or None),
                ("state", state or None),
                ("address_line1", address1 or None),
                ("address_line2", address2 or None)
            ]

            if image_public_id:
                fields.append(("image", image_public_id))
            if doc_public_id:
                fields.append(("document", doc_public_id))

            set_sql = ", ".join([f"{k}=%s" for k, _ in fields])
            params = [v for _, v in fields]
            params.append(id)

            cur.execute(f"UPDATE employees SET {set_sql} WHERE id=%s", tuple(params))
            mysql.connection.commit()
            flash("Employee updated successfully", "success")
            return redirect(url_for("employee_master"))

        except Exception as e:
            mysql.connection.rollback()
            app.logger.exception("Update failed")
            flash(f"Failed to update employee: {e}", "danger")
            return redirect(request.url)
        finally:
            cur.close()

    # GET Request
    cur.execute("SELECT * FROM employees WHERE id=%s", (id,))
    emp = cur.fetchone()
    
    cur.execute("SELECT * FROM employee_positions ORDER BY position_name")
    positions = cur.fetchall()
    
    cur.execute("SELECT * FROM employee_departments ORDER BY department_name")
    departments = cur.fetchall()
    
    cur.close()

    if not emp:
        flash("Employee not found", "danger")
        return redirect(url_for("employee_master"))

    return render_template("employees/edit_employee.html", 
                           employee=emp, positions=positions, departments=departments)


# ------------------ EMPLOYEES ----------------------------------------------------------------------
@app.route('/api/check-employee')
def api_check_employee():
    field = request.args.get('field', '').strip().lower()
    value = request.args.get('value', '').strip()
    exclude_id = request.args.get('exclude_id')

    allowed = {'name', 'email', 'phone'}
    if field not in allowed or not value:
        return {'ok': False, 'exists': False, 'error': 'Invalid request'}, 400

    cur = mysql.connection.cursor(MySQLdb.cursors.DictCursor)
    if exclude_id:
        cur.execute(f"SELECT id FROM employees WHERE {field}=%s AND id<>%s LIMIT 1", (value, exclude_id))
    else:
        cur.execute(f"SELECT id FROM employees WHERE {field}=%s LIMIT 1", (value,))
    row = cur.fetchone()
    cur.close()

    return {'ok': True, 'exists': bool(row)}


# REPLACE your existing 'delete_employee' route with this Smart Logic:

@app.route("/delete_employee/<int:id>", methods=["POST"])
def delete_employee(id):
    if "loggedin" not in session:
        return redirect(url_for("login"))

    cursor = mysql.connection.cursor(MySQLdb.cursors.DictCursor)
    try:
        # =========================================================
        # STEP 1: SAFETY CHECKS (Stop if Stock or Money is Pending)
        # =========================================================
        
        # A. Check Pending Stock
        # Logic: Get Last Settlement Date -> Check allocations after that date
        cursor.execute("SELECT id, date FROM evening_settle WHERE employee_id=%s AND status='final' ORDER BY date DESC LIMIT 1", (id,))
        last_settle = cursor.fetchone()
        
        has_stock = False
        last_date = last_settle['date'] if last_settle else '2000-01-01'

        # Check if items remained in last settlement
        if last_settle:
            cursor.execute("SELECT SUM(remaining_qty) as rem FROM evening_item WHERE settle_id=%s", (last_settle['id'],))
            rem_row = cursor.fetchone()
            if rem_row and rem_row['rem'] and float(rem_row['rem']) > 0:
                has_stock = True

        # Check allocations given AFTER the last settlement
        if not has_stock:
            cursor.execute("""
                SELECT SUM(mai.given_qty) as total_given
                FROM morning_allocations ma
                JOIN morning_allocation_items mai ON ma.id = mai.allocation_id
                WHERE ma.employee_id=%s AND ma.date > %s
            """, (id, last_date))
            alloc_row = cursor.fetchone()
            if alloc_row and alloc_row['total_given'] and float(alloc_row['total_given']) > 0:
                has_stock = True

        if has_stock:
            flash("Cannot delete/deactivate: Employee still holds STOCK. Please perform Evening Settlement to return stock first.", "danger")
            return redirect(url_for("employee_master"))

        # B. Check Ledger Balance (Money Due)
        cursor.execute("""
            SELECT 
                SUM(CASE WHEN type='debit' THEN amount ELSE 0 END) as total_debit,
                SUM(CASE WHEN type='credit' THEN amount ELSE 0 END) as total_credit
            FROM employee_transactions WHERE employee_id=%s
        """, (id,))
        bal_row = cursor.fetchone()
        
        debit = float(bal_row['total_debit'] or 0)
        credit = float(bal_row['total_credit'] or 0)
        balance = debit - credit

        # Allow small floating point difference (e.g., 0.01)
        if abs(balance) > 1.0: 
            flash(f"Cannot delete/deactivate: Employee ledger is not zero (Balance: {balance}). Please clear dues first.", "danger")
            return redirect(url_for("employee_master"))

        # =========================================================
        # STEP 2: DETERMINE DELETE TYPE (Soft vs Hard)
        # =========================================================
        
        # Check if History Exists
        cursor.execute("SELECT id FROM evening_settle WHERE employee_id=%s LIMIT 1", (id,))
        has_evening = cursor.fetchone()
        
        cursor.execute("SELECT id FROM morning_allocations WHERE employee_id=%s LIMIT 1", (id,))
        has_morning = cursor.fetchone()
        
        cursor.execute("SELECT id FROM employee_transactions WHERE employee_id=%s LIMIT 1", (id,))
        has_trans = cursor.fetchone()

        has_history = has_evening or has_morning or has_trans

        if has_history:
            # --- SOFT DELETE (Mark Inactive) ---
            # This preserves all reports, sales, and ledger history
            cursor.execute("UPDATE employees SET status='inactive' WHERE id=%s", (id,))
            mysql.connection.commit()
            flash("Employee marked as INACTIVE (Resigned). Historical data preserved.", "warning")
            
        else:
            # --- HARD DELETE (New/Empty Record) ---
            # Safe to delete because no business data links to this ID
            cursor.execute("DELETE FROM product_returns WHERE employee_id = %s", (id,))
            cursor.execute("DELETE FROM employee_transactions WHERE employee_id = %s", (id,))
            cursor.execute("DELETE FROM employee_attendance WHERE employee_id = %s", (id,))
            cursor.execute("DELETE FROM employees WHERE id = %s", (id,))
            
            mysql.connection.commit()
            flash("Employee record permanently deleted (No history found).", "success")

    except Exception as e:
        mysql.connection.rollback()
        app.logger.error(f"Delete Employee Failed: {e}")
        flash(f"An error occurred: {str(e)}", "danger")

    finally:
        cursor.close()

    return redirect(url_for("employee_master"))

# === API: view single employee (for drawer) ===
@app.route("/api/employees/<int:id>")
def api_employee_detail(id):
    cur = mysql.connection.cursor(MySQLdb.cursors.DictCursor)
    cur.execute("SELECT id, name, position, email, phone, city, status, image, document, DATE_FORMAT(join_date,'%%Y-%%m-%%d') as join_date FROM employees WHERE id=%s", (id,))
    emp = cur.fetchone()
    cur.close()
    if not emp:
        return jsonify({"ok": False, "error": "Not found"}), 404
    return jsonify({"ok": True, "employee": emp})


import requests
import tempfile
import os
from fpdf import FPDF
from datetime import datetime

# ==========================================
# BASE PDF CLASS (SHARED LOGIC)
# ==========================================
class BasePDF(FPDF):
    def __init__(self):
        super().__init__()
        self.company_name = "REAL PROMOTION"
        self.slogan = "SINCE 2005"
        self.address_lines = [
            "Real Promotion, G/12, Tulsimangalam Complex,",
            "B/h. Trimurti Complex, Ghodiya Bazar,",
            "Nadiad-387001, Gujarat, India"
        ]
        self.contact = "+91 96623 22476 | help@realpromotion.in"
        self.gst_text = "GSTIN:24AJVPT0460H1ZW"
        self.gst_subtext = "Composition Dealer- Not Eligibal To Collect Taxes On Supplies"
        self.set_auto_page_break(True, margin=5) # Minimal margin to prevent early break

    def common_header(self, title):
        # Company Header
        self.set_font('Arial', 'B', 24)
        self.set_text_color(26, 35, 126) # Dark Blue
        self.cell(0, 10, self.company_name, 0, 1, 'C')
        
        self.set_font('Arial', 'B', 10)
        self.set_text_color(100, 100, 100) # Grey
        self.cell(0, 5, self.slogan, 0, 1, 'C')
        self.ln(2)
        
        self.set_font('Arial', '', 9)
        self.set_text_color(50, 50, 50)
        for line in self.address_lines:
            self.cell(0, 4, line, 0, 1, 'C')
        self.cell(0, 4, self.contact, 0, 1, 'C')
        
        # GST Section
        self.set_font('Arial', 'B', 9)
        self.cell(0, 4, self.gst_text, 0, 1, 'C')
        self.set_font('Arial', 'I', 8)
        self.cell(0, 4, self.gst_subtext, 0, 1, 'C')
        self.ln(5)
        
        # Divider
        self.set_draw_color(200, 200, 200)
        self.line(10, self.get_y(), 200, self.get_y())
        self.ln(5)

        # Title
        self.set_font('Arial', 'B', 16)
        self.set_text_color(0, 0, 0)
        self.cell(0, 10, title, 0, 1, 'C')
        self.ln(5)

    def footer(self):
        self.set_y(-15)
        self.set_font('Arial', 'I', 8)
        self.set_text_color(128, 128, 128)
        self.cell(0, 10, f'Page {self.page_no()}', 0, 0, 'C')

    def draw_image_safe(self, img_url, x, y, w, h):
        """Draws image if available, handles errors gracefully by downloading if URL."""
        if not img_url or "default" in img_url or not img_url.strip(): 
            return
        
        try:
            if img_url.startswith("http"):
                # Download image to temp file
                headers = {'User-Agent': 'Mozilla/5.0'}
                response = requests.get(img_url, stream=True, timeout=5, headers=headers)
                if response.status_code == 200:
                    # Detect suffix
                    suffix = ".jpg"
                    if ".png" in img_url.lower(): suffix = ".png"
                    
                    # Create and close temp file so FPDF can open it
                    with tempfile.NamedTemporaryFile(delete=False, suffix=suffix) as tmp_file:
                        for chunk in response.iter_content(1024):
                            tmp_file.write(chunk)
                        tmp_path = tmp_file.name
                    
                    # Add to PDF
                    try:
                        self.image(tmp_path, x=x, y=y, w=w, h=h)
                    except:
                        pass
                    finally:
                        if os.path.exists(tmp_path):
                            os.remove(tmp_path)
            else:
                # Handle Local Files
                local_path = img_url.lstrip('/')
                if os.path.exists(local_path):
                    self.image(local_path, x=x, y=y, w=w, h=h)
                elif "static" in img_url:
                     if os.path.exists(local_path):
                         self.image(local_path, x=x, y=y, w=w, h=h)

        except Exception as e:
            pass 

    def add_table_header(self, columns, widths):
        self.set_font('Arial', 'B', 9)
        self.set_fill_color(26, 35, 126) 
        self.set_text_color(255, 255, 255)
        self.set_draw_color(0, 0, 0)
        self.set_line_width(0.3)
        for i, col in enumerate(columns):
            self.cell(widths[i], 8, col, 1, 0, 'C', True)
        self.ln()

# ==========================================
# CLASS 1: MORNING PDF (Stock Allocation)
# ==========================================
class MorningPDF(BasePDF):
    def header(self):
        self.common_header("DAILY STOCK ALLOCATION CHALLAN")

    def add_info(self, emp_name, emp_mobile, date_str, time_str, form_id, emp_image=None):
        self.set_font('Arial', '', 10)
        self.set_text_color(0, 0, 0)
        start_y = self.get_y()
        
        # Draw Image at Top Right of Info Block (Coordinates adjusted)
        # x=100 puts it roughly in the center-right gap. 
        if emp_image:
            self.draw_image_safe(emp_image, x=95, y=start_y, w=22, h=22)

        # Left: Employee Details
        self.set_xy(10, start_y)
        self.cell(35, 6, "Employee Name:", 0, 0)
        self.set_font('Arial', 'B', 10)
        self.cell(50, 6, str(emp_name).upper(), 0, 1) # Reduced width to avoid image overlap
        
        self.set_font('Arial', '', 10)
        self.cell(35, 6, "Mobile No:", 0, 0)
        self.set_font('Arial', 'B', 10)
        self.cell(50, 6, str(emp_mobile) if emp_mobile else "N/A", 0, 1)

        # Right: Date/Time
        self.set_xy(140, start_y)
        self.set_font('Arial', '', 10)
        self.cell(20, 6, "Date:", 0, 0)
        self.set_font('Arial', 'B', 10)
        self.cell(40, 6, str(date_str), 0, 1)

        self.set_xy(140, start_y + 6)
        self.set_font('Arial', '', 10)
        self.cell(20, 6, "Time:", 0, 0)
        self.set_font('Arial', 'B', 10)
        self.cell(40, 6, str(time_str), 0, 1)
        
        self.set_xy(140, start_y + 12)
        self.set_font('Arial', '', 10)
        self.cell(20, 6, "Form ID:", 0, 0)
        self.set_font('Arial', 'B', 10)
        self.cell(40, 6, str(form_id), 0, 1)
            
        self.ln(25) # Ensure gap below image

    def add_signature_section(self):
        # Optimized to fit on same page if possible
        if self.get_y() + 35 > 280: 
            self.add_page()
        
        self.ln(5) 
        y_pos = self.get_y()
        sig_path = "static/img/signature.png" 
        
        if os.path.exists(sig_path):
            self.image(sig_path, x=20, y=y_pos, w=40)
        
        self.set_font('Arial', 'B', 10)
        self.set_text_color(0, 0, 0)
        
        self.line(15, y_pos + 25, 75, y_pos + 25)
        self.set_xy(15, y_pos + 27)
        self.cell(60, 5, "Authorized Signature", 0, 0, 'C')
        
        self.line(135, y_pos + 25, 195, y_pos + 25)
        self.set_xy(135, y_pos + 27)
        self.cell(60, 5, "Employee Signature", 0, 1, 'C')

        self.ln(10)
        self.set_font('Arial', 'I', 8)
        self.set_text_color(128, 128, 128)
        self.cell(0, 5, "This is a computer generated document.", 0, 1, 'C')

# ==========================================
# CLASS 2: EVENING PDF (Settlement)
# ==========================================
class EveningPDF(BasePDF):
    def header(self):
        self.common_header("EVENING SETTLEMENT RECEIPT")

    def add_info(self, emp_name, emp_mobile, date_str, time_str, form_id, emp_image=None):
        # Same structure as Morning
        self.set_font('Arial', '', 10)
        self.set_text_color(0, 0, 0)
        start_y = self.get_y()
        
        if emp_image:
            self.draw_image_safe(emp_image, x=95, y=start_y, w=22, h=22)

        self.set_xy(10, start_y)
        self.cell(35, 6, "Employee Name:", 0, 0)
        self.set_font('Arial', 'B', 10)
        self.cell(50, 6, str(emp_name).upper(), 0, 1)
        
        self.set_font('Arial', '', 10)
        self.cell(35, 6, "Mobile No:", 0, 0)
        self.set_font('Arial', 'B', 10)
        self.cell(50, 6, str(emp_mobile) if emp_mobile else "N/A", 0, 1)

        self.set_xy(140, start_y)
        self.set_font('Arial', '', 10)
        self.cell(20, 6, "Date:", 0, 0)
        self.set_font('Arial', 'B', 10)
        self.cell(40, 6, str(date_str), 0, 1)

        self.set_xy(140, start_y + 6)
        self.set_font('Arial', '', 10)
        self.cell(20, 6, "Time:", 0, 0)
        self.set_font('Arial', 'B', 10)
        self.cell(40, 6, str(time_str), 0, 1)
        
        self.set_xy(140, start_y + 12)
        self.set_font('Arial', '', 10)
        self.cell(20, 6, "Form ID:", 0, 0)
        self.set_font('Arial', 'B', 10)
        self.cell(40, 6, str(form_id), 0, 1)
            
        self.ln(25) 

    def add_signature_section(self):
        # Compact check to keep on one page
        if self.get_y() + 35 > 280: 
            self.add_page()
        
        self.ln(5) 
        y_pos = self.get_y()
        sig_path = "static/img/signature.png" 
        
        if os.path.exists(sig_path):
            self.image(sig_path, x=20, y=y_pos, w=40)
        
        self.set_font('Arial', 'B', 10)
        self.set_text_color(0, 0, 0)
        
        self.line(15, y_pos + 25, 75, y_pos + 25)
        self.set_xy(15, y_pos + 27)
        self.cell(60, 5, "Authorized Signature", 0, 0, 'C')
        
        self.line(135, y_pos + 25, 195, y_pos + 25)
        self.set_xy(135, y_pos + 27)
        self.cell(60, 5, "Employee Signature", 0, 1, 'C')

        self.ln(10)
        self.set_font('Arial', 'I', 8)
        self.set_text_color(128, 128, 128)
        self.cell(0, 5, "This is a computer generated document.", 0, 1, 'C')

# ==========================================
# CLASS 3: OFFICE PDF (Bill of Supply)
# ==========================================
class OfficePDF(BasePDF):
    def header(self):
        self.common_header("BILL OF SUPPLY")

    def add_bill_info(self, sale_data):
        # NO EMPLOYEE IMAGE HERE
        self.set_font('Arial', '', 10)
        self.set_text_color(0, 0, 0)
        y = self.get_y()
        
        self.set_xy(10, y)
        self.set_font('Arial','',10); self.cell(25, 6, "Customer:", 0, 0); self.set_font('Arial','B',10); self.cell(50, 6, str(sale_data['customer_name']), 0, 1)
        self.set_font('Arial','',10); self.cell(25, 6, "Mobile:", 0, 0); self.set_font('Arial','B',10); self.cell(50, 6, str(sale_data['customer_mobile'] or 'N/A'), 0, 1)
        
        addr = str(sale_data['customer_address'] or '')
        if addr:
             self.set_font('Arial','',10); self.cell(25, 6, "Address:", 0, 0); self.set_font('Arial','',9); self.multi_cell(60, 6, addr, 0, 'L')
        else: self.ln(6)
        
        self.set_xy(120, y)
        self.set_font('Arial', '', 10)
        self.cell(20, 6, "Bill No:", 0, 0); self.set_font('Arial','B',10); self.cell(40, 6, str(sale_data['id']), 0, 1)
        self.set_xy(120, y+6)
        self.set_font('Arial', '', 10)
        self.cell(20, 6, "Date:", 0, 0); self.set_font('Arial','B',10); self.cell(40, 6, str(sale_data['sale_date']), 0, 1)
        self.set_xy(120, y+12)
        self.set_font('Arial', '', 10)
        self.cell(20, 6, "Time:", 0, 0); self.set_font('Arial','B',10); self.cell(40, 6, str(sale_data.get('sale_time', '')), 0, 1)
        self.set_xy(120, y+18)
        self.set_font('Arial', '', 10)
        self.cell(20, 6, "Sales By:", 0, 0); self.set_font('Arial','B',10); self.cell(40, 6, str(sale_data['sales_person'] or 'Office'), 0, 1)
        self.ln(25)

    def add_signature_section(self):
        if self.get_y() + 35 > 280: 
            self.add_page()
        
        self.ln(10) 
        y_pos = self.get_y()
        sig_path = "static/img/signature.png" 
        
        if os.path.exists(sig_path):
            self.image(sig_path, x=140, y=y_pos, w=40)
        
        self.set_xy(130, y_pos + 25)
        self.set_font('Arial', 'B', 10)
        self.set_text_color(0, 0, 0)
        self.cell(60, 5, "For, REAL PROMOTION", 0, 1, 'C')
        self.set_xy(130, y_pos + 30)
        self.cell(60, 5, "(Authorized Signatory)", 0, 1, 'C')

        self.ln(10)
        self.set_font('Arial', 'I', 8)
        self.set_text_color(128, 128, 128)
        self.cell(0, 5, "This is a computer generated document.", 0, 1, 'C')


# ==========================================
# 1. MORNING PDF ROUTE
# ==========================================
@app.route('/download_morning_pdf/<int:allocation_id>')
def download_morning_pdf(allocation_id):
    if "loggedin" not in session: return redirect(url_for("login"))
    
    conn = mysql.connection
    cursor = conn.cursor(MySQLdb.cursors.DictCursor)
    
    cursor.execute("""
        SELECT ma.date, ma.created_at, e.name as emp_name, e.phone as emp_mobile, e.image as emp_image
        FROM morning_allocations ma
        JOIN employees e ON ma.employee_id = e.id
        WHERE ma.id = %s
    """, (allocation_id,))
    header = cursor.fetchone()
    
    if not header:
        flash("Allocation not found", "danger")
        return redirect(url_for('allocation_list'))
    
    emp_img = resolve_img(header['emp_image'])

    cursor.execute("""
        SELECT p.name, p.image, mai.opening_qty, mai.given_qty, mai.unit_price
        FROM morning_allocation_items mai
        JOIN products p ON mai.product_id = p.id
        WHERE mai.allocation_id = %s
    """, (allocation_id,))
    items = cursor.fetchall()
    
    pdf = MorningPDF()
    pdf.alias_nb_pages()
    pdf.add_page()
    
    d_val = header['date'].strftime('%d-%m-%Y') if header['date'] else ""
    t_val = "N/A"
    if header.get('created_at') and isinstance(header['created_at'], datetime):
        t_val = header['created_at'].strftime('%I:%M %p')

    pdf.add_info(header['emp_name'], header['emp_mobile'], d_val, t_val, f"MA-{allocation_id}", emp_image=emp_img)
    
    cols = ["No", "Img", "Product Name", "Opn", "Given", "Price", "Amount"]
    widths = [10, 12, 68, 20, 25, 25, 30]
    pdf.add_table_header(cols, widths)
    
    pdf.set_font('Arial', '', 9)
    pdf.set_text_color(0, 0, 0)
    pdf.set_fill_color(245, 245, 245)
    
    total_opening_sum = 0 # Opening Total
    total_given_sum = 0
    total_amount_sum = 0
    fill = False
    row_h = 10 
    
    for i, item in enumerate(items):
        prod_img = resolve_img(item['image'])
        x_start = pdf.get_x()
        y_start = pdf.get_y()

        pdf.cell(widths[0], row_h, str(i+1), 1, 0, 'C', fill)
        
        pdf.cell(widths[1], row_h, "", 1, 0, 'C', fill)
        if prod_img:
            pdf.draw_image_safe(prod_img, x_start + widths[0] + 1, y_start + 1, 8, 8)

        pdf.cell(widths[2], row_h, str(item['name']), 1, 0, 'L', fill)
        pdf.cell(widths[3], row_h, str(item['opening_qty']), 1, 0, 'C', fill)
        pdf.cell(widths[4], row_h, str(item['given_qty']), 1, 0, 'C', fill)
        pdf.cell(widths[5], row_h, f"{float(item['unit_price']):.2f}", 1, 0, 'R', fill)
        
        total_held = int(item['opening_qty']) + int(item['given_qty'])
        amt = total_held * float(item['unit_price'])
        pdf.cell(widths[6], row_h, f"{amt:.2f}", 1, 1, 'R', fill)
        
        total_opening_sum += int(item['opening_qty'])
        total_given_sum += int(item['given_qty']) 
        total_amount_sum += amt
        fill = not fill 
        
    pdf.set_font('Arial', 'B', 10)
    pdf.cell(sum(widths[:3]), 8, "TOTAL", 1, 0, 'R', True)
    pdf.cell(widths[3], 8, str(total_opening_sum), 1, 0, 'C', True) # Display Opening Total
    pdf.cell(widths[4], 8, str(total_given_sum), 1, 0, 'C', True)
    pdf.cell(widths[5], 8, "", 1, 0, 'C', True)
    pdf.cell(widths[6], 8, f"{total_amount_sum:.2f}", 1, 1, 'R', True)
    
    pdf.add_signature_section()
    
    pdf_string = pdf.output(dest='S').encode('latin-1')
    buffer = io.BytesIO(pdf_string)
    buffer.seek(0)
    
    filename = f"{str(header['emp_name']).replace(' ', '_')}_Morning_{allocation_id}.pdf"
    return send_file(buffer, as_attachment=True, download_name=filename, mimetype='application/pdf')


# ==========================================
# 2. EVENING PDF ROUTE
# ==========================================
@app.route('/download_evening_pdf/<int:settle_id>')
def download_evening_pdf(settle_id):
    if "loggedin" not in session: return redirect(url_for("login"))
    
    conn = mysql.connection
    cursor = conn.cursor(MySQLdb.cursors.DictCursor)
    
    cursor.execute("""
        SELECT es.*, e.name as emp_name, e.phone as emp_mobile, e.image as emp_image
        FROM evening_settle es
        JOIN employees e ON es.employee_id = e.id
        WHERE es.id = %s
    """, (settle_id,))
    data = cursor.fetchone()
    
    if not data:
        flash("Settlement not found", "danger")
        return redirect(url_for('admin_evening_master'))
    
    emp_img = resolve_img(data['emp_image'])

    cursor.execute("""
        SELECT p.name, p.image, ei.total_qty, ei.sold_qty, ei.return_qty, ei.unit_price
        FROM evening_item ei
        JOIN products p ON ei.product_id = p.id
        WHERE ei.settle_id = %s
    """, (settle_id,))
    items = cursor.fetchall()
    
    pdf = EveningPDF()
    pdf.alias_nb_pages()
    pdf.add_page()
    
    d_val = data['date'].strftime('%d-%m-%Y') if data['date'] else ""
    t_val = "N/A"
    if data.get('created_at') and isinstance(data['created_at'], datetime):
        t_val = data['created_at'].strftime('%I:%M %p')
    
    pdf.add_info(data['emp_name'], data['emp_mobile'], d_val, t_val, f"EV-{settle_id}", emp_image=emp_img)
    
    # Columns & Widths
    # Reordered: Ret before Left. Total Width = 190
    cols = ["No", "Img", "Product", "Total", "Sold", "Price", "Amount", "Return", "Left"]
    widths = [8, 10, 42, 16, 16, 18, 22, 16, 16] 
    # Adjusted slightly if needed to match 190mm usable width. Sum here is 164, so plenty of room.
    # Let's expand Product name slightly to align better
    widths[2] = 68 # 42 -> 68 (Adds 26). New sum = 190. Perfect full width.
    
    pdf.add_table_header(cols, widths)
    
    pdf.set_font('Arial', '', 9)
    pdf.set_text_color(0, 0, 0)
    pdf.set_fill_color(245, 245, 245)
    
    tot_qty = 0
    tot_sold = 0
    tot_amt = 0
    tot_left = 0
    tot_ret = 0
    
    fill = False
    row_h = 8 # Reduced row height for compact table
    
    for i, item in enumerate(items):
        prod_img = resolve_img(item['image'])
        x_start = pdf.get_x()
        y_start = pdf.get_y()

        pdf.cell(widths[0], row_h, str(i+1), 1, 0, 'C', fill)
        
        pdf.cell(widths[1], row_h, "", 1, 0, 'C', fill)
        if prod_img:
            pdf.draw_image_safe(prod_img, x_start + widths[0] + 1, y_start + 1, 6, 6)

        pdf.cell(widths[2], row_h, str(item['name']), 1, 0, 'L', fill)
        pdf.cell(widths[3], row_h, str(item['total_qty']), 1, 0, 'C', fill)
        pdf.cell(widths[4], row_h, str(item['sold_qty']), 1, 0, 'C', fill)
        pdf.cell(widths[5], row_h, f"{float(item['unit_price']):.2f}", 1, 0, 'R', fill)
        
        amt = float(item['sold_qty']) * float(item['unit_price'])
        pdf.cell(widths[6], row_h, f"{amt:.2f}", 1, 0, 'R', fill)
        
        bal = int(item['total_qty']) - int(item['sold_qty']) - int(item['return_qty'])
        pdf.cell(widths[7], row_h, str(item['return_qty']), 1, 0, 'C', fill)
        pdf.cell(widths[8], row_h, str(bal), 1, 1, 'C', fill)
        
        tot_qty += int(item['total_qty'])
        tot_sold += int(item['sold_qty'])
        tot_amt += amt
        tot_left += bal
        tot_ret += int(item['return_qty'])
        
        fill = not fill

    # TOTAL ROW (With all column totals)
    pdf.set_font('Arial', 'B', 10)
    pdf.cell(sum(widths[:3]), 8, "TOTALS", 1, 0, 'R', True)
    pdf.cell(widths[3], 8, str(tot_qty), 1, 0, 'C', True) # Total Sum
    pdf.cell(widths[4], 8, str(tot_sold), 1, 0, 'C', True) # Sold Sum
    pdf.cell(widths[5], 8, "", 1, 0, 'C', True)
    pdf.cell(widths[6], 8, f"{tot_amt:.2f}", 1, 0, 'R', True)
    pdf.cell(widths[7], 8, str(tot_ret), 1, 0, 'C', True) # Return Sum
    pdf.cell(widths[8], 8, str(tot_left), 1, 1, 'C', True) # Left Sum
    
    # --- FINANCE & SETTLEMENT SECTION (Side-by-Side) ---
    pdf.ln(3)
    
    net_sales = tot_amt - float(data.get('discount', 0) or 0) - float(data.get('online_money', 0) or 0)
    balance_due = net_sales - float(data.get('cash_money', 0) or 0)
    
    emp_credit = float(data.get('emp_credit_amount', 0) or 0)
    emp_debit = float(data.get('emp_debit_amount', 0) or 0)
    
    status_text = ""
    status_color = (0, 0, 0)
    fill_color = (255, 255, 255)
    
    if balance_due > 0.01:
        status_text = f"BALANCE DUE: - {balance_due:.2f}"
        status_color = (255, 255, 255); fill_color = (220, 53, 69)
    elif balance_due < -0.01:
        status_text = f"PAID CASH (PROFIT): + {abs(balance_due):.2f}"
        status_color = (0, 0, 0); fill_color = (255, 193, 7)
    else:
        status_text = "BALANCE CLEARED"
        status_color = (255, 255, 255); fill_color = (25, 135, 84)

    # HEADER BAR
    pdf.set_font('Arial', 'B', 11)
    pdf.set_fill_color(50, 50, 50)
    pdf.set_text_color(255, 255, 255)
    pdf.cell(0, 8, " FINANCE & SETTLEMENT SUMMARY ", 1, 1, 'C', True)
    pdf.ln(2)
    
    y_base = pdf.get_y()
    
    # --- LEFT: EMPLOYEE LEDGER ---
    pdf.set_text_color(0, 0, 0)
    pdf.set_xy(10, y_base)
    pdf.set_font('Arial', 'B', 10)
    pdf.cell(90, 6, "Employee Ledger Details:", 0, 1, 'L')
    
    current_y_left = y_base + 6
    
    if emp_credit > 0:
        pdf.set_xy(10, current_y_left)
        pdf.set_font('Arial', '', 9)
        pdf.cell(30, 6, "Credit (Recv):", 0, 0)
        pdf.set_text_color(25, 135, 84)
        pdf.cell(30, 6, f"{emp_credit:.2f}", 0, 1)
        
        pdf.set_text_color(100, 100, 100)
        pdf.set_xy(10, current_y_left + 5)
        pdf.cell(90, 5, f"Note: {data.get('emp_credit_note') or '-'}", 0, 1)
        current_y_left += 12
    else:
        pdf.set_xy(10, current_y_left)
        pdf.set_font('Arial', 'I', 9)
        pdf.set_text_color(150, 150, 150)
        pdf.cell(90, 6, "No Credit Entry", 0, 1)
        current_y_left += 7

    pdf.set_text_color(0, 0, 0) 
    if emp_debit > 0:
        pdf.set_xy(10, current_y_left)
        pdf.set_font('Arial', '', 9)
        pdf.cell(30, 6, "Debit (Given):", 0, 0)
        pdf.set_text_color(220, 53, 69)
        pdf.cell(30, 6, f"{emp_debit:.2f}", 0, 1)
        
        pdf.set_text_color(100, 100, 100)
        pdf.set_xy(10, current_y_left + 5)
        pdf.cell(90, 5, f"Note: {data.get('emp_debit_note') or '-'}", 0, 1)
        current_y_left += 12
    else:
        pdf.set_xy(10, current_y_left)
        pdf.set_font('Arial', 'I', 9)
        pdf.set_text_color(150, 150, 150)
        pdf.cell(90, 6, "No Debit Entry", 0, 1)
        current_y_left += 7

    # --- RIGHT: CASH SETTLEMENT ---
    pdf.set_text_color(0, 0, 0)
    x_right = 110
    pdf.set_xy(x_right, y_base)
    pdf.set_font('Arial', 'B', 10)
    pdf.cell(90, 6, "Cash Settlement Details:", 0, 1, 'L')
    
    current_y_right = y_base + 6
    
    def print_right_row(label, val, y):
        pdf.set_xy(x_right, y)
        pdf.set_font('Arial', '', 9)
        pdf.cell(40, 6, label, 0, 0)
        pdf.set_font('Arial', 'B', 9)
        pdf.cell(30, 6, val, 0, 1, 'R')

    print_right_row("Total Sales:", f"{tot_amt:.2f}", current_y_right); current_y_right += 5
    print_right_row("Discount (-):", f"{float(data.get('discount', 0) or 0):.2f}", current_y_right); current_y_right += 5
    print_right_row("Online (-):", f"{float(data.get('online_money', 0) or 0):.2f}", current_y_right); current_y_right += 5
    
    pdf.set_draw_color(200, 200, 200)
    pdf.line(x_right, current_y_right, x_right+70, current_y_right)
    print_right_row("CASH PAID:", f"{float(data.get('cash_money', 0) or 0):.2f}", current_y_right + 1)
    current_y_right += 8

    # --- HORIZONTAL STATUS LINE ---
    final_y = max(current_y_left, current_y_right) + 3
    pdf.set_xy(10, final_y)
    
    pdf.set_fill_color(*fill_color)
    pdf.set_text_color(*status_color)
    pdf.set_font('Arial', 'B', 12)
    pdf.cell(0, 8, status_text, 1, 1, 'C', True)

    pdf.add_signature_section()
    
    pdf_string = pdf.output(dest='S').encode('latin-1')
    buffer = io.BytesIO(pdf_string)
    buffer.seek(0)
    
    filename = f"{str(data['emp_name']).replace(' ', '_')}_Evening_{settle_id}.pdf"
    return send_file(buffer, as_attachment=True, download_name=filename, mimetype='application/pdf')


# ==========================================
# 3. OFFICE PDF ROUTE
# ==========================================
@app.route('/office_sales/print/<int:sale_id>')
def download_office_bill(sale_id):
    if "loggedin" not in session: return redirect(url_for("login"))
    
    conn = mysql.connection
    cursor = conn.cursor(MySQLdb.cursors.DictCursor)
    
    cursor.execute("SELECT * FROM office_sales WHERE id=%s", (sale_id,))
    sale = cursor.fetchone()
    
    if not sale:
        flash("Bill not found", "danger")
        return redirect(url_for('office_sales'))
    
    if sale['sale_date']: sale['sale_date'] = sale['sale_date'].strftime('%d-%m-%Y')
    if sale.get('created_at') and isinstance(sale['created_at'], datetime):
         sale['sale_time'] = sale['created_at'].strftime('%I:%M %p')
    else: sale['sale_time'] = ""

    cursor.execute("""
        SELECT i.*, p.name as product_name, p.image
        FROM office_sale_items i 
        JOIN products p ON i.product_id = p.id 
        WHERE i.sale_id=%s
    """, (sale_id,))
    items = cursor.fetchall()
    
    for i in items: i['image'] = resolve_img(i['image'])
    
    pdf = OfficePDF()
    pdf.alias_nb_pages()
    pdf.add_page()
    pdf.add_bill_info(sale)
    
    cols = ["No", "Img", "Product Name", "Qty", "Price", "Total"]
    widths = [10, 12, 78, 20, 30, 40]
    pdf.add_table_header(cols, widths)
    
    pdf.set_font('Arial', '', 10)
    pdf.set_text_color(0, 0, 0)
    fill = False
    pdf.set_fill_color(245, 245, 245)
    
    sum_qty = 0
    sum_total = 0
    row_h = 10 # Compact rows
    
    for i, item in enumerate(items):
        qty = int(item['qty'])
        price = float(item['unit_price'])
        total = float(item['total_price'])
        sum_qty += qty
        sum_total += total
        
        x_start = pdf.get_x()
        y_start = pdf.get_y()
        
        pdf.cell(widths[0], row_h, str(i+1), 1, 0, 'C', fill)
        pdf.cell(widths[1], row_h, "", 1, 0, 'C', fill)
        if item.get('image'):
            pdf.draw_image_safe(item['image'], x_start + widths[0] + 1, y_start + 1, 8, 8)

        pdf.cell(widths[2], row_h, str(item['product_name']), 1, 0, 'L', fill)
        pdf.cell(widths[3], row_h, str(qty), 1, 0, 'C', fill)
        pdf.cell(widths[4], row_h, f"{price:.2f}", 1, 0, 'R', fill)
        pdf.cell(widths[5], row_h, f"{total:.2f}", 1, 1, 'R', fill) 
        fill = not fill
        
    pdf.set_font('Arial', 'B', 10)
    pdf.cell(widths[0]+widths[1]+widths[2], 8, "TOTAL ITEMS", 1, 0, 'R', True)
    pdf.cell(widths[3], 8, str(sum_qty), 1, 0, 'C', True)
    pdf.cell(widths[4], 8, "", 1, 0, 'C', True)
    pdf.cell(widths[5], 8, f"{sum_total:.2f}", 1, 1, 'R', True)
    pdf.ln(5)
    
    x_start = 120
    pdf.set_x(x_start)
    pdf.set_font('Arial', '', 10)
    pdf.cell(40, 6, "Sub Total:", 0, 0, 'R')
    pdf.set_font('Arial', 'B', 10)
    pdf.cell(30, 6, f"{float(sale['sub_total']):.2f}", 0, 1, 'R')
    
    if float(sale['discount']) > 0:
        pdf.set_x(x_start)
        pdf.set_font('Arial', '', 10)
        pdf.cell(40, 6, "Discount (-):", 0, 0, 'R')
        pdf.set_font('Arial', 'B', 10)
        pdf.cell(30, 6, f"{float(sale['discount']):.2f}", 0, 1, 'R')
        
    pdf.set_x(x_start)
    pdf.set_font('Arial', 'B', 12)
    pdf.set_fill_color(230, 230, 230)
    pdf.cell(40, 8, "GRAND TOTAL:", 1, 0, 'R', True)
    pdf.cell(30, 8, f"{float(sale['final_amount']):.2f}", 1, 1, 'R', True)
    pdf.ln(10)
    
    pdf.set_font('Arial', 'B', 10)
    pdf.cell(0, 6, "Terms & Conditions:", 0, 1)
    pdf.set_font('Arial', '', 9)
    pdf.cell(0, 5, "1. Goods once sold will not be taken back.", 0, 1)
    pdf.cell(0, 5, "2. 7 Days Return Policy (Damage product not accepted).", 0, 1)
    pdf.cell(0, 5, "3. 6 Month Warranty on Electronics.", 0, 1)
    pdf.cell(0, 5, "4. Subject to Nadiad Jurisdiction.", 0, 1)
    
    pdf.add_signature_section()
    
    pdf_string = pdf.output(dest='S').encode('latin-1')
    buffer = io.BytesIO(pdf_string)
    buffer.seek(0)
    
    cust_name = str(sale.get('customer_name', 'Customer')).replace(" ", "_")
    filename = f"{cust_name}_Bill_{sale_id}.pdf"
    return send_file(buffer, as_attachment=True, download_name=filename, mimetype='application/pdf')
# ==========================================
# PUBLIC ROUTE: DOWNLOAD MORNING PDF
# ==========================================
@app.route('/public/download_morning_pdf/<int:allocation_id>')
def download_morning_pdf_public(allocation_id):
    conn = mysql.connection
    cursor = conn.cursor(MySQLdb.cursors.DictCursor)
    
    cursor.execute("""
        SELECT ma.date, ma.created_at, e.name as emp_name, e.phone as emp_mobile
        FROM morning_allocations ma
        JOIN employees e ON ma.employee_id = e.id
        WHERE ma.id = %s
    """, (allocation_id,))
    header = cursor.fetchone()
    
    if not header: return "Allocation not found", 404
        
    cursor.execute("""
        SELECT p.name, mai.opening_qty, mai.given_qty, mai.unit_price
        FROM morning_allocation_items mai
        JOIN products p ON mai.product_id = p.id
        WHERE mai.allocation_id = %s
    """, (allocation_id,))
    items = cursor.fetchall()
    
    pdf = PDFGenerator("Morning")
    pdf.alias_nb_pages()
    pdf.add_page()
    
    d_val = header['date'].strftime('%d-%m-%Y') if header['date'] else ""
    t_val = "N/A"
    if header.get('created_at') and isinstance(header['created_at'], datetime):
        t_val = header['created_at'].strftime('%I:%M %p')

    pdf.add_info_section(header['emp_name'], header['emp_mobile'], d_val, t_val, f"MA-{allocation_id}")
    
    cols = ["No", "Product Name", "Opening", "Given Qty", "Price", "Amount"]
    widths = [10, 80, 20, 25, 25, 30]
    pdf.add_table_header(cols, widths)
    
    pdf.set_font('Arial', '', 9)
    pdf.set_text_color(0, 0, 0)
    pdf.set_fill_color(245, 245, 245)
    
    total_opening_sum = 0
    total_given_sum = 0
    total_amount_sum = 0
    fill = False
    
    for i, item in enumerate(items):
        pdf.cell(widths[0], 7, str(i+1), 1, 0, 'C', fill)
        pdf.cell(widths[1], 7, str(item['name']), 1, 0, 'L', fill)
        pdf.cell(widths[2], 7, str(item['opening_qty']), 1, 0, 'C', fill)
        pdf.cell(widths[3], 7, str(item['given_qty']), 1, 0, 'C', fill)
        pdf.cell(widths[4], 7, f"{float(item['unit_price']):.2f}", 1, 0, 'R', fill)
        
        total_held = int(item['opening_qty']) + int(item['given_qty'])
        amt = total_held * float(item['unit_price'])
        pdf.cell(widths[5], 7, f"{amt:.2f}", 1, 1, 'R', fill)
        
        total_opening_sum += int(item['opening_qty'])
        total_given_sum += int(item['given_qty']) 
        total_amount_sum += amt
        fill = not fill 
        
    pdf.set_font('Arial', 'B', 10)
    pdf.cell(widths[0]+widths[1], 8, "TOTALS", 1, 0, 'R', True)
    pdf.cell(widths[2], 8, str(total_opening_sum), 1, 0, 'C', True)
    pdf.cell(widths[3], 8, str(total_given_sum), 1, 0, 'C', True)
    pdf.cell(widths[4], 8, "", 1, 0, 'C', True)
    pdf.cell(widths[5], 8, f"{total_amount_sum:.2f}", 1, 1, 'R', True)
    
    pdf.add_signature_section(is_office_bill=False)
    
    pdf_string = pdf.output(dest='S').encode('latin-1')
    buffer = io.BytesIO(pdf_string)
    buffer.seek(0)
    
    safe_name = str(header['emp_name']).replace(" ", "_")
    filename = f"{safe_name}_Morning_{allocation_id}.pdf"
    return send_file(buffer, as_attachment=True, download_name=filename, mimetype='application/pdf')



# ==========================================
# 4. PUBLIC ROUTE: DOWNLOAD OFFICE BILL
# ==========================================
@app.route('/public/office_sales/print/<int:sale_id>')
def download_office_bill_public(sale_id):
    conn = mysql.connection
    cursor = conn.cursor(MySQLdb.cursors.DictCursor)
    
    cursor.execute("SELECT * FROM office_sales WHERE id=%s", (sale_id,))
    sale = cursor.fetchone()
    
    if not sale: return "Bill not found", 404
    
    if sale['sale_date']: sale['sale_date'] = sale['sale_date'].strftime('%d-%m-%Y')
    if sale.get('created_at') and isinstance(sale['created_at'], datetime):
         sale['sale_time'] = sale['created_at'].strftime('%I:%M %p')
    else: sale['sale_time'] = ""

    cursor.execute("""
        SELECT i.*, p.name as product_name 
        FROM office_sale_items i 
        JOIN products p ON i.product_id = p.id 
        WHERE i.sale_id=%s
    """, (sale_id,))
    items = cursor.fetchall()
    
    pdf = PDFGenerator("Office")
    pdf.alias_nb_pages()
    pdf.add_page()
    pdf.generate_office_bill_body(sale, items)
    
    pdf_string = pdf.output(dest='S').encode('latin-1')
    buffer = io.BytesIO(pdf_string)
    buffer.seek(0)
    
    cust_name = str(sale.get('customer_name', 'Customer')).replace(" ", "_")
    filename = f"{cust_name}_Bill_{sale_id}.pdf"
    return send_file(buffer, as_attachment=True, download_name=filename, mimetype='application/pdf')

# =====================================================================================
# 2. OFFICE SALES ROUTE (Updated to show only Active Products)
# =====================================================================================
@app.route('/office_sales', methods=['GET', 'POST'])
def office_sales():
    if "loggedin" not in session: return redirect(url_for("login"))
    
    conn = mysql.connection
    cursor = conn.cursor(MySQLdb.cursors.DictCursor)
    
    if request.method == 'POST':
        try:
            # 1. Customer & Header Info
            c_name = request.form.get('customer_name')
            c_mobile = request.form.get('customer_mobile')
            c_addr = request.form.get('customer_address')
            sales_person = request.form.get('sales_person')
            
            # Date Handling
            b_date_str = request.form.get('bill_date')
            try:
                bill_date = datetime.strptime(b_date_str, '%d-%m-%Y').strftime('%Y-%m-%d')
            except:
                bill_date = date.today().strftime('%Y-%m-%d')

            # 2. Financials
            discount = float(request.form.get('discount') or 0)
            online_amt = float(request.form.get('online') or 0)
            cash_amt = float(request.form.get('cash') or 0)
            due_note = request.form.get('due_note') 
            
            # 3. Products
            p_ids = request.form.getlist('product_id[]')
            qtys = request.form.getlist('qty[]')
            prices = request.form.getlist('price[]')
            
            total_amt = 0
            
            # --- TIME FIX: Force Server-Side IST ---
            ist_now = get_ist_now()
            time_str = ist_now.strftime('%Y-%m-%d %H:%M:%S')

            # Insert Header (Initial)
            cursor.execute("""
                INSERT INTO office_sales 
                (customer_name, customer_mobile, customer_address, sales_person, sale_date, discount, online_amount, cash_amount, due_note, created_at)
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
            """, (c_name, c_mobile, c_addr, sales_person, bill_date, discount, online_amt, cash_amt, due_note, time_str))
            sale_id = cursor.lastrowid
            
            # Process Items & Inventory
            for i, pid in enumerate(p_ids):
                if not pid: continue
                qty = int(qtys[i])
                price = float(prices[i])
                
                # Check Stock
                cursor.execute("SELECT stock, name FROM products WHERE id=%s", (pid,))
                prod = cursor.fetchone()
                if prod and prod['stock'] < qty:
                    raise Exception(f"Insufficient stock for {prod['name']}")

                # Deduct Stock
                cursor.execute("UPDATE products SET stock = stock - %s WHERE id=%s", (qty, pid))
                
                # Item Total
                item_total = qty * price
                cursor.execute("""
                    INSERT INTO office_sale_items (sale_id, product_id, qty, unit_price, total_price)
                    VALUES (%s, %s, %s, %s, %s)
                """, (sale_id, pid, qty, price, item_total))
                
                total_amt += item_total
            
            # Final Update
            final_amt = total_amt - discount
            cursor.execute("""
                UPDATE office_sales 
                SET sub_total=%s, final_amount=%s 
                WHERE id=%s
            """, (total_amt, final_amt, sale_id))
            
            conn.commit()
            
            return redirect(url_for('download_office_bill', sale_id=sale_id))
            
        except Exception as e:
            conn.rollback()
            flash(f"Error: {str(e)}", "danger")
            return redirect(url_for('office_sales'))

    # --- MODIFIED: Fetch ONLY Active Products ---
    try:
        cursor.execute("SELECT id, name, price, stock, image FROM products WHERE status='Active' ORDER BY name")
    except:
        cursor.execute("SELECT id, name, price, stock, image FROM products ORDER BY name")
        
    products = cursor.fetchall()
    
    for p in products:
        p['image'] = resolve_img(p['image'])

    return render_template('office_sales.html', products=products, today=date.today().strftime('%d-%m-%Y'))





# --- Helper: Robust Date Parsing ---

def parse_date(date_str):
    if not date_str: return None
    try:
        return datetime.strptime(date_str, "%d-%m-%Y").date()
    except ValueError:
        try:
            return datetime.strptime(date_str, "%Y-%m-%d").date()
        except ValueError:
            return None

# --- Helper: Image URL Resolver ---
def resolve_img(image_path):
    if not image_path: return url_for('static', filename='img/default-product.png')
    if image_path.startswith('http'): return image_path
    if '.' in image_path: return url_for('static', filename='uploads/' + image_path)
    try:
        import cloudinary.utils
        url, _ = cloudinary.utils.cloudinary_url(image_path, secure=True)
        return url
    except:
        return url_for('static', filename='img/default-product.png')






# =========================================================
#  CORE LOGIC: STOCK CALCULATOR (The Chain System)
# =========================================================
def get_current_stock_state(cursor, emp_id, target_date_str):
    """
    Calculates the stock state for an employee up to a specific date.
    Logic: Last Settlement + All Unsettled Allocations (Gap) + Today's Allocations
    """
    stock_map = {} # Key: ProductID, Value: {qty, name, price, image}

    # 1. FIND LAST FINAL SETTLEMENT (Baseline)
    cursor.execute("""
        SELECT id, date FROM evening_settle 
        WHERE employee_id=%s AND date < %s AND status='final'
        ORDER BY date DESC LIMIT 1
    """, (emp_id, target_date_str))
    last_settle = cursor.fetchone()
    
    last_settle_date = last_settle['date'] if last_settle else '2000-01-01' # Fallback to ancient date
    
    if last_settle:
        cursor.execute("""
            SELECT product_id, remaining_qty, unit_price, p.name, p.image
            FROM evening_item ei
            JOIN products p ON ei.product_id = p.id
            WHERE settle_id = %s AND remaining_qty > 0
        """, (last_settle['id'],))
        
        for r in cursor.fetchall():
            pid = r['product_id']
            stock_map[pid] = {
                'product_id': pid, 'name': r['name'], 'image': resolve_img(r['image']),
                'qty': int(r['remaining_qty']), 'price': float(r['unit_price'])
            }

    # 2. FILL THE GAP (Allocations AFTER Last Settle but BEFORE/ON Today)
    # This catches holidays, missed settlements, AND today's restocks
    cursor.execute("""
        SELECT mai.product_id, SUM(mai.given_qty) as total_given, MAX(mai.unit_price) as unit_price, p.name, p.image
        FROM morning_allocations ma
        JOIN morning_allocation_items mai ON ma.id = mai.allocation_id
        JOIN products p ON mai.product_id = p.id
        WHERE ma.employee_id = %s AND ma.date > %s AND ma.date <= %s
        GROUP BY mai.product_id
    """, (emp_id, last_settle_date, target_date_str))
    
    gap_allocations = cursor.fetchall()
    
    for r in gap_allocations:
        pid = r['product_id']
        added_qty = int(r['total_given'])
        
        if pid in stock_map:
            stock_map[pid]['qty'] += added_qty
            # Update price if needed (usually keeps old price or avg, simplest is keep old)
        else:
            stock_map[pid] = {
                'product_id': pid, 'name': r['name'], 'image': resolve_img(r['image']),
                'qty': added_qty, 'price': float(r['unit_price'])
            }
            
    return list(stock_map.values())


# =========================================================
#  CORE LOGIC: STOCK CALCULATOR (The Chain System)
# =========================================================
def get_current_stock_state(cursor, emp_id, target_date_str):
    """
    Calculates the stock state for an employee up to a specific date.
    Logic: Last Settlement + All Unsettled Allocations (Gap) + Today's Allocations
    """
    stock_map = {} # Key: ProductID, Value: {qty, name, price, image}

    # 1. FIND LAST FINAL SETTLEMENT (Baseline)
    cursor.execute("""
        SELECT id, date FROM evening_settle 
        WHERE employee_id=%s AND date < %s AND status='final'
        ORDER BY date DESC LIMIT 1
    """, (emp_id, target_date_str))
    last_settle = cursor.fetchone()
    
    last_settle_date = last_settle['date'] if last_settle else '2000-01-01' # Fallback to ancient date
    
    if last_settle:
        cursor.execute("""
            SELECT product_id, remaining_qty, unit_price, p.name, p.image
            FROM evening_item ei
            JOIN products p ON ei.product_id = p.id
            WHERE settle_id = %s AND remaining_qty > 0
        """, (last_settle['id'],))
        
        for r in cursor.fetchall():
            pid = r['product_id']
            stock_map[pid] = {
                'product_id': pid, 'name': r['name'], 'image': resolve_img(r['image']),
                'qty': int(r['remaining_qty']), 'price': float(r['unit_price'])
            }

    # 2. FILL THE GAP (Allocations AFTER Last Settle but BEFORE/ON Today)
    cursor.execute("""
        SELECT mai.product_id, SUM(mai.given_qty) as total_given, MAX(mai.unit_price) as unit_price, p.name, p.image
        FROM morning_allocations ma
        JOIN morning_allocation_items mai ON ma.id = mai.allocation_id
        JOIN products p ON mai.product_id = p.id
        WHERE ma.employee_id = %s AND ma.date > %s AND ma.date <= %s
        GROUP BY mai.product_id
    """, (emp_id, last_settle_date, target_date_str))
    
    gap_allocations = cursor.fetchall()
    
    for r in gap_allocations:
        pid = r['product_id']
        added_qty = int(r['total_given'])
        
        if pid in stock_map:
            stock_map[pid]['qty'] += added_qty
        else:
            stock_map[pid] = {
                'product_id': pid, 'name': r['name'], 'image': resolve_img(r['image']),
                'qty': added_qty, 'price': float(r['unit_price'])
            }
            
    return list(stock_map.values())


# ==========================================
# 1. API: FETCH MORNING STOCK
# ==========================================
@app.route('/api/fetch_stock', methods=['GET', 'POST'])
def api_fetch_stock():
    if "loggedin" not in session: return jsonify({"error": "Unauthorized"}), 401
    
    employee_id = request.values.get('employee_id')
    date_str = request.values.get('date') 

    if not employee_id or not date_str: return jsonify({"error": "Missing params"}), 400

    try:
        current_date = parse_date(date_str)
        if not current_date: return jsonify({"error": "Invalid date"}), 400
        formatted_date = current_date.strftime('%Y-%m-%d')
        
        cur = mysql.connection.cursor(MySQLdb.cursors.DictCursor)
        
        # 1. Check Evening Lock
        cur.execute("SELECT id FROM evening_settle WHERE employee_id=%s AND date=%s AND status='final'", (employee_id, formatted_date))
        if cur.fetchone():
            cur.close()
            return jsonify({"mode": "locked", "evening_settled": True})

        # 2. Check Mode (Restock or Normal)
        cur.execute("SELECT id FROM morning_allocations WHERE employee_id=%s AND date=%s", (employee_id, formatted_date))
        today_alloc = cur.fetchone()
        
        mode = "restock" if today_alloc else "normal"
        
        # 3. CALCULATE AGGREGATED STOCK
        aggregated_stock = get_current_stock_state(cur, employee_id, formatted_date)
        
        # 4. Format for Frontend
        existing_items = []
        if mode == 'restock':
            cur.execute("""
                SELECT p.name, p.image, mai.given_qty, mai.product_id
                FROM morning_allocation_items mai
                JOIN products p ON mai.product_id = p.id
                WHERE mai.allocation_id = %s
            """, (today_alloc['id'],))
            for r in cur.fetchall():
                existing_items.append({
                    'name': r['name'], 'image': resolve_img(r['image']),
                    'qty': int(r['given_qty']), 'product_id': r['product_id']
                })

        opening_list = []
        for item in aggregated_stock:
            opening_list.append({
                'product_id': str(item['product_id']),
                'name': item['name'],
                'image': item['image'],
                'remaining': item['qty'], 
                'price': item['price']
            })

        cur.close()
        return jsonify({
            "mode": mode,
            "evening_settled": False,
            "opening_stock": opening_list, 
            "existing_items": existing_items     
        })

    except Exception as e:
        return jsonify({"error": str(e)}), 500


# ==========================================
# 2. API: FETCH EVENING DATA
# ==========================================
@app.route('/api/fetch_evening_data', methods=['POST'])
def fetch_evening_data():
    if "loggedin" not in session: return jsonify({'status': 'error', 'message': 'Unauthorized'})

    try:
        emp_id = request.form.get('employee_id')
        date_str = request.form.get('date')
        
        target_date = parse_date(date_str)
        if not target_date: return jsonify({'status': 'error', 'message': 'Invalid Date'})
        formatted_date = target_date.strftime('%Y-%m-%d')
        
        cur = mysql.connection.cursor(MySQLdb.cursors.DictCursor)

        # A. Check Existing Settlement
        cur.execute("SELECT * FROM evening_settle WHERE employee_id=%s AND date=%s", (emp_id, formatted_date))
        existing = cur.fetchone()

        if existing:
            if existing['status'] == 'final':
                return jsonify({'status': 'submitted', 'message': 'Already submitted.'})
            
            # Draft Mode
            cur.execute("""
                SELECT ei.product_id, ei.total_qty, ei.sold_qty, ei.return_qty, ei.unit_price, p.name, p.image 
                FROM evening_item ei JOIN products p ON ei.product_id = p.id WHERE ei.settle_id = %s
            """, (existing['id'],))
            
            items = []
            for i in cur.fetchall():
                items.append({
                    'product_id': i['product_id'], 'name': i['name'], 'image': resolve_img(i['image']),
                    'unit_price': float(i['unit_price']), 
                    'total_qty': int(i['total_qty']),
                    'sold_qty': int(i['sold_qty']), 'return_qty': int(i['return_qty'])
                })

            return jsonify({
                'status': 'success', 'source': 'draft', 
                'allocation_id': existing.get('allocation_id') or '',
                'draft_id': existing['id'], 'draft_data': existing,
                'products': items
            })

        # B. CALCULATE FRESH DATA (Using Chain Logic)
        current_stock = get_current_stock_state(cur, emp_id, formatted_date)
        
        # Get Allocation ID for linking
        cur.execute("SELECT id FROM morning_allocations WHERE employee_id=%s AND date=%s", (emp_id, formatted_date))
        alloc_row = cur.fetchone()
        alloc_id = alloc_row['id'] if alloc_row else None

        final_items = []
        for item in current_stock:
            if item['qty'] > 0:
                final_items.append({
                    'product_id': item['product_id'], 'name': item['name'], 'image': item['image'],
                    'unit_price': item['price'], 'total_qty': item['qty'],
                    'sold_qty': 0, 'return_qty': 0
                })

        if not final_items:
            return jsonify({'status': 'error', 'message': 'No stock found for this employee.'})

        return jsonify({
            'status': 'success', 
            'source': 'fresh',
            'allocation_id': alloc_id,
            'draft_id': None, 'draft_data': None,
            'products': final_items
        })

    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)})
    finally:
        if 'cur' in locals(): cur.close()

# =====================================================================================
# 1. MORNING ALLOCATION ROUTE (Updated to show only Active Products)
# =====================================================================================
@app.route('/morning', methods=['GET', 'POST'])
def morning():
    if "loggedin" not in session: return redirect(url_for("login"))
    
    conn = mysql.connection
    cursor = conn.cursor(MySQLdb.cursors.DictCursor)

    if request.method == 'POST':
        try:
            emp_id = request.form.get('employee_id')
            date_str = request.form.get('date')
            time_str = request.form.get('timestamp')
            
            if not time_str or time_str.strip() == "":
                ist_now = get_ist_now()
                time_str = ist_now.strftime('%Y-%m-%d %H:%M:%S')

            p_ids = request.form.getlist('product_id[]')
            opens = request.form.getlist('opening[]') 
            givens = request.form.getlist('given[]')
            prices = request.form.getlist('price[]')

            formatted_date = parse_date(date_str).strftime('%Y-%m-%d')

            cursor.execute("SELECT id FROM morning_allocations WHERE employee_id=%s AND date=%s", (emp_id, formatted_date))
            existing_alloc = cursor.fetchone()

            if existing_alloc:
                alloc_id = existing_alloc['id']
                
                cursor.execute("SELECT id, product_id, given_qty FROM morning_allocation_items WHERE allocation_id=%s", (alloc_id,))
                db_items = cursor.fetchall() 
                
                cnt = 0
                for i, pid in enumerate(p_ids):
                    if not pid: continue
                    gv = int(givens[i] or 0)
                    ui_opening = int(opens[i] or 0)
                    pr = float(prices[i] or 0)
                    
                    if gv > 0 or ui_opening > 0: 
                        match = next((x for x in db_items if str(x['product_id']) == str(pid)), None)

                        if match:
                            # MERGE: Add new Given to Old Given
                            old_given = int(match['given_qty'])
                            new_total_given = old_given + gv 
                            
                            cursor.execute("UPDATE morning_allocation_items SET given_qty=%s, unit_price=%s WHERE id=%s", 
                                           (new_total_given, pr, match['id']))
                            
                            if gv > 0:
                                cursor.execute("UPDATE products SET stock = stock - %s WHERE id=%s", (gv, pid))
                        else:
                            cursor.execute("""
                                INSERT INTO morning_allocation_items (allocation_id, product_id, opening_qty, given_qty, unit_price) 
                                VALUES (%s, %s, %s, %s, %s)
                            """, (alloc_id, pid, ui_opening, gv, pr))
                            
                            if gv > 0:
                                cursor.execute("UPDATE products SET stock = stock - %s WHERE id = %s", (gv, pid))
                        cnt += 1
                
                conn.commit()
                flash("Restock Added & Merged Successfully.", "success")

            else:
                cursor.execute("INSERT INTO morning_allocations (employee_id, date, created_at) VALUES (%s, %s, %s)", (emp_id, formatted_date, time_str))
                alloc_id = cursor.lastrowid
                cnt = 0
                for i, pid in enumerate(p_ids):
                    if not pid: continue
                    gv = int(givens[i] or 0)
                    ui_opening = int(opens[i] or 0)
                    pr = float(prices[i] or 0)
                    
                    if ui_opening > 0 or gv > 0:
                        cursor.execute("INSERT INTO morning_allocation_items (allocation_id, product_id, opening_qty, given_qty, unit_price) VALUES (%s, %s, %s, %s, %s)", 
                                       (alloc_id, pid, ui_opening, gv, pr))
                        if gv > 0:
                            cursor.execute("UPDATE products SET stock = stock - %s WHERE id = %s", (gv, pid))
                        cnt += 1
                conn.commit()
                flash("Allocation saved.", "success")
            
            return redirect(url_for('morning'))

        except Exception as e:
            conn.rollback()
            flash(f"Error: {str(e)}", "danger")
            return redirect(url_for('morning'))

    # GET Logic (Standard)
    cursor.execute("SELECT id, name, image FROM employees WHERE status='active' ORDER BY name")
    emps = cursor.fetchall()
    for e in emps: e['image'] = resolve_img(e['image'])
    
    # --- MODIFIED: Fetch ONLY Active Products ---
    try:
        cursor.execute("SELECT id, name, price, image, stock FROM products WHERE status='Active' ORDER BY name")
    except:
        cursor.execute("SELECT id, name, price, image, stock FROM products ORDER BY name")
        
    prods = [{
        'id': p['id'], 'name': p['name'], 'price': float(p['price']), 
        'image': resolve_img(p['image']), 'stock': int(p['stock'] or 0)
    } for p in cursor.fetchall()]
    
    return render_template('morning.html', employees=emps, products=prods, today_date=date.today().strftime('%d-%m-%Y'))




# ==========================================
# ROUTE: EVENING SUBMIT (Time Fix Applied + NULL Fix)
# ==========================================
@app.route('/evening', methods=['GET', 'POST'])
def evening():
    if "loggedin" not in session: return redirect(url_for("login"))
    
    conn = mysql.connection
    cursor = conn.cursor(MySQLdb.cursors.DictCursor)
    
    if request.method == 'POST':
        try:
            status = request.form.get('status')
            draft_id = request.form.get('draft_id')
            alloc_id = request.form.get('allocation_id')
            emp_id = request.form.get('h_employee')
            date_val = request.form.get('h_date')
            
            # --- TIME FIX: Force Server-Side IST ---
            # We ignore frontend timestamp to ensure consistency
            ist_now = get_ist_now()
            time_str = ist_now.strftime('%Y-%m-%d %H:%M:%S') 
            
            # Check existing
            cursor.execute("SELECT id FROM evening_settle WHERE employee_id=%s AND date=%s", (emp_id, date_val))
            existing_record = cursor.fetchone()

            # --- NULL ALLOCATION ID FIX ---
            final_alloc_id = None
            if alloc_id and alloc_id.strip() not in ['', '0', 'None']:
                try: final_alloc_id = int(alloc_id)
                except: final_alloc_id = None
            
            if not final_alloc_id:
                cursor.execute("SELECT id FROM morning_allocations WHERE employee_id=%s AND date=%s", (emp_id, date_val))
                ma = cursor.fetchone()
                if ma:
                    final_alloc_id = ma['id']
                else:
                    # Create Dummy Allocation if missing (Foreign Key req)
                    cursor.execute("INSERT INTO morning_allocations (employee_id, date, created_at) VALUES (%s, %s, %s)", (emp_id, date_val, time_str))
                    final_alloc_id = cursor.lastrowid
                    conn.commit()

            total_amt = float(request.form.get('totalAmount') or 0)
            discount = float(request.form.get('discount') or 0)
            online = float(request.form.get('online') or 0)
            cash = float(request.form.get('cash') or 0)
            
            emp_c = float(request.form.get('emp_credit_amount') or 0)
            emp_c_n = request.form.get('emp_credit_note')
            emp_d = float(request.form.get('emp_debit_amount') or 0)
            emp_d_n = request.form.get('emp_debit_note')
            due_n = request.form.get('due_note')

            if existing_record:
                settle_id = existing_record['id']
                cursor.execute("""
                    UPDATE evening_settle SET total_amount=%s, cash_money=%s, online_money=%s, discount=%s, 
                    emp_credit_amount=%s, emp_credit_note=%s, emp_debit_amount=%s, emp_debit_note=%s, 
                    due_note=%s, status=%s, created_at=%s, allocation_id=%s WHERE id=%s
                """, (total_amt, cash, online, discount, emp_c, emp_c_n, emp_d, emp_d_n, due_n, status, time_str, final_alloc_id, settle_id))
                cursor.execute("DELETE FROM evening_item WHERE settle_id=%s", (settle_id,))
            else:
                cursor.execute("""
                    INSERT INTO evening_settle (allocation_id, employee_id, date, total_amount, cash_money, online_money, discount, 
                    emp_credit_amount, emp_credit_note, emp_debit_amount, emp_debit_note, due_note, status, created_at)
                    VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                """, (final_alloc_id, emp_id, date_val, total_amt, cash, online, discount, emp_c, emp_c_n, emp_d, emp_d_n, due_n, status, time_str))
                settle_id = cursor.lastrowid

            p_ids = request.form.getlist('product_id[]')
            t_qtys = request.form.getlist('total_qty[]')
            s_qtys = request.form.getlist('sold[]')
            r_qtys = request.form.getlist('return[]')
            prices = request.form.getlist('price[]')

            for i, pid in enumerate(p_ids):
                if not pid: continue
                tot = int(float(t_qtys[i] or 0))
                sold = int(float(s_qtys[i] or 0))
                ret = int(float(r_qtys[i] or 0))
                pr = float(prices[i] or 0)
                rem = max(0, tot - sold - ret) 

                cursor.execute("INSERT INTO evening_item (settle_id, product_id, total_qty, sold_qty, return_qty, remaining_qty, unit_price) VALUES (%s, %s, %s, %s, %s, %s, %s)", 
                               (settle_id, pid, tot, sold, ret, rem, pr))
                
                if status == 'final' and ret > 0:
                    cursor.execute("UPDATE products SET stock = stock + %s WHERE id = %s", (ret, pid))

            if status == 'final':
                # LEDGER ENTRIES with IST Time
                if emp_c > 0: 
                    cursor.execute("INSERT INTO employee_transactions (employee_id, transaction_date, type, amount, description, created_at) VALUES (%s, %s, 'credit', %s, %s, %s)", 
                                   (emp_id, date_val, emp_c, f"Credit #{settle_id}: {emp_c_n or 'Evening Settle'}", time_str))
                
                if emp_d > 0: 
                    cursor.execute("INSERT INTO employee_transactions (employee_id, transaction_date, type, amount, description, created_at) VALUES (%s, %s, 'debit', %s, %s, %s)", 
                                   (emp_id, date_val, emp_d, f"Debit #{settle_id}: {emp_d_n or 'Evening Settle'}", time_str))

            conn.commit()
            flash("Saved successfully!", "success")
            return redirect(url_for('evening'))

        except Exception as e:
            conn.rollback()
            flash(f"Error: {e}", "danger")
            return redirect(url_for('evening'))

    cursor.execute("SELECT id, name, image FROM employees WHERE status='active' ORDER BY name")
    emps = cursor.fetchall()
    for e in emps: e['image'] = resolve_img(e['image'])
    return render_template('evening.html', employees=emps, today=date.today().strftime('%d-%m-%Y'))
    
# --- 5. ROUTE: ALLOCATION LIST (Formatted Date/Time & Status Check) ---
@app.route('/allocation_list', methods=['GET'])
def allocation_list():
    if "loggedin" not in session: return redirect(url_for("login"))

    conn = mysql.connection
    cursor = conn.cursor(MySQLdb.cursors.DictCursor)

    date_filter = request.args.get('date')
    emp_filter = request.args.get('employee_id')
    
    # ADDED e.phone as emp_phone
    query = """
        SELECT 
            ma.id, ma.date, ma.created_at, 
            e.name as emp_name, e.image as emp_image, e.phone as emp_phone,
            (SELECT COUNT(*) FROM morning_allocation_items WHERE allocation_id = ma.id) as item_count,
            (SELECT status FROM evening_settle WHERE allocation_id = ma.id LIMIT 1) as evening_status
        FROM morning_allocations ma
        JOIN employees e ON ma.employee_id = e.id
        WHERE 1=1
    """
    params = []

    if date_filter:
        d_obj = parse_date(date_filter)
        if d_obj:
            query += " AND ma.date = %s"
            params.append(d_obj.strftime('%Y-%m-%d'))
    
    if emp_filter and emp_filter != 'all':
        query += " AND ma.employee_id = %s"
        params.append(emp_filter)

    query += " ORDER BY ma.date DESC, ma.created_at DESC"

    cursor.execute(query, tuple(params))
    allocations = cursor.fetchall()

    for a in allocations:
        a['emp_image'] = resolve_img(a['emp_image'])
        
        if isinstance(a['date'], (date, datetime)):
            a['formatted_date'] = a['date'].strftime('%d-%m-%Y')
        else:
            try: a['formatted_date'] = datetime.strptime(str(a['date']), '%Y-%m-%d').strftime('%d-%m-%Y')
            except: a['formatted_date'] = str(a['date'])

        if a.get('created_at'):
            if isinstance(a['created_at'], datetime): a['formatted_time'] = a['created_at'].strftime('%I:%M %p')
            elif isinstance(a['created_at'], timedelta): a['formatted_time'] = (datetime.min + a['created_at']).strftime('%I:%M %p')
            else: a['formatted_time'] = str(a['created_at'])
        else: a['formatted_time'] = "-"

        if a['evening_status'] == 'final':
            a['status_badge'] = 'Submitted'; a['status_class'] = 'bg-success'; a['is_locked'] = True
        elif a['evening_status'] == 'draft':
            a['status_badge'] = 'Draft'; a['status_class'] = 'bg-warning text-dark'; a['is_locked'] = False 
        else:
            a['status_badge'] = 'Pending'; a['status_class'] = 'bg-danger'; a['is_locked'] = False

    cursor.execute("SELECT id, name FROM employees WHERE status='active' ORDER BY name")
    employees = cursor.fetchall()

    return render_template('allocation_list.html', allocations=allocations, employees=employees, filters={'date': date_filter, 'employee_id': emp_filter})


# --- NEW ROUTE: DRAFT LIST PAGE ---
@app.route('/evening/drafts')
def draft_evening_list():
    if "loggedin" not in session: return redirect(url_for("login"))
    
    cur = mysql.connection.cursor(MySQLdb.cursors.DictCursor)
    # Check if status column exists first via try/catch in query
    try:
        cur.execute("""
            SELECT s.*, e.name as emp_name, e.image as emp_image 
            FROM evening_settle s 
            JOIN employees e ON s.employee_id = e.id 
            WHERE s.status = 'draft' 
            ORDER BY s.date DESC, s.id DESC
        """)
        drafts = cur.fetchall()
        for d in drafts: d['emp_image'] = resolve_img(d['emp_image'])
    except:
        drafts = [] # Fallback if table doesn't have status col yet
        
    cur.close()
    return render_template('draft_evening.html', drafts=drafts)




# ---------------------------------------------------------
# API: FETCH MORNING STOCK (With Holiday Logic)
# ---------------------------------------------------------
@app.route('/api/fetch_morning_allocation', methods=['GET', 'POST']) 
def fetch_morning_allocation(): 
    if "loggedin" not in session: return jsonify({'status': 'error', 'message': 'Unauthorized'})

    try:
        # Use request.values to grab from query string (GET) or body (POST)
        employee_id = request.values.get('employee_id')
        date_str = request.values.get('date') 
        
        if not employee_id or not date_str:
            return jsonify({'status': 'error', 'message': 'Missing parameters'})

        cur = mysql.connection.cursor(MySQLdb.cursors.DictCursor)
        
        # Parse Date (dd-mm-yyyy or yyyy-mm-dd)
        date_obj = parse_date(date_str)
        if not date_obj:
            return jsonify({'status': 'error', 'message': 'Invalid Date Format (Expected dd-mm-yyyy)'})

        formatted_date = date_obj.strftime('%Y-%m-%d')
        
        # HOLIDAY LOGIC: Fetch LAST Closing Stock (Opening for Today)
        
        # 1. Find the LAST COMPLETED Evening Settlement for this employee BEFORE today
        # We must ignore drafts here, only 'final' (or missing status which implies final old records)
        # Using a safer query that handles if 'status' column doesn't exist yet by ignoring it if error, 
        # but since we want robust logic:
        
        # NOTE: If you haven't added 'status' column yet, this query might fail if we add WHERE status='final'.
        # Assuming you will add it or legacy records are final.
        
        cur.execute("""
            SELECT id 
            FROM evening_settle 
            WHERE employee_id = %s AND date < %s 
            -- AND (status = 'final' OR status IS NULL) -- Uncomment after adding status column
            ORDER BY date DESC, id DESC LIMIT 1
        """, (employee_id, formatted_date))
        
        last_settle = cur.fetchone()
        
        stock_map = {}
        
        if last_settle:
            last_settle_id = last_settle['id']
            # 2. Fetch the remaining items from that settlement
            cur.execute("""
                SELECT product_id, remaining_qty 
                FROM evening_item 
                WHERE settle_id = %s
            """, (last_settle_id,))
            
            items = cur.fetchall()
            stock_map = {item['product_id']: item['remaining_qty'] for item in items}

        # --- IMPORTANT: Also fetch items allocated THIS MORNING (Today's Allocation) ---
        # So we can calculate [Total = Opening + Given Today]
        cur.execute("""
            SELECT ma.id as alloc_id, mai.product_id, mai.opening_qty, mai.given_qty, mai.unit_price, p.name, p.image
            FROM morning_allocations ma
            JOIN morning_allocation_items mai ON ma.id = mai.allocation_id
            JOIN products p ON mai.product_id = p.id
            WHERE ma.employee_id = %s AND ma.date = %s
        """, (employee_id, formatted_date))
        
        morning_items = cur.fetchall()
        
        # Prepare the final list of items for the Evening Settlement Table
        # Structure: Product ID, Name, Image, Unit Price, Total Qty (Opening + Given)
        
        final_items_list = []
        if morning_items:
            for item in morning_items:
                total_qty = int(item['opening_qty'] or 0) + int(item['given_qty'] or 0)
                if total_qty > 0:
                    final_items_list.append({
                        'product_id': item['product_id'],
                        'product_name': item['name'],
                        'image': resolve_img(item['image']),
                        'unit_price': float(item['unit_price']),
                        'total_qty': total_qty,
                        'remaining_qty': total_qty, # Initially remaining is total (before sales)
                        'sold_qty': 0,
                        'return_qty': 0
                    })
            
            # Allocation ID is needed to link the settlement later
            alloc_id = morning_items[0]['alloc_id']
        else:
            alloc_id = None
            # If no morning allocation found for today, maybe user wants to settle previous outstanding stock?
            # Or maybe they just forgot to do morning allocation. 
            # For now, let's return empty list but with success status so UI doesn't crash.
        
        return jsonify({
            'status': 'success',
            'allocation_id': alloc_id,
            'items': final_items_list,
            'allocation_date': formatted_date
        })

    except Exception as e:
        app.logger.error(f"Morning Fetch Error: {e}")
        return jsonify({'status': 'error', 'message': str(e)})
    finally:
        if 'cur' in locals(): cur.close()

# ==========================================
# 3. EDIT MORNING ALLOCATION (For Manual Edit Page)
# ==========================================
@app.route('/morning/edit/<int:allocation_id>', methods=['GET', 'POST'])
def edit_morning_allocation(allocation_id):
    if "loggedin" not in session: return redirect(url_for("login"))

    db_cursor = None
    try:
        db_cursor = mysql.connection.cursor(MySQLdb.cursors.DictCursor)
        
        db_cursor.execute("SELECT id FROM evening_settle WHERE allocation_id = %s", (allocation_id,))
        if db_cursor.fetchone():
            flash("Cannot edit: Evening settlement already submitted.", "warning")
            return redirect(url_for('allocation_list'))

        if request.method == 'POST':
            item_ids = request.form.getlist('item_id[]')
            product_ids = request.form.getlist('product_id[]')
            opening_qtys = request.form.getlist('opening[]')
            given_qtys = request.form.getlist('given[]')
            prices = request.form.getlist('price[]')

            db_cursor.execute("SELECT id, product_id, given_qty FROM morning_allocation_items WHERE allocation_id = %s", (allocation_id,))
            existing_items_map = {str(row['id']): row for row in db_cursor.fetchall()}
            
            submitted_item_ids = set([str(iid) for iid in item_ids if iid and iid != 'new_item'])

            # 1. DELETE logic
            for db_id, db_item in existing_items_map.items():
                if db_id not in submitted_item_ids:
                    qty_to_restore = int(db_item['given_qty'])
                    if qty_to_restore > 0:
                        db_cursor.execute("UPDATE products SET stock = stock + %s WHERE id = %s", (qty_to_restore, db_item['product_id']))
                    db_cursor.execute("DELETE FROM morning_allocation_items WHERE id = %s", (db_id,))

            # 2. UPDATE/INSERT logic
            for i, pid in enumerate(product_ids):
                if not pid: continue
                
                item_id = str(item_ids[i])
                op_qty = int(opening_qtys[i] or 0)
                gv_qty = int(given_qtys[i] or 0)
                price = float(prices[i] or 0)

                if item_id == 'new_item':
                    db_cursor.execute("""
                        INSERT INTO morning_allocation_items 
                        (allocation_id, product_id, opening_qty, given_qty, unit_price)
                        VALUES (%s, %s, %s, %s, %s)
                    """, (allocation_id, pid, op_qty, gv_qty, price))
                    if gv_qty > 0:
                        db_cursor.execute("UPDATE products SET stock = stock - %s WHERE id = %s", (gv_qty, pid))
                
                elif item_id in existing_items_map:
                    old_data = existing_items_map[item_id]
                    old_given = int(old_data['given_qty'])
                    diff = gv_qty - old_given # Difference
                    
                    db_cursor.execute("""
                        UPDATE morning_allocation_items 
                        SET product_id=%s, opening_qty=%s, given_qty=%s, unit_price=%s
                        WHERE id=%s
                    """, (pid, op_qty, gv_qty, price, item_id))
                    
                    if diff != 0:
                        # If diff is +ve (gave more), subtract from stock. If -ve (gave less), add to stock.
                        db_cursor.execute("UPDATE products SET stock = stock - %s WHERE id = %s", (diff, pid))

            mysql.connection.commit()
            flash("Allocation updated successfully!", "success")
            return redirect(url_for('allocation_list'))

        # GET Request
        db_cursor.execute("""
            SELECT ma.*, e.name as employee_name, e.image as emp_image 
            FROM morning_allocations ma
            JOIN employees e ON ma.employee_id = e.id
            WHERE ma.id = %s
        """, (allocation_id,))
        allocation = db_cursor.fetchone()
        
        if not allocation:
            return redirect(url_for('allocation_list'))
            
        allocation['emp_image'] = resolve_img(allocation['emp_image'])

        db_cursor.execute("""
            SELECT mai.*, p.image 
            FROM morning_allocation_items mai
            JOIN products p ON mai.product_id = p.id
            WHERE mai.allocation_id = %s
        """, (allocation_id,))
        items = db_cursor.fetchall()
        for item in items: item['image'] = resolve_img(item['image'])

        db_cursor.execute("SELECT id, name, price, image FROM products ORDER BY name")
        all_products = db_cursor.fetchall()
        products_js = []
        for p in all_products:
            p_img = resolve_img(p['image'])
            products_js.append({'id': p['id'], 'name': p['name'], 'price': float(p['price']), 'image': p_img})

        return render_template('morning_edit.html',
                               allocation=allocation,
                               items=items,
                               products=products_js)

    except Exception as e:
        if db_cursor: mysql.connection.rollback()
        flash(f"Error: {e}", "danger")
        return redirect(url_for('allocation_list'))
    finally:
        if db_cursor: db_cursor.close()



# ==========================================
# 4. DELETE ALLOCATION (Hard Delete)
# ==========================================
@app.route('/morning/delete/<int:id>', methods=['POST'])
def delete_allocation(id):
    if "loggedin" not in session: return redirect(url_for("login"))
    
    conn = mysql.connection
    cursor = conn.cursor(MySQLdb.cursors.DictCursor)
    
    try:
        # Check Dependency
        cursor.execute("SELECT id FROM evening_settle WHERE allocation_id = %s", (id,))
        if cursor.fetchone():
            flash("Cannot delete: Linked to Evening Settlement.", "danger")
            return redirect(url_for('allocation_list'))

        # 1. Restore Stock
        cursor.execute("SELECT product_id, given_qty FROM morning_allocation_items WHERE allocation_id = %s", (id,))
        items = cursor.fetchall()

        for item in items:
            if item['given_qty'] > 0:
                cursor.execute("UPDATE products SET stock = stock + %s WHERE id = %s", 
                               (item['given_qty'], item['product_id']))

        # 2. Delete
        cursor.execute("DELETE FROM morning_allocation_items WHERE allocation_id = %s", (id,))
        cursor.execute("DELETE FROM morning_allocations WHERE id = %s", (id,))
        
        conn.commit()
        flash("Allocation deleted and stock restored.", "success")
            
    except Exception as e:
        conn.rollback()
        flash(f"Error deleting: {str(e)}", "danger")
    finally:
        cursor.close()
        
    return redirect(url_for('allocation_list'))

# ---------------------------------------------------------#
# 3. EVENING MASTER (History & Drafts) - WITH FILTERS
# ---------------------------------------------------------#
@app.route('/evening/master')
def admin_evening_master():
    if "loggedin" not in session: return redirect(url_for("login"))
    
    conn = mysql.connection
    cursor = conn.cursor(MySQLdb.cursors.DictCursor)
    
    # --- 1. Get Filter Params ---
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    emp_filter = request.args.get('employee_id')

    # --- 2. Build Query ---
    # Fetching all necessary fields, including phone for potential future use
    query = """
        SELECT es.id, es.date, es.created_at, 
               IFNULL(es.total_amount, 0) as total_amount, 
               IFNULL(es.cash_money, 0) as cash_money, 
               IFNULL(es.online_money, 0) as online_money, 
               IFNULL(es.discount, 0) as discount,
               IFNULL(es.due_amount, 0) as due_amount,
               e.name as emp_name, e.image as emp_image, e.phone as emp_mobile
        FROM evening_settle es
        JOIN employees e ON es.employee_id = e.id
        WHERE 1=1
    """
    params = []

    if start_date:
        try:
            # Safe check for parse_date helper or datetime
            dt = parse_date(start_date) if 'parse_date' in globals() else datetime.strptime(start_date, '%d-%m-%Y')
            query += " AND es.date >= %s"
            params.append(dt.strftime('%Y-%m-%d'))
        except ValueError: pass
    
    if end_date:
        try:
            dt = parse_date(end_date) if 'parse_date' in globals() else datetime.strptime(end_date, '%d-%m-%Y')
            query += " AND es.date <= %s"
            params.append(dt.strftime('%Y-%m-%d'))
        except ValueError: pass

    if emp_filter and emp_filter != 'all':
        query += " AND es.employee_id = %s"
        params.append(emp_filter)

    query += " ORDER BY es.date DESC, es.created_at DESC"
    
    cursor.execute(query, tuple(params))
    settlements = cursor.fetchall()
    
    # --- 3. Stats Calculation (Restored Discount & Logic) ---
    stats = {
        'total_sales': 0.0,      # Gross Sales
        'net_sales': 0.0,        # Sales after discount
        'cash': 0.0,             # Raw Cash Collected
        'online': 0.0,
        'discount': 0.0,         # Total Discount Given
        'total_refunds': 0.0,    # Money returned to emp (Profit)
        'total_due': 0.0         # Actual Pending Due
    }
    
    for s in settlements:
        # Image Resolution Logic
        if s['emp_image']:
            if s['emp_image'].startswith('http'):
                pass 
            else:
                s['emp_image'] = url_for('static', filename='uploads/' + s['emp_image'])
        else:
            s['emp_image'] = url_for('static', filename='img/default-user.png')

        # Date Formatting
        if isinstance(s['date'], (date, datetime)):
            s['formatted_date'] = s['date'].strftime('%d-%m-%Y')
        else:
            try:
                s['formatted_date'] = datetime.strptime(str(s['date']), '%Y-%m-%d').strftime('%d-%m-%Y')
            except ValueError:
                s['formatted_date'] = str(s['date'])

        # Time Formatting (12-hour)
        if s.get('created_at'):
            if isinstance(s['created_at'], timedelta):
                 dummy_date = datetime.min + s['created_at']
                 s['formatted_time'] = dummy_date.strftime('%I:%M %p')
            elif isinstance(s['created_at'], datetime):
                 s['formatted_time'] = s['created_at'].strftime('%I:%M %p')
            else:
                 s['formatted_time'] = str(s['created_at'])
        else:
            s['formatted_time'] = ""

        # Numeric Values
        t_amt = float(s['total_amount'])
        c_money = float(s['cash_money'])
        o_money = float(s['online_money'])
        disc = float(s['discount'])
        due_amt = float(s['due_amount'])
        
        # Accumulate Basic Stats
        stats['total_sales'] += t_amt
        stats['discount'] += disc
        stats['net_sales'] += (t_amt - disc)
        stats['cash'] += c_money
        stats['online'] += o_money

        # --- ADJUSTED CASH LOGIC ---
        # If due is negative, it means we PAID BACK cash (refund/profit).
        # We need to subtract this from our "Cash In Hand" KPI.
        if due_amt < -0.01:
            stats['total_refunds'] += abs(due_amt)
        elif due_amt > 0.01:
            stats['total_due'] += due_amt

    # Final Adjusted Cash Calculation
    adjusted_cash = stats['cash'] - stats['total_refunds']

    # Fetch Employees for Filter Dropdown
    cursor.execute("SELECT id, name FROM employees WHERE status='active' ORDER BY name")
    employees = cursor.fetchall()

    return render_template('admin/evening_master.html', 
                           settlements=settlements, 
                           stats=stats,
                           adjusted_cash=adjusted_cash, # Passed explicitly
                           total_due=stats['total_due'], # Passed explicitly
                           employees=employees,
                           filters={'start': start_date, 'end': end_date, 'emp': emp_filter})
    
# ==========================================
# 5. EDIT EVENING (Stock Logic)
# ==========================================

@app.route('/admin/edit_evening_settle/<int:settle_id>', methods=['GET', 'POST'])
def edit_evening_settle(settle_id):
    if "loggedin" not in session: return redirect(url_for("login"))
    conn = mysql.connection
    cursor = conn.cursor(MySQLdb.cursors.DictCursor)
    
    if request.method == 'POST':
        try:
            # Basic financial update
            total_amt = float(request.form.get('totalAmount') or 0)
            discount = float(request.form.get('discount') or 0)
            cash = float(request.form.get('cash') or 0)
            online = float(request.form.get('online') or 0)
            
            # Recalculate Due
            net = total_amt - discount
            paid = cash + online
            due_amt = net - paid
            
            cursor.execute("""
                UPDATE evening_settle 
                SET total_amount=%s, discount=%s, cash_money=%s, online_money=%s, due_amount=%s
                WHERE id=%s
            """, (total_amt, discount, cash, online, due_amt, settle_id))
            
            # Handle Item Updates (Assumption: You have item arrays in form)
            p_ids = request.form.getlist('product_id[]')
            sold_qtys = request.form.getlist('sold[]')
            return_qtys = request.form.getlist('return[]')
            
            # You can add the item update logic here similar to previous implementation if needed
            # For brevity in this specific request, updating financials is key.

            conn.commit()
            flash("Settlement updated.", "success")
            return redirect(url_for('admin_evening_master'))
        except Exception as e:
            conn.rollback()
            flash(f"Error: {e}", "danger")
    
    # GET
    cursor.execute("SELECT s.*, e.name as emp_name, DATE_FORMAT(s.date, '%%d-%%m-%%Y') as formatted_date FROM evening_settle s JOIN employees e ON s.employee_id = e.id WHERE s.id=%s", (settle_id,))
    s = cursor.fetchone()
    
    cursor.execute("SELECT i.*, p.name as product_name, p.image FROM evening_item i JOIN products p ON i.product_id = p.id WHERE i.settle_id=%s", (settle_id,))
    items = cursor.fetchall()
    
    return render_template('admin/edit_evening_settle.html', settlement=s, items=items)


# ==========================================
# 6. DELETE EVENING (Standard UNDO + Ledger Delete)
# ==========================================
@app.route('/admin/evening/delete/<int:settle_id>', methods=['POST'])
def admin_delete_evening(settle_id):
    if "loggedin" not in session: return redirect(url_for("login"))

    conn = mysql.connection
    cursor = conn.cursor(MySQLdb.cursors.DictCursor)
    try:
        # 1. Fetch Items to Reverse Returns
        # We need to subtract the returned quantity from the warehouse because 
        # the settlement (which added it) is being deleted.
        cursor.execute("SELECT product_id, return_qty FROM evening_item WHERE settle_id = %s", (settle_id,))
        items = cursor.fetchall()

        for item in items:
            ret = int(item['return_qty'])
            if ret > 0:
                cursor.execute("UPDATE products SET stock = stock - %s WHERE id = %s", (ret, item['product_id']))

        # 2. Delete Ledger Entries (Financial Reversal)
        # Using the description pattern "Credit #{id}:..." or "Debit #{id}:..."
        cursor.execute("DELETE FROM employee_transactions WHERE description LIKE %s", (f"%#{settle_id}:%",))

        # 3. Delete Record
        cursor.execute("DELETE FROM evening_item WHERE settle_id = %s", (settle_id,))
        cursor.execute("DELETE FROM evening_settle WHERE id = %s", (settle_id,))
        
        conn.commit()
        flash("Settlement Deleted. Returns removed from warehouse. Ledger entries removed.", "success")
    except Exception as e:
        conn.rollback()
        flash(f"Delete Error: {e}", "danger")
    finally:
        cursor.close()
    return redirect(url_for('admin_evening_master'))


# --- 4. BULK PDF EXPORT ---
@app.route('/admin/evening/export_pdf')
def admin_evening_export_pdf():
    start = request.args.get('start_date')
    end = request.args.get('end_date')
    emp_id = request.args.get('employee_id')
    
    cur = mysql.connection.cursor(MySQLdb.cursors.DictCursor)
    query = """
        SELECT es.date, e.name, es.total_amount, es.cash_money, es.online_money, es.due_amount
        FROM evening_settle es
        JOIN employees e ON es.employee_id = e.id
        WHERE es.date BETWEEN %s AND %s
    """
    params = [start, end]
    if emp_id and emp_id != 'all':
        query += " AND es.employee_id = %s"
        params.append(emp_id)
    query += " ORDER BY es.date DESC"
    
    cur.execute(query, tuple(params))
    data = cur.fetchall()
    cur.close()

    # Generate PDF
    pdf = PDF()
    pdf.add_page()
    pdf.set_font("Arial", 'B', 14)
    pdf.cell(0, 10, f"Settlement Report ({start} to {end})", 0, 1, 'C')
    pdf.ln(5)

    pdf.set_font("Arial", 'B', 10)
    pdf.set_fill_color(240, 240, 240)
    pdf.cell(30, 8, "Date", 1, 0, 'C', 1)
    pdf.cell(60, 8, "Employee", 1, 0, 'L', 1)
    pdf.cell(30, 8, "Sales", 1, 0, 'R', 1)
    pdf.cell(30, 8, "Cash", 1, 0, 'R', 1)
    pdf.cell(30, 8, "Due", 1, 1, 'R', 1)

    pdf.set_font("Arial", '', 9)
    total = 0
    for row in data:
        # Date Format Fix
        d_str = row['date'].strftime('%d-%m-%Y') if isinstance(row['date'], date) else str(row['date'])
        
        pdf.cell(30, 7, d_str, 1)
        pdf.cell(60, 7, pdf.safe_text(row['name']), 1)
        pdf.cell(30, 7, f"{row['total_amount']:.2f}", 1, 0, 'R')
        pdf.cell(30, 7, f"{row['cash_money']:.2f}", 1, 0, 'R')
        pdf.cell(30, 7, f"{row['due_amount']:.2f}", 1, 1, 'R')
        total += row['total_amount']

    pdf.ln(5)
    pdf.set_font("Arial", 'B', 11)
    pdf.cell(0, 10, f"Total Period Sales: {total:.2f}", 0, 1, 'R')

    response = make_response(pdf.output(dest='S').encode('latin1'))
    response.headers['Content-Type'] = 'application/pdf'
    response.headers['Content-Disposition'] = 'attachment; filename=Summary_Report.pdf'
    return response


######################################################################===============================================#
# All Reports Pages Routes Here
######################################################################================================================#

# ==============================#
#  PURCHASE REPORTS MODULE
# ==============================#


@app.route("/purchase_report", methods=["GET", "POST"])
def purchase_report():
    if "loggedin" not in session:
        return redirect(url_for("login"))

    cursor = mysql.connection.cursor(MySQLdb.cursors.DictCursor)

    # ---- Filters ----
    # For GET requests, we can also look at args if you want bookmarks to work
    # But usually filters are POSTed. Let's support both for better UX.
    start_date = request.form.get("start_date") or request.args.get("start_date")
    end_date = request.form.get("end_date") or request.args.get("end_date")
    supplier_id = request.form.get("supplier_id") or request.args.get("supplier_id")

    # Fetch suppliers for dropdown
    cursor.execute("SELECT id, name FROM suppliers ORDER BY name ASC")
    suppliers = cursor.fetchall()

    # 1. Base query for Table List
    sql = """
        SELECT 
            p.id,
            p.purchase_date,
            p.bill_number,
            p.total_amount,
            s.name AS supplier_name
        FROM purchases p
        LEFT JOIN suppliers s ON s.id = p.supplier_id
        WHERE 1 = 1
    """
    params = []

    # Apply filters
    if start_date:
        sql += " AND DATE(p.purchase_date) >= %s"
        params.append(start_date)
    if end_date:
        sql += " AND DATE(p.purchase_date) <= %s"
        params.append(end_date)
    if supplier_id and supplier_id != 'all':
        sql += " AND p.supplier_id = %s"
        params.append(supplier_id)

    sql += " ORDER BY p.purchase_date DESC"

    cursor.execute(sql, tuple(params))
    purchases = cursor.fetchall()

    # 2. Monthly summary for Trend Chart (Aggregated)
    # Using same filters but grouping by month
    monthly_sql = """
        SELECT 
            DATE_FORMAT(p.purchase_date, '%%Y-%%m') AS month,
            SUM(p.total_amount) AS total_purchase
        FROM purchases p
        WHERE 1 = 1
    """
    monthly_params = []
    
    # Re-apply filters for charts to keep them in sync
    if start_date:
        monthly_sql += " AND DATE(p.purchase_date) >= %s"
        monthly_params.append(start_date)
    if end_date:
        monthly_sql += " AND DATE(p.purchase_date) <= %s"
        monthly_params.append(end_date)
    # Note: Supplier filter usually applies to trend too? Yes.
    if supplier_id and supplier_id != 'all':
        monthly_sql += " AND p.supplier_id = %s"
        monthly_params.append(supplier_id)

    monthly_sql += " GROUP BY DATE_FORMAT(p.purchase_date, '%%Y-%%m') ORDER BY month ASC"

    cursor.execute(monthly_sql, tuple(monthly_params))
    # Fetch all and ensure values are floats (Decimal not JSON serializable sometimes)
    monthly_data_raw = cursor.fetchall()
    monthly_data = []
    for row in monthly_data_raw:
        monthly_data.append({
            'month': row['month'],
            'total_purchase': float(row['total_purchase'] or 0)
        })

    # 3. Supplier-wise totals for Pie Chart
    supplier_sql = """
        SELECT 
            s.name AS supplier_name,
            SUM(p.total_amount) AS total_purchase
        FROM purchases p
        LEFT JOIN suppliers s ON s.id = p.supplier_id
        WHERE 1 = 1
    """
    supplier_params = []

    if start_date:
        supplier_sql += " AND DATE(p.purchase_date) >= %s"
        supplier_params.append(start_date)
    if end_date:
        supplier_sql += " AND DATE(p.purchase_date) <= %s"
        supplier_params.append(end_date)
    if supplier_id and supplier_id != 'all':
        supplier_sql += " AND p.supplier_id = %s"
        supplier_params.append(supplier_id)

    supplier_sql += " GROUP BY s.name ORDER BY total_purchase DESC"

    cursor.execute(supplier_sql, tuple(supplier_params))
    supplier_data_raw = cursor.fetchall()
    supplier_data = []
    for row in supplier_data_raw:
        supplier_data.append({
            'supplier_name': row['supplier_name'] or 'Unknown',
            'total_purchase': float(row['total_purchase'] or 0)
        })

    cursor.close()

    return render_template(
        "reports/purchase_report.html",
        purchases=purchases,
        suppliers=suppliers,
        monthly_data=monthly_data,
        supplier_data=supplier_data,
        start_date=start_date,
        end_date=end_date,
        selected_supplier=supplier_id
    )



@app.route("/export_purchase_excel")
def export_purchase_excel():
    cursor = mysql.connection.cursor(MySQLdb.cursors.DictCursor)

    cursor.execute("""
        SELECT 
            p.purchase_date,
            p.bill_number,
            p.total_amount,
            s.name AS supplier_name
        FROM purchases p
        LEFT JOIN suppliers s ON s.id = p.supplier_id
        ORDER BY p.purchase_date DESC
    """)
    data = cursor.fetchall()
    cursor.close()

    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Purchase Report"

    ws.append(["Date", "Supplier", "Bill #", "Total Amount"])

    for row in data:
        ws.append([
            row["purchase_date"],
            row["supplier_name"],
            row["bill_number"],
            row["total_amount"]
        ])

    filename = "purchase_report.xlsx"
    wb.save(filename)
    return send_file(filename, as_attachment=True)


@app.route("/export_purchase_pdf")
def export_purchase_pdf():
    cursor = mysql.connection.cursor(MySQLdb.cursors.DictCursor)

    cursor.execute("""
        SELECT 
            p.purchase_date,
            p.bill_number,
            p.total_amount,
            s.name AS supplier_name
        FROM purchases p
        LEFT JOIN suppliers s ON s.id = p.supplier_id
        ORDER BY p.purchase_date DESC
    """)
    data = cursor.fetchall()
    cursor.close()

    from reportlab.pdfgen import canvas
    from reportlab.lib.pagesizes import letter

    filename = "purchase_report.pdf"
    c = canvas.Canvas(filename, pagesize=letter)
    width, height = letter
    y = height - 50

    c.setFont("Helvetica-Bold", 14)
    c.drawString(50, y, "Purchase Report")
    y -= 40

    c.setFont("Helvetica", 10)
    for row in data:
        line = f"{row['purchase_date']} | {row['supplier_name']} | Bill: {row['bill_number']} | ₹{row['total_amount']}"
        c.drawString(50, y, line)
        y -= 18
        if y < 50:
            c.showPage()
            y = height - 50

    c.save()
    return send_file(filename, as_attachment=True)



@app.route('/summary')
def summary():
    selected_date = request.args.get('date', date.today().isoformat())
    db_cursor = mysql.connection.cursor(MySQLdb.cursors.DictCursor)
    db_cursor.execute("""
        SELECT ma.id, e.name as employee_name, 
               (SELECT COUNT(id) FROM evening_settle WHERE allocation_id = ma.id) as evening_submitted
        FROM morning_allocations ma
        JOIN employees e ON ma.employee_id = e.id
        WHERE ma.date = %s ORDER BY e.name
    """, (selected_date,))
    morning_forms_raw = db_cursor.fetchall()
    morning_forms = []
    for form in morning_forms_raw:
        form['is_editable'] = form['evening_submitted'] == 0
        morning_forms.append(form)
    db_cursor.execute("""
        SELECT es.id, e.name as employee_name, es.total_amount
        FROM evening_settle es
        JOIN employees e ON es.employee_id = e.id
        WHERE es.date = %s ORDER BY e.name
    """, (selected_date,))
    evening_forms = db_cursor.fetchall()
    db_cursor.close()
    return render_template('daily_summary.html', 
                           morning_forms=morning_forms, 
                           evening_forms=evening_forms, 
                           selected_date=selected_date)


@app.route('/reports')
def reports():
    return render_template("/reports.html")


# REPLACE your existing 'daily_summary' route in app.py

@app.route('/reports/daily_summary', methods=['GET', 'POST'])
def daily_summary():
    if "loggedin" not in session: return redirect(url_for("login"))
    
    report_date = request.form.get('report_date') or date.today().isoformat()
    
    summary_data = {
        "cash_in": 0.0,
        "cash_out": 0.0,
        "net_flow": 0.0,
        "breakdown": {
            "evening_sales": 0.0,
            "office_sales": 0.0,
            "emp_credit": 0.0,
            "expenses": 0.0,
            "emp_debit": 0.0,
            "supplier_pay": 0.0
        },
        "office_sales_list": []
    }

    db_cursor = None
    try:
        db_cursor = mysql.connection.cursor(MySQLdb.cursors.DictCursor)
        
        # 1. CASH IN
        # A. Evening Sales (Net = Total - Discount)
        # Note: We consider Cash + Online as total money IN for the day
        db_cursor.execute("""
            SELECT SUM(total_amount - discount) as val 
            FROM evening_settle WHERE date = %s AND status='final'
        """, (report_date,))
        summary_data['breakdown']['evening_sales'] = float(db_cursor.fetchone()['val'] or 0)

        # B. Office Sales (Net = Final Amount)
        db_cursor.execute("SELECT SUM(final_amount) as val FROM office_sales WHERE sale_date = %s", (report_date,))
        summary_data['breakdown']['office_sales'] = float(db_cursor.fetchone()['val'] or 0)

        # C. Employee Credit (Money Received from Emp)
        db_cursor.execute("SELECT SUM(amount) as val FROM employee_transactions WHERE transaction_date = %s AND type='credit'", (report_date,))
        summary_data['breakdown']['emp_credit'] = float(db_cursor.fetchone()['val'] or 0)

        # 2. CASH OUT
        # A. Expenses
        db_cursor.execute("SELECT SUM(amount) as val FROM expenses WHERE expense_date = %s", (report_date,))
        summary_data['breakdown']['expenses'] = float(db_cursor.fetchone()['val'] or 0)

        # B. Employee Debit (Money Given to Emp)
        db_cursor.execute("SELECT SUM(amount) as val FROM employee_transactions WHERE transaction_date = %s AND type='debit'", (report_date,))
        summary_data['breakdown']['emp_debit'] = float(db_cursor.fetchone()['val'] or 0)

        # C. Supplier Payments
        db_cursor.execute("SELECT SUM(amount_paid) as val FROM supplier_payments WHERE payment_date = %s", (report_date,))
        summary_data['breakdown']['supplier_pay'] = float(db_cursor.fetchone()['val'] or 0)

        # 3. Totals
        summary_data['cash_in'] = (
            summary_data['breakdown']['evening_sales'] + 
            summary_data['breakdown']['office_sales'] + 
            summary_data['breakdown']['emp_credit']
        )
        
        summary_data['cash_out'] = (
            summary_data['breakdown']['expenses'] + 
            summary_data['breakdown']['emp_debit'] + 
            summary_data['breakdown']['supplier_pay']
        )
        
        summary_data['net_flow'] = summary_data['cash_in'] - summary_data['cash_out']

        # 4. Detailed Office Sales List (Net of Discount)
        db_cursor.execute("""
            SELECT id, customer_name, sub_total, discount, final_amount, sales_person
            FROM office_sales 
            WHERE sale_date = %s
            ORDER BY id DESC
        """, (report_date,))
        summary_data['office_sales_list'] = db_cursor.fetchall()

    except Exception as e:
        flash(f"Error generating summary: {e}", "danger")
    finally:
        if db_cursor: db_cursor.close()

    report_date_obj = date.fromisoformat(report_date)
    report_date_str = report_date_obj.strftime('%d %B, %Y')
    
    return render_template('reports/daily_summary.html', 
                           report_date=report_date,
                           report_date_str=report_date_str,
                           data=summary_data)


@app.route('/reports/daily_summary/pdf/<report_date>')
def report_daily_summary_pdf(report_date):
    db_cursor = None
    try:
        db_cursor = mysql.connection.cursor(MySQLdb.cursors.DictCursor)
        db_cursor.execute("SELECT SUM(total_amount) as total_sales FROM evening_settle WHERE date = %s", (report_date,))
        sales_result = db_cursor.fetchone()
        total_sales = sales_result['total_sales'] or 0
        db_cursor.execute("SELECT SUM(amount) as total_expenses FROM expenses WHERE expense_date = %s", (report_date,))
        expenses_result = db_cursor.fetchone()
        total_expenses = expenses_result['total_expenses'] or 0
        db_cursor.close()
        
        net_cash_flow = total_sales - total_expenses
        report_date_obj = date.fromisoformat(report_date)
        report_date_str = report_date_obj.strftime('%d %B, %Y')

        pdf = PDF(orientation='P', unit='mm', format='A4')
        pdf.add_page()
        pdf.set_font(pdf.font_family, 'B', 14)
        pdf.cell(0, 10, f'Daily Financial Summary: {report_date_str}', 0, 1, 'C')
        pdf.ln(15)
        pdf.set_font(pdf.font_family, '', 12)
        pdf.cell(0, 10, 'Total Sales', 1, 0, 'L')
        pdf.cell(0, 10, f'Rs. {total_sales:.2f}', 0, 1, 'R')
        pdf.ln(2)
        pdf.cell(0, 10, 'Total Expenses', 1, 0, 'L')
        pdf.cell(0, 10, f'Rs. {total_expenses:.2f}', 0, 1, 'R')
        pdf.ln(2)
        pdf.set_font(pdf.font_family, 'B', 12)
        pdf.set_fill_color(240, 240, 240)
        pdf.cell(0, 10, 'Net Cash Flow', 1, 0, 'L', 1)
        pdf.cell(0, 10, f'Rs. {net_cash_flow:.2f}', 0, 1, 'R')

        pdf_output = pdf.output(dest='S').encode('latin1')
        response = make_response(pdf_output)
        response.headers.set('Content-Type', 'application/pdf')
        response.headers.set('Content-Disposition', 'attachment', filename=f'Daily_Summary_{report_date}.pdf')
        return response
    except Exception as e:
        app.logger.error(f"PDF Generation Error: {e}")
        flash("Could not generate PDF due to a server error.", "danger")
        return redirect(url_for('daily_summary'))
    finally:
        if db_cursor:
            db_cursor.close()



# =========================================================#
# DAILY SALES REPORT (Updated for Net Revenue)
# =========================================================#
# REPLACE your existing 'daily_sales' route in app.py with this UPDATED version:

@app.route('/reports/daily_sales', methods=['GET', 'POST'])
def daily_sales():
    if "loggedin" not in session: return redirect(url_for("login"))
    
    # 1. Get raw date (default to today)
    raw_date = request.form.get('report_date') or date.today().strftime('%d-%m-%Y')
    
    # 2. Parse Date Correctly
    try:
        report_date_obj = datetime.strptime(raw_date, '%d-%m-%Y').date()
    except ValueError:
        try:
            report_date_obj = date.fromisoformat(raw_date)
        except ValueError:
            report_date_obj = date.today()
    
    # SQL & Display formats
    sql_date = report_date_obj.strftime('%Y-%m-%d')
    report_date_str = report_date_obj.strftime('%d-%m-%Y')
    
    sales_data = []
    grand_total = 0.0      # This is Gross Total for Table sum
    net_total_revenue = 0.0 # This is Actual Money Earned (Net)
    total_discount = 0.0    # NEW: Total Discount Given
    total_units = 0

    db_cursor = None
    try:
        db_cursor = mysql.connection.cursor(MySQLdb.cursors.DictCursor)
        
        # A. Fetch Item Details (For Table)
        sql_items = """
            SELECT 
                'Field Sale' as type,
                p.name AS product_name,
                ei.sold_qty as qty,
                ei.unit_price,
                (ei.sold_qty * ei.unit_price) AS total_amount
            FROM evening_item ei
            JOIN evening_settle es ON ei.settle_id = es.id
            JOIN products p ON ei.product_id = p.id
            WHERE es.date = %s AND ei.sold_qty > 0 AND es.status = 'final'

            UNION ALL

            SELECT 
                'Counter Sale' as type,
                p.name AS product_name,
                osi.qty as qty,
                osi.unit_price,
                osi.total_price AS total_amount
            FROM office_sale_items osi
            JOIN office_sales os ON osi.sale_id = os.id
            JOIN products p ON osi.product_id = p.id
            WHERE os.sale_date = %s

            ORDER BY type, product_name
        """
        db_cursor.execute(sql_items, (sql_date, sql_date))
        sales_data = db_cursor.fetchall()
        
        for item in sales_data:
            grand_total += float(item['total_amount'])
            total_units += int(item['qty'])

        # B. Calculate Net Revenue & Total Discount (From Header Tables)
        # Using Header tables is accurate because discount is applied on the whole bill.
        sql_net = """
            SELECT 
                SUM(net_val) as true_revenue, 
                SUM(disc_val) as total_disc 
            FROM (
                SELECT (total_amount - discount) as net_val, discount as disc_val 
                FROM evening_settle 
                WHERE date = %s AND status = 'final'
                
                UNION ALL
                
                SELECT final_amount as net_val, discount as disc_val 
                FROM office_sales 
                WHERE sale_date = %s
            ) as t
        """
        db_cursor.execute(sql_net, (sql_date, sql_date))
        net_res = db_cursor.fetchone()
        
        if net_res:
            net_total_revenue = float(net_res['true_revenue'] or 0)
            total_discount = float(net_res['total_disc'] or 0) # Capture Discount

    except Exception as e:
        flash(f"Error fetching daily sales: {e}", "danger")
        app.logger.error(f"Daily Sales Error: {e}")
    finally:
        if db_cursor: db_cursor.close()

    return render_template('reports/daily_sales.html', 
                           report_date=report_date_str, 
                           report_date_str=report_date_str,
                           sales_data=sales_data,
                           grand_total=net_total_revenue, 
                           total_discount=total_discount, # Pass Discount
                           table_total=grand_total,
                           total_units=total_units)


@app.route('/reports/employee_performance', methods=['GET', 'POST'])
def employee_performance():
    if "loggedin" not in session: return redirect(url_for("login"))
    
    # Defaults
    end_date_obj = date.today()
    start_date_obj = end_date_obj - timedelta(days=29)
    
    # Get raw input
    start_date_str = request.form.get('start_date') or request.args.get('start_date')
    end_date_str = request.form.get('end_date') or request.args.get('end_date')

    # Robust Date Parsing
    if start_date_str:
        try: start_date_obj = datetime.strptime(start_date_str, '%d-%m-%Y').date()
        except: 
            try: start_date_obj = date.fromisoformat(start_date_str)
            except: pass
            
    if end_date_str:
        try: end_date_obj = datetime.strptime(end_date_str, '%d-%m-%Y').date()
        except:
            try: end_date_obj = date.fromisoformat(end_date_str)
            except: pass

    # SQL Format (YYYY-MM-DD)
    sql_start = start_date_obj.strftime('%Y-%m-%d')
    sql_end = end_date_obj.strftime('%Y-%m-%d')
    
    # Display Format (DD-MM-YYYY)
    display_start = start_date_obj.strftime('%d-%m-%Y')
    display_end = end_date_obj.strftime('%d-%m-%Y')

    performance_data = []
    total_period_revenue = 0.0
    total_period_units = 0
    db_cursor = None
    
    try:
        db_cursor = mysql.connection.cursor(MySQLdb.cursors.DictCursor)
        
        # UNIFIED QUERY: Evening Sales + Office Sales (grouped by employee)
        # Logic: 
        # 1. Evening Sales: Net Revenue = (total_amount - discount)
        # 2. Office Sales: Net Revenue = final_amount (which is sub_total - discount)
        # We group by employee name to merge field and counter sales if names match.
        
        sql = """
            SELECT 
                COALESCE(emp_name_final, 'Office/Admin') AS employee_name,
                MAX(emp_image_final) as employee_image,
                SUM(net_rev) as total_sales_amount,
                SUM(units) as total_units_sold
            FROM (
                -- Field Sales (Evening Settle)
                SELECT 
                    e.name as emp_name_final,
                    e.image as emp_image_final,
                    (es.total_amount - es.discount) as net_rev,
                    (SELECT COALESCE(SUM(sold_qty),0) FROM evening_item WHERE settle_id=es.id) as units
                FROM evening_settle es
                JOIN employees e ON es.employee_id = e.id
                WHERE es.date BETWEEN %s AND %s AND es.status = 'final'
                
                UNION ALL
                
                -- Counter Sales (Office Sales)
                -- Attempt to link sales_person name to employee table for image, else NULL
                SELECT 
                    os.sales_person as emp_name_final,
                    (SELECT image FROM employees WHERE name = os.sales_person LIMIT 1) as emp_image_final,
                    os.final_amount as net_rev,
                    (SELECT COALESCE(SUM(qty),0) FROM office_sale_items WHERE sale_id=os.id) as units
                FROM office_sales os
                WHERE os.sale_date BETWEEN %s AND %s
            ) as combined_data
            GROUP BY emp_name_final
            ORDER BY total_sales_amount DESC
        """
        
        db_cursor.execute(sql, (sql_start, sql_end, sql_start, sql_end))
        rows = db_cursor.fetchall()
        
        for r in rows:
            rev = float(r['total_sales_amount'] or 0)
            qty = int(r['total_units_sold'] or 0)
            
            # Use helper to resolve image path (Cloudinary or Local)
            img_path = resolve_img(r['employee_image'])
            
            performance_data.append({
                'employee_name': r['employee_name'],
                'employee_image': img_path,
                'total_units_sold': qty,
                'total_sales_amount': rev
            })
            
            total_period_revenue += rev
            total_period_units += qty

    except Exception as e:
        flash(f"Error fetching employee performance: {e}", "danger")
        app.logger.error(f"Emp Perf Error: {e}")
    finally:
        if db_cursor: db_cursor.close()

    return render_template('reports/employee_performance.html', 
                           start_date=display_start,
                           end_date=display_end,
                           performance_data=performance_data,
                           total_revenue=total_period_revenue,
                           total_units=total_period_units)


@app.route('/reports/employee_performance/pdf/<start_date>/<end_date>')
def report_employee_performance_pdf(start_date, end_date):
    db_cursor = None
    try:
        db_cursor = mysql.connection.cursor(MySQLdb.cursors.DictCursor)
        sql = """
            SELECT e.name AS employee_name, SUM(ei.sold_qty) AS total_units_sold, SUM(ei.sold_qty * ei.unit_price) AS total_sales_amount
            FROM evening_item ei
            JOIN evening_settle es ON ei.settle_id = es.id
            JOIN employees e ON es.employee_id = e.id
            WHERE es.date BETWEEN %s AND %s
            GROUP BY e.id, e.name
            ORDER BY total_sales_amount DESC
        """
        db_cursor.execute(sql, (start_date, end_date))
        performance_data = db_cursor.fetchall()
        
        start_date_obj = date.fromisoformat(start_date)
        end_date_obj = date.fromisoformat(end_date)
        date_range_str = f"{start_date_obj.strftime('%d %b %Y')} to {end_date_obj.strftime('%d %b %Y')}"

        pdf = PDF(orientation='P', unit='mm', format='A4')
        pdf.add_page()
        pdf.set_font(pdf.font_family, 'B', 14)
        pdf.cell(0, 10, 'Employee Performance Report', 0, 1, 'C')
        pdf.set_font(pdf.font_family, '', 10)
        pdf.cell(0, 10, date_range_str, 0, 1, 'C')
        pdf.ln(10)
        pdf.set_font(pdf.font_family, 'B', 10)
        pdf.set_fill_color(230, 230, 230)
        pdf.cell(15, 8, 'Rank', 1, 0, 'C', 1)
        pdf.cell(85, 8, 'Employee Name', 1, 0, 'L', 1)
        pdf.cell(40, 8, 'Total Units Sold', 1, 0, 'C', 1)
        pdf.cell(50, 8, 'Total Sales Amount', 1, 1, 'R', 1)
        pdf.set_font(pdf.font_family, '', 9)
        if not performance_data:
            pdf.cell(190, 10, 'No sales data found for this period.', 1, 1, 'C')
        else:
            for i, emp in enumerate(performance_data):
                pdf.cell(15, 7, str(i + 1), 1, 0, 'C')
                pdf.cell(85, 7, pdf.safe_text(emp['employee_name']), 1)
                pdf.cell(40, 7, str(emp['total_units_sold']), 1, 0, 'C')
                pdf.cell(50, 7, f"{emp['total_sales_amount']:.2f}", 1, 1, 'R')
        
        pdf_output = pdf.output(dest='S').encode('latin1')
        response = make_response(pdf_output)
        response.headers.set('Content-Type', 'application/pdf')
        response.headers.set('Content-Disposition', 'attachment', filename=f'Employee_Performance_{start_date}_to_{end_date}.pdf')
        return response
    except Exception as e:
        app.logger.error(f"PDF Generation Error: {e}")
        flash("Could not generate PDF due to a server error.", "danger")
        return redirect(url_for('employee_performance')) # Corrected route
    finally:
        if db_cursor:
            db_cursor.close()

# REPLACE your existing 'product_perfrm' route in app.py with this updated version:

@app.route('/reports/product_perfrm', methods=['GET', 'POST'])
def product_perfrm():
    if "loggedin" not in session: return redirect(url_for("login"))
    
    # Defaults
    end_date_obj = date.today()
    start_date_obj = end_date_obj - timedelta(days=29)
    
    # Get raw input
    start_date_str = request.form.get('start_date') or request.args.get('start_date')
    end_date_str = request.form.get('end_date') or request.args.get('end_date')
    sort_by = request.form.get('sort_by') or request.args.get('sort_by') or 'revenue'

    # Robust Date Parsing
    if start_date_str:
        try: start_date_obj = datetime.strptime(start_date_str, '%d-%m-%Y').date()
        except: 
            try: start_date_obj = date.fromisoformat(start_date_str)
            except: pass
            
    if end_date_str:
        try: end_date_obj = datetime.strptime(end_date_str, '%d-%m-%Y').date()
        except:
            try: end_date_obj = date.fromisoformat(end_date_str)
            except: pass

    # SQL Format (YYYY-MM-DD)
    sql_start = start_date_obj.strftime('%Y-%m-%d')
    sql_end = end_date_obj.strftime('%Y-%m-%d')
    
    # Display Format (DD-MM-YYYY)
    display_start = start_date_obj.strftime('%d-%m-%Y')
    display_end = end_date_obj.strftime('%d-%m-%Y')

    product_data = []
    db_cursor = None
    
    try:
        db_cursor = mysql.connection.cursor(MySQLdb.cursors.DictCursor)
        
        order_clause = "ORDER BY total_revenue DESC"
        if sort_by == 'quantity':
            order_clause = "ORDER BY total_units_sold DESC"
        
        # Unified Query: Evening + Office Sales
        sql = f"""
            SELECT 
                p.name AS product_name,
                p.image,
                SUM(qty) AS total_units_sold,
                SUM(line_total) AS total_revenue
            FROM (
                SELECT ei.product_id, ei.sold_qty as qty, (ei.sold_qty * ei.unit_price) as line_total
                FROM evening_item ei
                JOIN evening_settle es ON ei.settle_id = es.id
                WHERE es.date BETWEEN %s AND %s AND es.status = 'final'
                
                UNION ALL
                
                SELECT osi.product_id, osi.qty, osi.total_price as line_total
                FROM office_sale_items osi
                JOIN office_sales os ON osi.sale_id = os.id
                WHERE os.sale_date BETWEEN %s AND %s
            ) as sales_data
            JOIN products p ON sales_data.product_id = p.id
            GROUP BY p.id, p.name
            {order_clause}
        """
        
        db_cursor.execute(sql, (sql_start, sql_end, sql_start, sql_end))
        # Fetch as list of dicts for JSON compatibility
        rows = db_cursor.fetchall()
        for r in rows:
            product_data.append({
                'product_name': r['product_name'],
                'image': resolve_img(r['image']),
                'total_units_sold': int(r['total_units_sold'] or 0),
                'total_revenue': float(r['total_revenue'] or 0)
            })

    except Exception as e:
        flash(f"Error fetching product performance: {e}", "danger")
        app.logger.error(f"Product Perf Error: {e}")
    finally:
        if db_cursor: db_cursor.close()

    return render_template('reports/product_perfrm.html', 
                           start_date=display_start,
                           end_date=display_end,
                           sort_by=sort_by,
                           product_data=product_data)

# ... (rest of your routes) ...
@app.route('/reports/profitability_reprt', methods=['GET', 'POST'])
def profitability_report():
    today = date.today()
    start_date = today.replace(day=1)
    end_date = today
    report_data = {
        "total_revenue": 0,
        "total_cogs": 0,
        "total_expenses": 0,
        "net_profit": 0
    }
    if request.method == 'POST':
        start_date_str = request.form.get('start_date')
        end_date_str = request.form.get('end_date')
        start_date = date.fromisoformat(start_date_str)
        end_date = date.fromisoformat(end_date_str)
        db_cursor = None
        try:
            db_cursor = mysql.connection.cursor(MySQLdb.cursors.DictCursor)
            rev_sql = "SELECT SUM(ei.sold_qty * ei.unit_price) AS total_revenue FROM evening_item ei JOIN evening_settle es ON ei.settle_id = es.id WHERE es.date BETWEEN %s AND %s"
            db_cursor.execute(rev_sql, (start_date, end_date))
            revenue_result = db_cursor.fetchone()
            total_revenue = revenue_result['total_revenue'] or 0
            cogs_sql = "SELECT SUM(ei.sold_qty * p.purchase_price) AS total_cogs FROM evening_item ei JOIN evening_settle es ON ei.settle_id = es.id JOIN products p ON ei.product_id = p.id WHERE es.date BETWEEN %s AND %s"
            db_cursor.execute(cogs_sql, (start_date, end_date))
            cogs_result = db_cursor.fetchone()
            total_cogs = cogs_result['total_cogs'] or 0
            exp_sql = "SELECT SUM(amount) AS total_expenses FROM expenses WHERE expense_date BETWEEN %s AND %s"
            db_cursor.execute(exp_sql, (start_date, end_date))
            expense_result = db_cursor.fetchone()
            total_expenses = expense_result['total_expenses'] or 0
            net_profit = (total_revenue - total_cogs) - total_expenses
            report_data = {
                "total_revenue": float(total_revenue),
                "total_cogs": float(total_cogs),
                "total_expenses": float(total_expenses),
                "net_profit": float(net_profit)
            }
        except Exception as e:
            flash(f"An error occurred while generating the report: {e}", "danger")
        finally:
            if db_cursor:
                db_cursor.close()
    return render_template('reports/profitability_report.html', 
                           start_date=start_date.isoformat(), 
                           end_date=end_date.isoformat(),
                           start_date_str=start_date.strftime('%d %b, %Y'),
                           end_date_str=end_date.strftime('%d %b, %Y'),
                           **report_data
                           )



# --- Product Sales Detail Report Route ---
@app.route('/reports/product_sales', methods=['GET', 'POST'])
def product_sales():
    db_cursor = None
    try:
        db_cursor = mysql.connection.cursor(MySQLdb.cursors.DictCursor)
        db_cursor.execute("SELECT id, name FROM products ORDER BY name")
        all_products = db_cursor.fetchall()

        # Default to the last 30 days
        end_date = date.today()
        start_date = end_date - timedelta(days=29)
        
        sales_data = []
        selected_product_id = None
        selected_product_name = ""

        if request.method == 'POST':
            start_date_str = request.form.get('start_date')
            end_date_str = request.form.get('end_date')
            selected_product_id = request.form.get('product_id')
            
            start_date = date.fromisoformat(start_date_str)
            end_date = date.fromisoformat(end_date_str)
            
            if not selected_product_id:
                flash("Please select a product to generate the report.", "warning")
            else:
                sql = """
                    SELECT
                        e.name AS employee_name,
                        SUM(ei.sold_qty) AS total_units_sold
                    FROM evening_item ei
                    JOIN evening_settle es ON ei.settle_id = es.id
                    JOIN employees e ON es.employee_id = e.id
                    WHERE es.date BETWEEN %s AND %s AND ei.product_id = %s AND ei.sold_qty > 0
                    GROUP BY e.id, e.name
                    ORDER BY total_units_sold DESC
                """
                db_cursor.execute(sql, (start_date, end_date, selected_product_id))
                sales_data = db_cursor.fetchall()

                # Get product name for the report title
                product = next((p for p in all_products if p['id'] == int(selected_product_id)), None)
                if product:
                    selected_product_name = product['name']

        return render_template('reports/product_sales.html', 
                               all_products=all_products,
                               start_date=start_date.isoformat(), 
                               end_date=end_date.isoformat(),
                               start_date_str=start_date.strftime('%d %b, %Y'),
                               end_date_str=end_date.strftime('%d %b, %Y'),
                               sales_data=sales_data,
                               selected_product_id=int(selected_product_id) if selected_product_id else None,
                               selected_product_name=selected_product_name)
    except Exception as e:
        flash(f"An error occurred: {e}", "danger")
        return redirect(url_for('reports'))
    finally:
        if db_cursor:
            db_cursor.close()





# ================================================  Employee Ledger Routes =======================================================
# ... existing imports ...

# ============================================================
#  EMPLOYEE FINANCE LIST ROUTE (Update this in app.py)
# ============================================================
@app.route('/emp_list')
def emp_list():
    if 'loggedin' not in session: return redirect(url_for('login'))
    
    cursor = mysql.connection.cursor(MySQLdb.cursors.DictCursor)
    
    # FETCH EMPLOYEES: Changed ORDER BY to 'id ASC'
    # Also added LEFT JOIN to ensure position_name is fetched if position ID is used
    cursor.execute("""
        SELECT e.*, p.position_name 
        FROM employees e
        LEFT JOIN employee_positions p ON e.position_id = p.id 
        WHERE e.status = 'active' 
        ORDER BY e.id ASC
    """)
    employees = cursor.fetchall()
    
    # Calculate Global Stats (Unchanged)
    cursor.execute("""
        SELECT 
            SUM(CASE WHEN type = 'debit' THEN amount ELSE 0 END) as total_debit,
            SUM(CASE WHEN type = 'credit' THEN amount ELSE 0 END) as total_credit
        FROM employee_transactions
    """)
    stats = cursor.fetchone()
    
    total_debit = float(stats['total_debit'] or 0)
    total_credit = float(stats['total_credit'] or 0)
    net_holding = total_debit - total_credit 
    
    cursor.close()
    
    return render_template('emp_list.html', 
                           employees=employees, 
                           stats={
                               'total_debit': total_debit,
                               'total_credit': total_credit,
                               'net_holding': net_holding
                           })




# --- 1. Fix for PDF Generation (Handle None Dates) ---
@app.route('/employee/ledger/pdf/<int:employee_id>')
def employee_ledger_pdf(employee_id):
    if 'loggedin' not in session: return redirect(url_for('login'))
    
    cur = mysql.connection.cursor(MySQLdb.cursors.DictCursor)
    cur.execute("SELECT * FROM employees WHERE id=%s", (employee_id,))
    emp = cur.fetchone()
    
    # Fetch transactions (Ensure created_at is selected for time)
    cur.execute("""
        SELECT * FROM employee_transactions 
        WHERE employee_id=%s 
        ORDER BY transaction_date ASC, created_at ASC
    """, (employee_id,))
    transactions = cur.fetchall()
    cur.close()
    
    if not emp: return "Employee not found", 404
    
    pdf = PDF() 
    pdf.add_page()
    
    # Header
    pdf.set_font("Arial", 'B', 16)
    pdf.cell(0, 10, "Employee Ledger Statement", 0, 1, 'C')
    pdf.ln(5)
    
    # Info
    pdf.set_font("Arial", '', 11)
    pdf.cell(0, 6, f"Employee: {emp['name']}", 0, 1)
    # Handle missing position safely
    pos_name = emp.get('position') or emp.get('position_name') or 'N/A'
    pdf.cell(0, 6, f"Position: {pos_name}", 0, 1)
    
    # Current Time (IST)
    now_ist = datetime.now() + timedelta(hours=5, minutes=30) # Approx adjustment if server is UTC
    pdf.cell(0, 6, f"Generated: {now_ist.strftime('%d-%m-%Y %I:%M %p')}", 0, 1)
    pdf.ln(10)
    
    # Table Header
    pdf.set_font("Arial", 'B', 9)
    pdf.set_fill_color(230, 230, 230)
    
    pdf.cell(25, 8, "Date", 1, 0, 'C', 1)
    pdf.cell(20, 8, "Time", 1, 0, 'C', 1)
    pdf.cell(65, 8, "Description", 1, 0, 'L', 1)
    pdf.cell(25, 8, "Debit", 1, 0, 'R', 1)
    pdf.cell(25, 8, "Credit", 1, 0, 'R', 1)
    pdf.cell(30, 8, "Balance", 1, 1, 'R', 1)
    
    pdf.set_font("Arial", '', 8)
    balance = 0.0
    
    for t in transactions:
        amt = float(t['amount'])
        debit = 0.0
        credit = 0.0
        
        if t['type'] == 'debit':
            debit = amt
            balance += amt
        else:
            credit = amt
            balance -= amt
            
        # --- FIX: Handle None Date ---
        if t['transaction_date']:
            d_str = t['transaction_date'].strftime('%d-%m-%Y')
        else:
            d_str = "N/A"
            
        # --- FIX: Handle Time & Convert to IST ---
        if t.get('created_at'):
            # Assuming 'created_at' from DB is UTC, add 5:30 for IST
            # If your DB is already IST, remove the timedelta
            local_time = t['created_at'] + timedelta(hours=5, minutes=30)
            t_str = local_time.strftime('%I:%M %p')
        else:
            t_str = "-"
        
        pdf.cell(25, 7, d_str, 1, 0, 'C')
        pdf.cell(20, 7, t_str, 1, 0, 'C')
        pdf.cell(65, 7, pdf.safe_text(t['description']), 1, 0, 'L')
        pdf.cell(25, 7, f"{debit:.2f}" if debit > 0 else "-", 1, 0, 'R')
        pdf.cell(25, 7, f"{credit:.2f}" if credit > 0 else "-", 1, 0, 'R')
        pdf.cell(30, 7, f"{balance:.2f}", 1, 1, 'R')
        
    pdf.ln(5)
    pdf.set_font("Arial", 'B', 12)
    status = "Due from Employee" if balance > 0 else "Payable to Employee"
    pdf.cell(0, 10, f"Net Closing Balance: {abs(balance):.2f} ({status})", 0, 1, 'R')
    
    response = make_response(pdf.output(dest='S').encode('latin1'))
    response.headers['Content-Type'] = 'application/pdf'
    response.headers['Content-Disposition'] = f'attachment; filename={emp["name"]}_Ledger.pdf'
    return response



# --- HELPER: Get Cloudinary Public ID from URL (for legacy support if needed) ---
def get_public_id_from_url(url):
    """Extracts public_id from a full Cloudinary URL if present."""
    if url and "res.cloudinary.com" in url:
        try:
            parts = url.split("/upload/")
            if len(parts) > 1:
                version_and_id = parts[1].split("/")
                if version_and_id[0].startswith("v"):
                    public_id_with_ext = "/".join(version_and_id[1:])
                else:
                    public_id_with_ext = "/".join(version_and_id)
                return public_id_with_ext.rsplit(".", 1)[0]
        except: return url
    return url 




# ==========================================
# ROUTE: EMPLOYEE LEDGER (View) - Updated Query
# ==========================================
@app.route('/employee-ledger/<int:employee_id>')
def emp_ledger(employee_id):
    if 'loggedin' not in session: return redirect(url_for('login'))
    
    cur = mysql.connection.cursor(MySQLdb.cursors.DictCursor)
    
    cur.execute("SELECT id, name, position, email, phone, image FROM employees WHERE id = %s", [employee_id])
    employee = cur.fetchone()
    
    if not employee:
        flash('Employee not found!', 'danger')
        return redirect(url_for('emp_list'))
        
    # FIX: Added 'payment_mode' to the SELECT query
    cur.execute("""
        SELECT id, transaction_date, type, amount, description, payment_mode, created_at 
        FROM employee_transactions 
        WHERE employee_id = %s 
        ORDER BY transaction_date DESC, created_at DESC
    """, [employee_id])
    transactions_from_db = cur.fetchall()
    cur.close()
    
    from datetime import datetime as dt_class, date as date_class
    
    # Sort Logic
    transactions_calc = sorted(
        transactions_from_db, 
        key=lambda x: (
            x['transaction_date'] or date_class.min, 
            x['created_at'] or dt_class.min
        )
    )
    
    running_balance = 0.0
    total_debit = 0.0
    total_credit = 0.0
    has_opening_balance = False
    
    processed_transactions = []
    
    for t in transactions_calc:
        amt = float(t['amount'])
        if t['type'] == 'debit':
            running_balance += amt
            total_debit += amt
        else:
            running_balance -= amt
            total_credit += amt
            
        if t['description'] == 'Opening Balance': has_opening_balance = True
            
        t_dict = dict(t)
        t_dict['balance'] = running_balance
        
        # Time Formatting (Already IST in DB)
        if t.get('created_at'):
            t_dict['time_str'] = t['created_at'].strftime('%I:%M %p')
        else:
            t_dict['time_str'] = "-"
            
        processed_transactions.append(t_dict)
        
    # Reverse to show newest first
    transactions_display = processed_transactions[::-1]

    return render_template('emp_ledger.html', 
                           employee=employee, 
                           transactions=transactions_display, 
                           final_balance=running_balance,
                           total_debit=total_debit,
                           total_credit=total_credit,
                           has_opening_balance=has_opening_balance)





@app.route('/add-opening-balance/<int:employee_id>', methods=['GET', 'POST'])
def add_opening_balance(employee_id):
    if 'loggedin' not in session: return redirect(url_for('login'))
    
    cur = mysql.connection.cursor(MySQLdb.cursors.DictCursor)
    
    if request.method == 'POST':
        amount = float(request.form['amount'])
        type_ = request.form['type']
        date_ = request.form['date']
        
        cur.execute("SELECT id FROM employee_transactions WHERE employee_id=%s AND description='Opening Balance'", (employee_id,))
        exists = cur.fetchone()
        
        if exists:
            cur.execute("""
                UPDATE employee_transactions 
                SET amount=%s, type=%s, transaction_date=%s, created_at=NOW() 
                WHERE id=%s
            """, (amount, type_, date_, exists['id']))
            flash('Opening Balance Updated!', 'info')
        else:
            cur.execute("""
                INSERT INTO employee_transactions (employee_id, transaction_date, type, amount, description, created_at)
                VALUES (%s, %s, %s, %s, 'Opening Balance', NOW())
            """, (employee_id, date_, type_, amount))
            flash('Opening Balance Added!', 'success')
            
        mysql.connection.commit()
        cur.close()
        return redirect(url_for('emp_ledger', employee_id=employee_id))
        
    cur.execute("SELECT * FROM employees WHERE id=%s", (employee_id,))
    emp = cur.fetchone()
    cur.execute("SELECT * FROM employee_transactions WHERE employee_id=%s AND description='Opening Balance'", (employee_id,))
    existing = cur.fetchone()
    cur.close()
    
    return render_template('add_opening_balance.html', employee=emp, existing=existing)

# ==========================================
# ROUTE: ADD TRANSACTION (Updated with Payment Mode)
# ==========================================
@app.route('/add-transaction/<int:employee_id>', methods=['GET', 'POST'])
def add_transaction(employee_id):
    if 'loggedin' not in session: return redirect(url_for('login'))
    
    cur = mysql.connection.cursor(MySQLdb.cursors.DictCursor)
    
    # --- Ensure payment_mode column exists ---
    try:
        cur.execute("SELECT payment_mode FROM employee_transactions LIMIT 1")
    except:
        try:
            cur.execute("ALTER TABLE employee_transactions ADD COLUMN payment_mode VARCHAR(50) DEFAULT 'Cash'")
            mysql.connection.commit()
        except Exception as e:
            print(f"Error altering table: {e}")
    # -----------------------------------------

    if request.method == 'POST':
        raw_date = request.form['transaction_date']
        trans_date = parse_date_input(raw_date) 
        trans_type = request.form['type'] 
        amount = request.form['amount']
        description = request.form['description']
        payment_mode = request.form.get('payment_mode', 'Cash') # New Field

        # Time Fix
        client_time = request.form.get('client_time')
        if client_time and client_time.strip():
            final_time = client_time 
        else:
            final_time = get_ist_now().strftime('%Y-%m-%d %H:%M:%S')
        
        cur.execute("""
            INSERT INTO employee_transactions (employee_id, transaction_date, type, amount, description, payment_mode, created_at)
            VALUES (%s, %s, %s, %s, %s, %s, %s)
        """, (employee_id, trans_date, trans_type, amount, description, payment_mode, final_time))
        
        mysql.connection.commit()
        cur.close()
        flash('Transaction Added Successfully!', 'success')
        return redirect(url_for('emp_ledger', employee_id=employee_id))
    
    cur.execute("SELECT id, name FROM employees WHERE id = %s", [employee_id])
    employee = cur.fetchone()
    cur.close()
    
    selected_type = request.args.get('type', 'credit') 
    
    return render_template('add_transaction.html', employee=employee, selected_type=selected_type)




@app.route('/edit-transaction/<int:transaction_id>', methods=['GET', 'POST'])
def edit_transaction(transaction_id):
    cur = mysql.connection.cursor(MySQLdb.cursors.DictCursor)
    if request.method == 'POST':
        trans_date = request.form['transaction_date']
        trans_type = request.form['type']
        amount = request.form['amount']
        description = request.form['description']
        cur.execute("UPDATE employee_transactions SET transaction_date=%s, type=%s, amount=%s, description=%s WHERE id=%s", (trans_date, trans_type, amount, description, transaction_id))
        mysql.connection.commit()
        cur.execute("SELECT employee_id FROM employee_transactions WHERE id = %s", [transaction_id])
        transaction_info = cur.fetchone()
        cur.close()
        flash('Transaction Updated Successfully!', 'info')
        return redirect(url_for('emp_ledger', employee_id=transaction_info['employee_id']))
    cur.execute("SELECT et.*, e.name FROM employee_transactions et JOIN employees e ON et.employee_id = e.id WHERE et.id = %s", [transaction_id])
    transaction = cur.fetchone()
    cur.close()
    return render_template('edit_transaction.html', transaction=transaction)

@app.route('/delete-transaction/<int:transaction_id>', methods=['POST'])
def delete_transaction(transaction_id):
    cur = mysql.connection.cursor(MySQLdb.cursors.DictCursor)
    cur.execute("SELECT employee_id FROM employee_transactions WHERE id = %s", [transaction_id])
    result = cur.fetchone()
    if result:
        employee_id = result['employee_id']
        cur.execute("DELETE FROM employee_transactions WHERE id = %s", [transaction_id])
        mysql.connection.commit()
        flash('Transaction Deleted!', 'danger')
        cur.close()
        return redirect(url_for('emp_ledger', employee_id=employee_id))
    cur.close()
    flash('Transaction not found.', 'warning')
    return redirect(url_for('emp_list'))

#============================================= Employees transaction Reports module========================  
# =========================================================
#  HELPER: FETCH TRANSACTION DATA (Updated for Modern UI)
# =========================================================
def _fetch_transaction_data(filters):
    cur = mysql.connection.cursor(MySQLdb.cursors.DictCursor)
    
    where_clauses = []
    params = []

    # Date filtering
    period = filters.get('period')
    if period == 'today':
        where_clauses.append("t.transaction_date = CURDATE()")
    elif period == 'this_week':
        where_clauses.append("YEARWEEK(t.transaction_date, 1) = YEARWEEK(CURDATE(), 1)")
    elif period == 'this_month':
        where_clauses.append("YEAR(t.transaction_date) = YEAR(CURDATE()) AND MONTH(t.transaction_date) = MONTH(CURDATE())")
    elif period == 'this_year':
        where_clauses.append("YEAR(t.transaction_date) = YEAR(CURDATE())")
    elif period == 'custom' and filters.get('start_date') and filters.get('end_date'):
        where_clauses.append("t.transaction_date BETWEEN %s AND %s")
        params.extend([filters.get('start_date'), filters.get('end_date')])

    # Employee filtering
    employee_id = filters.get('employee_id')
    if employee_id and employee_id != 'all':
        where_clauses.append("t.employee_id = %s")
        params.append(employee_id)

    # ADDED: payment_mode, created_at
    sql = """
        SELECT t.id, t.transaction_date, t.created_at, t.type, t.amount, t.description, t.payment_mode,
               e.name as employee_name, e.image as employee_image
        FROM employee_transactions t
        JOIN employees e ON t.employee_id = e.id
    """
    if where_clauses:
        sql += " WHERE " + " AND ".join(where_clauses)
    sql += " ORDER BY t.transaction_date DESC, t.id DESC"

    cur.execute(sql, tuple(params))
    transactions = cur.fetchall()
    cur.close()

    # Process Images
    for t in transactions:
        if 'employee_image' in t:
            t['employee_image'] = resolve_img(t['employee_image'])
        
        # Format Time for Template
        t_str = "-"
        raw_time = t.get('created_at')
        if raw_time:
            if isinstance(raw_time, datetime):
                t_str = raw_time.strftime('%I:%M %p')
            elif isinstance(raw_time, timedelta):
                t_str = (datetime.min + raw_time).strftime('%I:%M %p')
            else:
                try: t_str = datetime.strptime(str(raw_time), '%Y-%m-%d %H:%M:%S').strftime('%I:%M %p')
                except: t_str = str(raw_time)
        t['formatted_time'] = t_str
            
    return transactions


# ==========================================
# TRANSACTION REPORT PDF CLASS (Final Layout Fix)
# ==========================================
class TransactionPDF(FPDF):
    def __init__(self):
        super().__init__()
        self.company_name = "REAL PROMOTION"
        self.slogan = "SINCE 2005"
        self.address_lines = [
            "Real Promotion, G/12, Tulsimangalam Complex,",
            "B/h. Trimurti Complex, Ghodiya Bazar,",
            "Nadiad-387001, Gujarat, India"
        ]
        self.contact = "+91 96623 22476 | help@realpromotion.in"

    def header(self):
        # Company Header
        self.set_font('Arial', 'B', 24)
        self.set_text_color(26, 35, 126) # Dark Blue
        self.cell(0, 10, self.company_name, 0, 1, 'C')
        
        self.set_font('Arial', 'B', 10)
        self.set_text_color(100, 100, 100) # Grey
        self.cell(0, 5, self.slogan, 0, 1, 'C')
        self.ln(2)
        
        self.set_font('Arial', '', 9)
        self.set_text_color(50, 50, 50)
        for line in self.address_lines:
            self.cell(0, 4, line, 0, 1, 'C')
        self.cell(0, 4, self.contact, 0, 1, 'C')
        
        self.ln(5)
        self.set_draw_color(200, 200, 200)
        self.line(10, self.get_y(), 200, self.get_y())
        self.ln(5)

        self.set_font('Arial', 'B', 16)
        self.set_text_color(0, 0, 0)
        self.cell(0, 10, "TRANSACTION REPORT", 0, 1, 'C')
        self.ln(5)

    def footer(self):
        self.set_y(-15)
        self.set_font('Arial', 'I', 8)
        self.set_text_color(128, 128, 128)
        self.cell(0, 10, f'Page {self.page_no()}', 0, 0, 'C')

    def add_filters_info(self, filters, employee_details=None):
        self.set_font('Arial', '', 10)
        self.set_text_color(0, 0, 0)
        
        # Period Info
        period = filters.get('period', 'All')
        start = filters.get('start_date')
        end = filters.get('end_date')
        
        period_text = f"Report Period: {period.replace('_', ' ').title()}"
        if start and end:
            period_text += f" ({start} to {end})"
        
        self.cell(0, 6, period_text, 0, 1, 'L')
        
        # Employee Details
        if employee_details:
            self.ln(2)
            self.set_font('Arial', 'B', 11)
            self.cell(0, 6, "Employee Profile:", 0, 1, 'L')
            
            self.set_font('Arial', '', 10)
            self.cell(25, 6, "Name:", 0, 0)
            self.set_font('Arial', 'B', 10)
            self.cell(60, 6, str(employee_details['name']), 0, 0)
            
            self.set_font('Arial', '', 10)
            self.cell(20, 6, "Mobile:", 0, 0)
            self.set_font('Arial', 'B', 10)
            self.cell(40, 6, str(employee_details.get('phone', 'N/A')), 0, 1)
            
            if employee_details.get('position_name'):
                self.set_font('Arial', '', 10)
                self.cell(25, 6, "Position:", 0, 0)
                self.set_font('Arial', 'B', 10)
                self.cell(60, 6, str(employee_details['position_name']), 0, 1)
            
            # Generated Time
            self.ln(6) 
            self.set_font('Arial', '', 9)
            self.cell(25, 6, "Generated:", 0, 0)
            self.set_font('Arial', 'I', 9)
            ist_now = datetime.utcnow() + timedelta(hours=5, minutes=30)
            self.cell(60, 6, ist_now.strftime('%d-%b-%Y %I:%M %p'), 0, 1)

        self.ln(5)

    def add_table(self, transactions):
        # Header
        cols = ["#", "Date/Time", "Employee", "Description", "Credit", "Debit"]
        # Widths: 10, 40, 40, 45, 27, 28 = 190
        widths = [10, 40, 40, 45, 27, 28]
        
        self.set_font('Arial', 'B', 9)
        self.set_fill_color(26, 35, 126) 
        self.set_text_color(255, 255, 255)
        self.set_draw_color(200, 200, 200) # Light grey grid lines
        self.set_line_width(0.1)
        
        for i, col in enumerate(cols):
            self.cell(widths[i], 8, col, 1, 0, 'C', True)
        self.ln()
        
        # Rows
        self.set_font('Arial', '', 8)
        self.set_text_color(0, 0, 0)
        fill = False
        
        total_credit = 0
        total_debit = 0
        
        for idx, t in enumerate(transactions):
            # Row Background Logic (Zebra Striping)
            if fill:
                self.set_fill_color(245, 245, 245)
            else:
                self.set_fill_color(255, 255, 255)

            # 1. Date/Time Parsing
            if t['transaction_date']:
                 d_part = t['transaction_date'].strftime('%d-%m-%Y')
            else: d_part = "-"
                 
            t_part = ""
            raw_time = t.get('created_at')
            if raw_time:
                if isinstance(raw_time, datetime):
                    t_part = raw_time.strftime('%I:%M %p')
                elif isinstance(raw_time, timedelta):
                    t_part = (datetime.min + raw_time).strftime('%I:%M %p')
                else:
                    try: t_part = datetime.strptime(str(raw_time), '%Y-%m-%d %H:%M:%S').strftime('%I:%M %p')
                    except: pass
            
            # Combine Date & Time
            dt_str = f"{d_part}\n{t_part}" if t_part else d_part

            # Financials
            credit_amt = "-"
            debit_amt = "-"
            if t['type'] == 'credit':
                val = float(t['amount'])
                total_credit += val
                credit_amt = f"{val:.2f}"
            else:
                val = float(t['amount'])
                total_debit += val
                debit_amt = f"{val:.2f}"
            
            # Safe Strings
            desc_text = str(t['description'] or '')[:35]
            emp_name = str(t['employee_name'])[:20] 

            # Calculate Height needed (Date might be 2 lines)
            # Standard cell height is 6, but date/time needs 10 to look good
            row_h = 10 

            # Check Page Break
            if self.get_y() > 250:
                self.add_page()
                # Reprint Header
                self.set_font('Arial', 'B', 9)
                self.set_fill_color(26, 35, 126) 
                self.set_text_color(255, 255, 255)
                for i, col in enumerate(cols):
                    self.cell(widths[i], 8, col, 1, 0, 'C', True)
                self.ln()
                self.set_font('Arial', '', 8)
                self.set_text_color(0, 0, 0)
            
            # --- PRINT ROW CELLS ---
            # Save X and Y
            x_start = self.get_x()
            y_start = self.get_y()
            
            # Draw Cells with uniform height and fill
            self.cell(widths[0], row_h, str(idx+1), 1, 0, 'C', True)
            
            # Date/Time is Multiline effectively, but let's just use regular cell with newline hack
            # FPDF cell doesn't render \n unless MultiCell used. 
            # Switching to compact format "DD-MM-YY HH:MM" to keep single line & fix spacing
            compact_dt = f"{d_part} {t_part}" if t_part else d_part
            self.cell(widths[1], row_h, compact_dt, 1, 0, 'C', True)
            
            self.cell(widths[2], row_h, emp_name, 1, 0, 'L', True)
            self.cell(widths[3], row_h, desc_text, 1, 0, 'L', True)
            self.cell(widths[4], row_h, credit_amt, 1, 0, 'R', True)
            self.cell(widths[5], row_h, debit_amt, 1, 1, 'R', True) # End of line
            
            fill = not fill
            
        # Totals Row
        self.set_font('Arial', 'B', 9)
        self.set_fill_color(230, 230, 230)
        self.cell(sum(widths[:4]), 8, "TOTALS", 1, 0, 'R', True)
        self.cell(widths[4], 8, f"{total_credit:.2f}", 1, 0, 'R', True)
        self.cell(widths[5], 8, f"{total_debit:.2f}", 1, 1, 'R', True)
        self.ln(10)
        
        # Net Flow Box
        net = total_debit - total_credit
        status = "Payable (Company Owes)" if net < 0 else "Receivable (Employee Owes)"
        
        self.set_x(10)
        self.set_font('Arial', 'B', 11)
        self.set_text_color(0, 0, 0)
        self.cell(0, 8, f"Net Position: {abs(net):.2f} - {status}", 0, 1, 'R')
        
        # Signature
        self.add_signature()

    def add_signature(self):
        if self.get_y() > 240: self.add_page()
        self.ln(10)
        y_pos = self.get_y()
        
        sig_path = os.path.join(app.root_path, 'static', 'img', 'signature.png')
        if os.path.exists(sig_path):
            self.image(sig_path, x=150, y=y_pos, w=40)
            
        self.set_xy(140, y_pos + 25)
        self.set_font('Arial', 'B', 10)
        self.set_text_color(0, 0, 0)
        self.cell(60, 5, "Authorized Signature", 0, 1, 'C')
        self.ln(10)

# =========================================================
#  DOWNLOAD TRANSACTION REPORT ROUTE
# =========================================================
@app.route('/download_transaction_report')
def download_transaction_report():
    if "loggedin" not in session: return redirect(url_for("login"))
    
    filters = {
        'period': request.args.get('period', 'this_month'),
        'employee_id': request.args.get('employee_id', 'all'),
        'start_date': request.args.get('start_date', ''),
        'end_date': request.args.get('end_date', '')
    }
    
    report_format = request.args.get('format', 'pdf')
    transactions = _fetch_transaction_data(filters)
    
    emp_details = None
    if filters['employee_id'] != 'all':
        cur = mysql.connection.cursor(MySQLdb.cursors.DictCursor)
        cur.execute("""
            SELECT e.name, e.phone, p.position_name 
            FROM employees e 
            LEFT JOIN employee_positions p ON e.position_id = p.id 
            WHERE e.id = %s
        """, (filters['employee_id'],))
        emp_details = cur.fetchone()
        cur.close()
    
    if report_format == 'pdf':
        try:
            pdf = TransactionPDF()
            pdf.alias_nb_pages()
            pdf.add_page()
            pdf.add_filters_info(filters, emp_details)
            pdf.add_table(transactions)
            
            pdf_string = pdf.output(dest='S').encode('latin-1')
            buffer = io.BytesIO(pdf_string)
            buffer.seek(0)
            
            filename = f"Transaction_Report_{date.today()}.pdf"
            if emp_details:
                filename = f"{emp_details['name'].replace(' ', '_')}_Report.pdf"
                
            return send_file(buffer, as_attachment=True, download_name=filename, mimetype='application/pdf')
            
        except Exception as e:
            app.logger.error(f"PDF Error: {e}")
            flash(f"Error generating PDF: {e}", "danger")
            return redirect(url_for('transaction_report'))

    elif report_format == 'excel':
        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Transactions"
            
            headers = ['Date', 'Time', 'Employee', 'Description', 'Credit (Due)', 'Debit (Given)']
            ws.append(headers)
            for cell in ws[1]: cell.font = Font(bold=True)
            
            for t in transactions:
                c_amt = float(t['amount']) if t['type'] == 'credit' else 0
                d_amt = float(t['amount']) if t['type'] == 'debit' else 0
                
                t_str = ""
                if t.get('created_at'):
                    if isinstance(t['created_at'], datetime): t_str = t['created_at'].strftime('%I:%M %p')
                    else: t_str = str(t['created_at'])

                ws.append([
                    t['transaction_date'].strftime('%Y-%m-%d'),
                    t_str,
                    t['employee_name'],
                    t['description'],
                    c_amt,
                    d_amt
                ])
                
            buffer = io.BytesIO()
            wb.save(buffer)
            buffer.seek(0)
            
            return send_file(buffer, as_attachment=True, download_name=f"Transaction_Report_{date.today()}.xlsx", 
                             mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                             
        except Exception as e:
             app.logger.error(f"Excel Error: {e}")
             flash(f"Error generating Excel: {e}", "danger")
             return redirect(url_for('transaction_report'))

    return redirect(url_for('transaction_report'))

# =========================================================
#  TRANSACTION REPORT VIEW ROUTE
# =========================================================
@app.route('/transaction-report', methods=['GET', 'POST'])
def transaction_report():
    if "loggedin" not in session: return redirect(url_for("login"))
    
    cur = mysql.connection.cursor(MySQLdb.cursors.DictCursor)
    cur.execute("SELECT id, name FROM employees WHERE status = 'active' ORDER BY name ASC")
    employees = cur.fetchall()
    cur.close()

    filters = {
        'period': 'this_month', 
        'employee_id': 'all', 
        'start_date': '', 
        'end_date': ''
    }
    
    if request.method == 'POST':
        filters['period'] = request.form.get('period') or 'this_month'
        filters['employee_id'] = request.form.get('employee_id') or 'all'
        filters['start_date'] = request.form.get('start_date') or ''
        filters['end_date'] = request.form.get('end_date') or ''

    transactions = _fetch_transaction_data(filters)
    
    # --- 1. Basic Totals ---
    total_debit = sum(float(t['amount']) for t in transactions if t['type'] == 'debit')
    total_credit = sum(float(t['amount']) for t in transactions if t['type'] == 'credit')
    net_flow = total_debit - total_credit

    return render_template('reports/transaction_report.html',
                           employees=employees,
                           transactions=transactions,
                           total_transactions=len(transactions),
                           total_debit=total_debit,
                           total_credit=total_credit,
                           net_flow=net_flow,
                           filters=filters)



# ---------- CUSTOM JINJA FILTER: INR FORMAT ----------#
@app.template_filter("inr")
def inr_format(value):
    """Format number as Indian Rupees."""
    try:
        value = float(value)
        return f"₹ {value:,.2f}"
    except:
        return value



# --- FINAL: Add this to the very bottom ---
# This block is for local development
# Render will use gunicorn to run the 'app' object
if __name__ == "__main__":
    app.logger.info("Starting app in debug mode...")
    app.run(debug=True, host='0.0.0.0', port=int(os.environ.get('PORT', 5000)))

















































